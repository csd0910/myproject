import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

def run_comparison():
    # 1. 中間ファイル（temp_stage1.xlsx）の読み込み。なければシミュレートして作成。
    temp_file = r"C:\Users\フォーレスト026\MyProject\UploadDataCreate\output\temp_stage1.xlsx"
    base_file = r"C:\Users\フォーレスト026\MyProject\UploadDataCreate\【テスト用】商品名作成 尾島手作業\元データ\result_260819_154419.xlsx"
    exclude_file = r"C:\Users\フォーレスト026\MyProject\UploadDataCreate\【テスト用】商品名作成 尾島手作業\元データ\【テスト用】キーワード追加用情報抽出_111_260819.xlsx"
    ojima_file = r"C:\Users\フォーレスト026\MyProject\UploadDataCreate\【テスト用】商品名作成 尾島手作業\【手順1】result_260819_154419.xlsx"
    
    if not os.path.exists(temp_file):
        print("中間ファイルが存在しないため、Pythonプログラムと同じ処理をシミュレートして生成します...")
        df_base = pd.read_excel(base_file)
        df_ex = pd.read_excel(exclude_file, sheet_name=0)
        
        # 品目cd(1), 注文番号(3)
        ex_keys = (df_ex.iloc[:, 1].fillna('').astype(str).str.strip() + "_" + df_ex.iloc[:, 3].fillna('').astype(str).str.strip()).tolist()
        base_keys = (df_base.iloc[:, 2].fillna('').astype(str).str.strip() + "_" + df_base.iloc[:, 3].fillna('').astype(str).str.strip())
        df_temp = df_base[~base_keys.isin(ex_keys)].copy()
    else:
        df_temp = pd.read_excel(temp_file)
        
    df_ojima = pd.read_excel(ojima_file)
    
    # 注文番号をキーにして比較
    ojima_orders = set(df_ojima.iloc[:, 3].astype(str).str.strip())
    temp_orders = set(df_temp.iloc[:, 3].astype(str).str.strip())
    
    # 全注文番号の和集合
    all_orders = list(ojima_orders.union(temp_orders))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Step1_比較結果"
    
    # ヘッダー作成
    headers = ["判定", "注文番号", "尾島様の手作業 (有無)", "Pythonプログラム (有無)"]
    ws.append(headers)
    
    fill_blue = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid") # 薄い青（緑がかった青）
    fill_red = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid") # 薄い赤
    
    row_idx = 2
    match_count = 0
    diff_count = 0
    
    for order in all_orders:
        in_ojima = "〇 (残っている)" if order in ojima_orders else "× (除外された)"
        in_temp = "〇 (残っている)" if order in temp_orders else "× (除外された)"
        
        if in_ojima == in_temp:
            judge = "一致"
            fill = fill_blue
            match_count += 1
        else:
            judge = "不一致"
            fill = fill_red
            diff_count += 1
            
        ws.cell(row=row_idx, column=1, value=judge).fill = fill
        ws.cell(row=row_idx, column=2, value=order).fill = fill
        ws.cell(row=row_idx, column=3, value=in_ojima).fill = fill
        ws.cell(row=row_idx, column=4, value=in_temp).fill = fill
        row_idx += 1
        
    out_path = r"C:\Users\フォーレスト026\MyProject\UploadDataCreate\output\Step1_比較結果.xlsx"
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"比較完了！ 一致: {match_count}件, 不一致: {diff_count}件")
    print(f"出力先: {out_path}")
    
if __name__ == "__main__":
    run_comparison()
