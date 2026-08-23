import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import os

FILL_GREEN = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
FILL_RED = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FONT_RED = Font(color="FF0000", bold=True)

def get_col_idx(df, col_name, fallback_idx):
    if col_name in df.columns:
        return df.columns.get_loc(col_name)
    for i, c in enumerate(df.columns):
        if col_name in str(c):
            return i
    return fallback_idx

def generate_diff_excel(prev_temp_path, temp_file_path, ojima_file_path, output_diff_path):
    if not os.path.exists(temp_file_path) or not os.path.exists(ojima_file_path):
        return False, "比較元のファイルが存在しません。"
        
    df_prev = pd.read_excel(prev_temp_path) if prev_temp_path and os.path.exists(prev_temp_path) else pd.DataFrame()
    df_temp = pd.read_excel(temp_file_path)
    df_ojima = pd.read_excel(ojima_file_path)
    
    # 尾島様データの辞書作成 (注文番号をキーとする)
    ojima_dict = {}
    if not df_ojima.empty:
        o_chumon_idx = get_col_idx(df_ojima, "注文番号", 3)
        for _, row in df_ojima.iterrows():
            k = str(row.iloc[o_chumon_idx]).strip()
            if k and k != 'nan':
                ojima_dict[k] = row
                
    wb = Workbook()
    ws = wb.active
    
    # ヘッダはそのまま
    ws.append(list(df_temp.columns))
    
    t_chumon_idx = get_col_idx(df_temp, "注文番号", 3)
    
    # Python出力のデータ行
    for _, row in df_temp.iterrows():
        order_num = str(row.iloc[t_chumon_idx]).strip()
        ws.append(list(row))
        current_row = ws.max_row
        
        if order_num not in ojima_dict:
            # 尾島ファイルに無い（Pythonの消し忘れ等）-> 行全体を赤背景
            for cell in ws[current_row]:
                cell.fill = FILL_RED
        else:
            # 尾島ファイルにある -> 基本は緑背景
            ojima_row = ojima_dict[order_num]
            for col_idx, cell in enumerate(ws[current_row], start=1):
                cell.fill = FILL_GREEN
                
                # 値が変更されているかチェック (列名で比較)
                col_name = df_temp.columns[col_idx - 1]
                if col_name in df_ojima.columns:
                    val_temp = str(cell.value).strip() if pd.notnull(cell.value) else ""
                    ojima_val_raw = ojima_row[col_name]
                    val_ojima = str(ojima_val_raw).strip() if pd.notnull(ojima_val_raw) else ""
                    
                    # 浮動小数点の微妙な違い（例: '1.0' vs '1'）を吸収するため、数値として比較できる場合はする
                    if val_temp != val_ojima:
                        try:
                            if float(val_temp) == float(val_ojima):
                                continue
                        except ValueError:
                            pass
                            
                        # 変更されたセルはフォントを赤にする
                        cell.font = FONT_RED

    # 削除された分（黄色い空セル）
    if not df_prev.empty:
        prev_orders = set(df_prev.iloc[:, get_col_idx(df_prev, "注文番号", 3)].dropna().astype(str).str.strip())
        temp_orders = set(df_temp.iloc[:, t_chumon_idx].dropna().astype(str).str.strip())
        deleted_orders = prev_orders - temp_orders
        
        # 削除された件数分、下に黄色い空行を追加
        for _ in range(len(deleted_orders)):
            ws.append([""] * len(df_temp.columns))
            for cell in ws[ws.max_row]:
                cell.fill = FILL_YELLOW

    wb.save(output_diff_path)
    return True, "色付け完了（追加カラム・別シートなし）"
