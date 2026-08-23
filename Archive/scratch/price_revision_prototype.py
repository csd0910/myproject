import os
import openpyxl
from openpyxl.utils import get_column_letter

# パスの設定
TARGET_DIR = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\価格改定についてのツール化検証\参考資料"
TARGET_FILE = "●20260710【114】7月末_アイリスオーヤマ価格改定_173件.xlsx"
OUTPUT_FILE = "●20260710【114】7月末_アイリスオーヤマ価格改定_173件_自動計算推奨版.xlsx"

def run_prototype():
    input_path = os.path.join(TARGET_DIR, TARGET_FILE)
    output_path = os.path.join(TARGET_DIR, OUTPUT_FILE)
    
    print(f"Loading workbook: {input_path}")
    wb = openpyxl.load_workbook(input_path, data_only=False) # 数式を壊さないために data_only=False
    ws = wb.active
    
    # 動的列検出のためのヘッダー行（通常は3行目）の解析
    header_row = 3
    col_mapping = {}
    
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row, column=col).value
        if cell_val:
            cell_val_clean = str(cell_val).replace("\n", "").replace(" ", "").strip()
            col_mapping[cell_val_clean] = col
            
    print("Detected headers:")
    for k, v in col_mapping.items():
        if k in ["売価", "仕切", "粗利率", "売価1", "AS換算売価", "換算金額", "アスクル", "たの", "カウ"]:
            print(f"  {k}: Column {get_column_letter(v)} ({v})")
            
    # 主要列のインデックス特定（検出できない場合は固定値や近似値で代替）
    # ※ メーカーやファイルによる表記ブレに対応
    def find_col(possible_names, default_idx):
        for name in possible_names:
            if name in col_mapping:
                return col_mapping[name]
        return default_idx

    # 列番号の決定
    col_old_price = find_col(["売価"], 15)  # 現行売価 (O)
    col_old_shikiri = find_col(["仕切"], 16)  # 現行仕切 (P)
    col_new_shikiri_calc = find_col(["仕切換算"], 24) # 改定後仕切換算 (X)
    col_competitor_as = find_col(["AS換算売価", "換算金額"], 29) # アスクル換算売価 (AC)
    col_new_price1 = find_col(["売価1"], 32) # 改定後売価1 (AF)
    
    print(f"\nTarget columns configured:")
    print(f"  現行売価: {get_column_letter(col_old_price)}")
    print(f"  現行仕切: {get_column_letter(col_old_shikiri)}")
    print(f"  新仕切換算: {get_column_letter(col_new_shikiri_calc)}")
    print(f"  アスクル換算: {get_column_letter(col_competitor_as)}")
    print(f"  新売価1 (書き込み先): {get_column_letter(col_new_price1)}")

    # 最低粗利率の設定（例: 15%）
    MIN_MARGIN = 0.15
    
    # 4行目から最終行までデータ処理
    modified_count = 0
    
    # データ読み込み用に別インスタンス(data_only=True)を開く
    wb_data = openpyxl.load_workbook(input_path, data_only=True)
    ws_data = wb_data.active
    
    max_row = ws.max_row
    for r in range(4, max_row + 1):
        # 現在の数値を読み込み
        old_price = ws_data.cell(row=r, column=col_old_price).value
        old_shikiri = ws_data.cell(row=r, column=col_old_shikiri).value
        new_shikiri = ws_data.cell(row=r, column=col_new_shikiri_calc).value
        competitor_price = ws_data.cell(row=r, column=col_competitor_as).value
        
        # 必要な数値（旧売価、旧仕切、新仕切）がない行はスキップ
        if old_price is None or old_shikiri is None or new_shikiri is None:
            continue
            
        try:
            old_price = float(old_price)
            old_shikiri = float(old_shikiri)
            new_shikiri = float(new_shikiri)
        except ValueError:
            continue
            
        # 1. 競合価格がある場合の仮値（競合価格 - 1円）
        # 2. 競合価格がない場合は、原価値上率スライド値上げ
        if competitor_price is not None and str(competitor_price).strip() not in ["-", ""]:
            try:
                comp_val = float(competitor_price)
                suggested_price = comp_val - 1
            except ValueError:
                # 競合価格が数値に変換できない場合はスライド
                rate = new_shikiri / old_shikiri if old_shikiri > 0 else 1.0
                suggested_price = round(old_price * rate)
        else:
            rate = new_shikiri / old_shikiri if old_shikiri > 0 else 1.0
            suggested_price = round(old_price * rate)
            
        # 粗利率下限チェック (15%)
        # 粗利率 = (売価 - 新仕切) / 売価
        if suggested_price > 0:
            margin = (suggested_price - new_shikiri) / suggested_price
            if margin < MIN_MARGIN:
                # 粗利率が15%を下回る場合は、粗利率15%を確保する価格に引き上げ
                suggested_price = round(new_shikiri / (1.0 - MIN_MARGIN))
                
        # 整数に丸める
        suggested_price = int(suggested_price)
        
        # 元の数式ブック(wb)の「売価1」に値を書き込む
        ws.cell(row=r, column=col_new_price1, value=suggested_price)
        modified_count += 1
        
        if r <= 10:
            print(f"Row {r}: 旧売価={old_price}, 旧仕切={old_shikiri} -> 新仕切={new_shikiri} | 競合={competitor_price} | 推奨売価={suggested_price}")

    wb_data.close()
    
    # 変更を保存
    wb.save(output_path)
    wb.close()
    print(f"\nProcessing completed. Modified {modified_count} rows.")
    print(f"Saved output to: {output_path}")

if __name__ == "__main__":
    run_prototype()
