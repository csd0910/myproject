import os
import openpyxl
from openpyxl.utils import get_column_letter

TARGET_DIR = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\価格改定についてのツール化検証\参考資料"

files = [
    "●20260710【114】7月末_アイリスオーヤマ価格改定_173件.xlsx",
    "2026.08.01改定_ﾌｼﾞｺﾋﾟｱﾝ_契約条件変更フォーマット 1件.xlsx",
    "20260713【797】7月末_カグクロ価格改定_30件.xlsx",
    "●20260701【291】7月末_今村紙工価格改定_47件.xlsx"
]

def inspect_file(filepath):
    print(f"\n========================================\nFile: {os.path.basename(filepath)}")
    if not os.path.exists(filepath):
        print("File does not exist.")
        return
        
    try:
        wb_val = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        wb_form = openpyxl.load_workbook(filepath, data_only=False, read_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return
    
    for sheetname in wb_val.sheetnames:
        print(f"\n--- Sheet: {sheetname} ---")
        ws_val = wb_val[sheetname]
        ws_form = wb_form[sheetname]
        
        # Get dimensions
        max_r = ws_val.max_row
        max_c = ws_val.max_column
        print(f"Dimensions: {max_r} rows x {max_c} cols")
        
        # Read header (first 5 rows)
        rows_val = list(ws_val.iter_rows(max_row=min(5, max_r or 5), max_col=max_c, values_only=True))
        rows_form = list(ws_form.iter_rows(max_row=min(5, max_r or 5), max_col=max_c, values_only=True))
        
        # Display header with column index/letter
        for r_idx in range(len(rows_val)):
            print(f"Row {r_idx + 1}:")
            for c_idx in range(len(rows_val[r_idx])):
                col_letter = get_column_letter(c_idx + 1)
                val = rows_val[r_idx][c_idx]
                form = rows_form[r_idx][c_idx]
                
                # Check if it has a formula or non-null value
                if val is not None or form is not None:
                    # If it's a formula, print it
                    if isinstance(form, str) and form.startswith('='):
                        print(f"  {col_letter}{r_idx+1}: [Formula] {form} (Value: {val})")
                    else:
                        print(f"  {col_letter}{r_idx+1}: {val}")
                        
        # Let's inspect some formulas in later rows if possible (e.g. up to row 30)
        # Search for columns: M-Q (13-17), AC-AF (29-32), BT-BX (72-76), Q-AB (17-28)
        interest_cols = list(range(13, 18)) + list(range(29, 33)) + list(range(72, 77)) + list(range(17, 29))
        interest_cols = sorted(list(set(interest_cols)))
        
        print("\nChecking formulas in row 6-30 for interest columns:")
        limit_r = min(30, max_r or 30)
        has_formulas = False
        # Create non-read-only check for formula scan to avoid max_row None issue if any
        for r in range(6, limit_r + 1):
            for c in interest_cols:
                if c <= max_c:
                    col_letter = get_column_letter(c)
                    val = ws_val.cell(row=r, column=c).value
                    form = ws_form.cell(row=r, column=c).value
                    if isinstance(form, str) and form.startswith('='):
                        print(f"  {col_letter}{r}: {form} -> {val}")
                        has_formulas = True
        if not has_formulas:
            print("  No formulas found in interest columns for rows 6-30.")
            
    wb_val.close()
    wb_form.close()

for f in files:
    inspect_file(os.path.join(TARGET_DIR, f))
