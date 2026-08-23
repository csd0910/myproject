import os
import openpyxl
from collections import Counter

TARGET_DIR = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\価格改定についてのツール化検証\参考資料"

files = [
    "●20260710【114】7月末_アイリスオーヤマ価格改定_173件.xlsx",
    "20260713【797】7月末_カグクロ価格改定_30件.xlsx",
    "●20260617【887】7月末_ジェムコ価格改定_105件.xlsx",
    "●20260626【982】7月末_ジョインテックス価格改定_64件.xlsx",
    "●20260701【291】7月末_今村紙工価格改定_47件.xlsx",
    "●20260702【740】7月末_共和価格改定_27件_売価2あり.xlsx",
    "●20260708【9323】7月末_福井価格改定_22件.xlsx"
]

def analyze_patterns():
    print("=== 各ファイルごとの端数パターン（下2桁・下3桁）の分析 ===")
    
    for filename in files:
        filepath = os.path.join(TARGET_DIR, filename)
        if not os.path.exists(filepath):
            continue
            
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            ws = wb.active
        except Exception:
            continue
            
        # ヘッダー検出
        header_row = 3
        col_mapping = {}
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col).value
            if cell_val:
                col_mapping[str(cell_val).replace("\n", "").strip()] = col
                
        col_new_price = col_mapping.get("売価1", col_mapping.get("新売価", col_mapping.get("新）売価", 32)))
        
        prices = []
        for r in range(4, ws.max_row + 1):
            val = ws.cell(row=r, column=col_new_price).value
            if val is not None:
                try:
                    prices.append(int(float(val)))
                except ValueError:
                    pass
                    
        wb.close()
        
        if not prices:
            continue
            
        # 下1桁、下2桁、下3桁の集計
        last_1 = [p % 10 for p in prices]
        last_2 = [p % 100 for p in prices]
        last_3 = [p % 1000 for p in prices]
        
        c_1 = Counter(last_1).most_common(3)
        c_2 = Counter(last_2).most_common(3)
        c_3 = Counter(last_3).most_common(3)
        
        print(f"\n■ ファイル名: {filename} (データ数: {len(prices)}件)")
        print(f"  価格帯: {min(prices)}円 〜 {max(prices)}円")
        print(f"  代表的な下1桁: " + ", ".join(f"末尾{k}円 ({v}件, {v/len(prices)*100:.1f}%)" for k, v in c_1))
        print(f"  代表的な下2桁: " + ", ".join(f"末尾{k:02d}円 ({v}件, {v/len(prices)*100:.1f}%)" for k, v in c_2))
        if max(prices) >= 1000:
            print(f"  代表的な下3桁: " + ", ".join(f"末尾{k:03d}円 ({v}件, {v/len(prices)*100:.1f}%)" for k, v in c_3))
        print("-" * 60)

if __name__ == "__main__":
    analyze_patterns()
