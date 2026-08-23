import os
import openpyxl
from collections import Counter
import sys

BASE_DIR = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\価格改定についてのツール化検証\参考資料"
SUB_DIRS = [
    "2026年5月1日改定",
    "2026年6月1日改定",
    "2026年7月1日改定",
    "2026年8月1日改定",
    "2026年9月1日改定"
]

def round_to_nearest_10(val):
    return round(val / 10) * 10

def round_to_nearest_100(val):
    return round(val / 100) * 100

def scan_and_analyze():
    all_files = []
    
    # サブフォルダ内のExcelファイルを再帰的に探索
    for sub in SUB_DIRS:
        sub_path = os.path.join(BASE_DIR, sub)
        if not os.path.exists(sub_path):
            continue
            
        for root, dirs, files_in_dir in os.walk(sub_path):
            for f in files_in_dir:
                if f.endswith(".xlsx") and not f.startswith("~$"):
                    full_path = os.path.join(root, f)
                    size_mb = os.path.getsize(full_path) / (1024 * 1024)
                    rel_path = os.path.relpath(full_path, BASE_DIR)
                    all_files.append((rel_path, full_path, size_mb))
                
    print(f"Found {len(all_files)} Excel files in subdirectories.")
    sys.stdout.flush()
    
    # 総合集計
    global_patterns = {
        "現行据え置き (価格差0)": 0,
        "競合価格対抗 (アスクル等 -1円〜-0円など)": 0,
        "アスクルスライド値 (切り捨て・端数処理含む)": 0,
        "自社原価スライド値 (スライド売価)": 0,
        "自社原価スライド値 (丸め調整)": 0,
        "粗利率基準 (15%, 20%, 25%, 30%張り付き)": 0,
        "その他個別調整": 0
    }
    
    total_valid_rows = 0
    all_prices = []
    
    out_file = r"c:\Users\フォーレスト026\MyProject\scratch\subfolders_analysis_report.txt"
    with open(out_file, "w", encoding="utf-8") as out:
        out.write("==================================================\n")
        out.write("改定時期別サブフォルダ内Excelデータ網羅的解析\n")
        out.write("==================================================\n\n")
        out.flush()
        
        file_idx = 0
        for rel_path, fpath, size_mb in all_files:
            file_idx += 1
            print(f"[{file_idx}/{len(all_files)}] Processing: {rel_path} ({size_mb:.2f} MB)...")
            sys.stdout.flush()
            
            # 2.5MB以上の巨大ファイルは解析除外
            if size_mb > 2.5:
                print(f"  Skipped (File too large)")
                sys.stdout.flush()
                out.write(f"■ ファイル: {rel_path} (スキップ: {size_mb:.2f} MB - 巨大ファイルのため除外)\n\n")
                out.flush()
                continue
                
            out.write(f"■ ファイル: {rel_path}\n")
            
            try:
                # read_only=True & data_only=True
                wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                ws = wb.active
            except Exception as e:
                print(f"  Error loading file: {e}")
                sys.stdout.flush()
                out.write(f"  読み込みエラー: {e}\n\n")
                out.flush()
                continue
                
            header_row = 3
            col_mapping = {}
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=header_row, column=col).value
                if cell_val:
                    col_mapping[str(cell_val).replace("\n", "").strip()] = col
                    
            # 1-indexedの列番号を取得
            col_old_price = col_mapping.get("売価", 15)
            col_old_shikiri = col_mapping.get("仕切", 16)
            col_new_shikiri = col_mapping.get("仕切換算", col_mapping.get("新仕切", col_mapping.get("改定仕切", 24)))
            col_new_price = col_mapping.get("売価1", col_mapping.get("新売価", col_mapping.get("新）売価", 32)))
            col_comp_as = col_mapping.get("AS換算売価", col_mapping.get("換算金額", col_mapping.get("ｱｽｸﾙｽﾗｲﾄﾞ", 29)))
            col_slide_p = col_mapping.get("スライド売価", 26)
            col_askul_slide_p = col_mapping.get("ｱｽｸﾙｽﾗｲﾄﾞ", 30)
            col_item_name = col_mapping.get("商品名", col_mapping.get("品目TEXT", 8))
            
            file_valid = 0
            file_patterns = {k: 0 for k in global_patterns.keys()}
            file_prices = []
            
            max_r = ws.max_row
            if max_r is None:
                max_r = 1000
                
            if max_r > 5000:
                print(f"  Warning: Row count is large ({max_r}). Limiting to first 2000 rows.")
                sys.stdout.flush()
                max_r = 2000
                
            # iter_rowsで一括高速取得（values_only=Trueにより高速）
            try:
                rows_data = list(ws.iter_rows(min_row=4, max_row=max_r, values_only=True))
            except Exception as e:
                print(f"  Error reading rows: {e}")
                sys.stdout.flush()
                wb.close()
                continue
                
            # インデックスの安全な取得
            def get_val_safe(row_tuple, col_1_indexed):
                idx = col_1_indexed - 1
                if idx >= 0 and idx < len(row_tuple):
                    return row_tuple[idx]
                return None
            
            # 空行連続カウント用
            consecutive_empty = 0
            
            for row in rows_data:
                old_p = get_val_safe(row, col_old_price)
                new_p = get_val_safe(row, col_new_price)
                
                # 空行検出
                if old_p is None and new_p is None:
                    consecutive_empty += 1
                    if consecutive_empty >= 10:
                        break
                    continue
                else:
                    consecutive_empty = 0
                    
                old_s = get_val_safe(row, col_old_shikiri)
                new_s = get_val_safe(row, col_new_shikiri)
                comp_p = get_val_safe(row, col_comp_as)
                slide_p = get_val_safe(row, col_slide_p)
                askul_slide = get_val_safe(row, col_askul_slide_p)
                
                try:
                    old_p = float(old_p)
                    new_p = float(new_p)
                    old_s = float(old_s) if old_s is not None else 0.0
                    new_s = float(new_s) if new_s is not None else 0.0
                    comp_p = float(comp_p) if (comp_p is not None and str(comp_p).strip() not in ["-", ""]) else None
                    slide_p = float(slide_p) if (slide_p is not None and str(slide_p).strip() not in ["-", ""]) else None
                    askul_slide = float(askul_slide) if (askul_slide is not None and str(askul_slide).strip() not in ["-", ""]) else None
                except (ValueError, TypeError):
                    continue
                    
                file_valid += 1
                total_valid_rows += 1
                file_prices.append(int(new_p))
                all_prices.append(int(new_p))
                
                if abs(new_p - old_p) < 0.1:
                    file_patterns["現行据え置き (価格差0)"] += 1
                    global_patterns["現行据え置き (価格差0)"] += 1
                    continue
                    
                if comp_p is not None:
                    if abs(new_p - (comp_p - 1)) <= 1 or abs(new_p - comp_p) <= 1:
                        file_patterns["競合価格対抗 (アスクル等 -1円〜-0円など)"] += 1
                        global_patterns["競合価格対抗 (アスクル等 -1円〜-0円など)"] += 1
                        continue
                        
                if askul_slide is not None:
                    if abs(new_p - askul_slide) <= 2 or abs(new_p - int(askul_slide)) == 0 or abs(new_p - round_to_nearest_10(askul_slide)) <= 2:
                        file_patterns["アスクルスライド値 (切り捨て・端数処理含む)"] += 1
                        global_patterns["アスクルスライド値 (切り捨て・端数処理含む)"] += 1
                        continue
                        
                if slide_p is not None:
                    if abs(new_p - slide_p) <= 2 or abs(new_p - int(slide_p)) == 0:
                        file_patterns["自社原価スライド値 (スライド売価)"] += 1
                        global_patterns["自社原価スライド値 (スライド売価)"] += 1
                        continue
                    elif abs(new_p - round_to_nearest_10(slide_p)) <= 2 or abs(new_p - round_to_nearest_100(slide_p)) <= 5:
                        file_patterns["自社原価スライド値 (丸め調整)"] += 1
                        global_patterns["自社原価スライド値 (丸め調整)"] += 1
                        continue
                        
                if new_p > 0:
                    margin = (new_p - new_s) / new_p
                    if any(abs(margin - m) <= 0.005 for m in [0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40]):
                        file_patterns["粗利率基準 (15%, 20%, 25%, 30%張り付き)"] += 1
                        global_patterns["粗利率基準 (15%, 20%, 25%, 30%張り付き)"] += 1
                        continue
                        
                file_patterns["その他個別調整"] += 1
                global_patterns["その他個別調整"] += 1
                
            out.write(f"  解析有効データ数: {file_valid}行\n")
            if file_valid > 0:
                last_2 = [p % 100 for p in file_prices]
                c_2 = Counter(last_2).most_common(2)
                out.write("  代表的な下2桁: " + ", ".join(f"末尾{k:02d}円 ({v}件)" for k, v in c_2) + "\n")
                for k, v in file_patterns.items():
                    out.write(f"    - {k}: {v}件 ({v/file_valid*100:.1f}%)\n")
            out.write("-" * 50 + "\n\n")
            out.flush()
            wb.close()
            
        out.write("\n==================================================\n")
        out.write("総合集計結果 (サブフォルダ全行合計)\n")
        out.write(f"総有効データ行数: {total_valid_rows}\n")
        out.write("==================================================\n")
        for k, v in global_patterns.items():
            rate = v / total_valid_rows * 100 if total_valid_rows > 0 else 0
            out.write(f"- {k}: {v}件 ({rate:.1f}%)\n")
        out.flush()
            
    print(f"Analysis completed for subfolders. Output written to subfolders_analysis_report.txt")
    sys.stdout.flush()

if __name__ == "__main__":
    scan_and_analyze()
