import openpyxl

path = r'C:\Users\フォーレスト026\Desktop\伊藤作業用\残業調査\20260427_ログ表示結果-20260416175217 - コピー.xlsx'
try:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['管理部提出用']
    print(f'Sheet: {ws.title}')
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        # 2列目(氏名)にお二人の名前があるか確認
        name = str(row[1]).replace(' ', '').replace('　', '') if len(row) > 1 else ""
        if '岸川' in name or '山本' in name:
            print(f"Row {r_idx} (Name: {name}):")
            # 全列の値を表示
            for c_idx, val in enumerate(row, 1):
                if val:
                    print(f"  Col {c_idx} ({openpyxl.utils.get_column_letter(c_idx)}): {val}")
except Exception as e:
    print(f"Error: {e}")
