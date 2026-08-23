import os
import openpyxl

TARGET_DIR = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\価格改定についてのツール化検証\参考資料"

files = [
    "●20260710【114】7月末_アイリスオーヤマ価格改定_173件.xlsx",
    "20260713【797】7月末_カグクロ価格改定_30件.xlsx",
    "●20260617【887】7月末_ジェムコ価格改定_105件.xlsx",
    "●20260626【982】7月末_ジョインテックス価格改定_64件.xlsx",
    "●20260701【291】7月末_今村紙工価格改定_47件.xlsx",
    "●20260702【740】7月末_共和価格改定_27件_売価2あり.xlsx",
    "●20260708【9323】7月末_福井価格改定_22件.xlsx"
]

def apply_smart_pricing_rule(base_val):
    """
    担当者の端数処理クセを再現するスマート丸めロジック
    """
    if base_val is None or base_val <= 0:
        return 0
        
    val = round(base_val)
    
    # 100円未満
    if val < 100:
        # 例: 46.6円 -> 46円 または 47円。下一桁をそのままか、キリ良く丸める
        # 下一桁が8, 9なら9にする、それ以外はそのまま
        last_digit = val % 10
        if last_digit in [8, 9]:
            return (val // 10) * 10 + 9
        return val
        
    # 100円以上 1,000円未満
    elif val < 1000:
        last_digit = val % 10
        # 下一桁が8, 9なら 9円にする (例: 729.54 -> 729)
        if last_digit in [8, 9]:
            return (val // 10) * 10 + 9
        # それ以外は10円単位に丸める (例: 671 -> 670)
        return round(val / 10) * 10
        
    # 1,000円以上
    else:
        # 1000円以上は、下二桁が「80円」「00円」「50円」に丸められる傾向が強い
        last_two = val % 100
        if last_two >= 75:
            # 80円か00円に丸める
            if last_two >= 90:
                return (val // 100) * 100 + 100 # 切り上げ00円
            return (val // 100) * 100 + 80 # 80円
        elif last_two >= 40 and last_two <= 60:
            return (val // 100) * 100 + 50 # 50円
        else:
            # 最も近い10円単位に丸める
            return round(val / 10) * 10

def simulate_pricing():
    print("=== 手作業の値（売価1）を再現するシミュレーション開始 ===")
    
    total_valid = 0
    exact_matches = 0
    near_matches_5 = 0 # 誤差5円以内
    near_matches_10 = 0 # 誤差10円以内
    
    for filename in files:
        filepath = os.path.join(TARGET_DIR, filename)
        if not os.path.exists(filepath):
            continue
            
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            ws = wb.active
        except Exception:
            continue
            
        # ヘッダー検出
        header_row = 3
        col_mapping = {}
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col).value
            if cell_val:
                col_mapping[str(cell_val).replace("\n", "").strip()] = col
                
        col_old_price = col_mapping.get("売価", 15)
        col_old_shikiri = col_mapping.get("仕切", 16)
        col_new_shikiri = col_mapping.get("仕切換算", col_mapping.get("新仕切", col_mapping.get("改定仕切", 24)))
        col_new_price = col_mapping.get("売価1", col_mapping.get("新売価", col_mapping.get("新）売価", 32)))
        col_comp_as = col_mapping.get("AS換算売価", col_mapping.get("換算金額", col_mapping.get("ｱｽｸﾙｽﾗｲﾄﾞ", 29)))
        col_slide_p = col_mapping.get("スライド売価", 26)
        col_askul_slide_p = col_mapping.get("ｱｽｸﾙｽﾗｲﾄﾞ", 30)
        
        file_valid = 0
        file_exact = 0
        file_near_5 = 0
        
        for r in range(4, ws.max_row + 1):
            old_p = ws.cell(row=r, column=col_old_price).value
            actual_new_p = ws.cell(row=r, column=col_new_price).value
            old_s = ws.cell(row=r, column=col_old_shikiri).value
            new_s = ws.cell(row=r, column=col_new_shikiri).value
            comp_p = ws.cell(row=r, column=col_comp_as).value
            slide_p = ws.cell(row=r, column=col_slide_p).value
            askul_slide = ws.cell(row=r, column=col_askul_slide_p).value
            
            if old_p is None or actual_new_p is None:
                continue
                
            try:
                old_p = float(old_p)
                actual_new_p = float(actual_new_p)
                old_s = float(old_s) if old_s is not None else 0.0
                new_s = float(new_s) if new_s is not None else 0.0
                comp_p = float(comp_p) if (comp_p is not None and str(comp_p).strip() not in ["-", ""]) else None
                slide_p = float(slide_p) if (slide_p is not None and str(slide_p).strip() not in ["-", ""]) else None
                askul_slide = float(askul_slide) if (askul_slide is not None and str(askul_slide).strip() not in ["-", ""]) else None
            except ValueError:
                continue
                
            # --- ロジックシミュレーション ---
            # 1. 競合スライド値を最優先で候補にする
            # 2. 次に自社スライド値を候補にする
            # 3. どちらもない場合は旧売価から計算
            
            candidate_price = None
            
            if askul_slide is not None and askul_slide > 0:
                candidate_price = askul_slide
            elif slide_p is not None and slide_p > 0:
                candidate_price = slide_p
            else:
                # 代替計算
                rate = new_s / old_s if old_s > 0 else 1.0
                candidate_price = old_p * rate
                
            # スマート丸めルールの適用
            simulated_price = apply_smart_pricing_rule(candidate_price)
            
            # 粗利下限チェック (15%)
            if simulated_price > 0:
                margin = (simulated_price - new_s) / simulated_price
                if margin < 0.15:
                    simulated_price = apply_smart_pricing_rule(new_s / 0.85)
            
            # 統計
            file_valid += 1
            total_valid += 1
            
            error = abs(simulated_price - actual_new_p)
            if error < 0.1:
                file_exact += 1
                exact_matches += 1
            if error <= 5:
                file_near_5 += 1
                near_matches_5 += 1
            if error <= 10:
                near_matches_10 += 1
                
        wb.close()
        
        if file_valid > 0:
            print(f"■ {filename}:")
            print(f"  解析数: {file_valid}行")
            print(f"  完全一致率: {file_exact/file_valid*100:.1f}% ({file_exact}件)")
            print(f"  誤差5円以内率: {file_near_5/file_valid*100:.1f}% ({file_near_5}件)")
            print("-" * 50)
            
    print("\n================総合結果=================")
    print(f"総検証データ行数: {total_valid}行")
    print(f"完全一致率: {exact_matches/total_valid*100:.1f}% ({exact_matches}件)")
    print(f"誤差5円以内率 (ほぼ手作業と同等): {near_matches_5/total_valid*100:.1f}% ({near_matches_5}件)")
    print(f"誤差10円以内率: {near_matches_10/total_valid*100:.1f}% ({near_matches_10}件)")
    print("=========================================")

if __name__ == "__main__":
    simulate_pricing()
