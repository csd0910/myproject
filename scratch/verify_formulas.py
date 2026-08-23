import os
import openpyxl

TARGET_DIR = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\価格改定についてのツール化検証\参考資料"
TARGET_FILE = "●20260710【114】7月末_アイリスオーヤマ価格改定_173件.xlsx"

wb = openpyxl.load_workbook(os.path.join(TARGET_DIR, TARGET_FILE), data_only=True, read_only=True)
ws = wb.active

# 3行目の列名の逆引き
header_row = 3
col_mapping = {}
for col in range(1, ws.max_column + 1):
    cell_val = ws.cell(row=header_row, column=col).value
    if cell_val:
        col_mapping[str(cell_val).replace("\n", "").strip()] = col

# 主要列
col_item_name = col_mapping.get("商品名", 8)
col_new_price = col_mapping.get("売価1", 32) # AF
col_slide_price = col_mapping.get("スライド売価", 26) # Z
col_askul_slide = col_mapping.get("ｱｽｸﾙｽﾗｲﾄﾞ", 30) # AD
col_competitor_as = col_mapping.get("AS換算売価", 29) # AC
col_new_shikiri = col_mapping.get("仕切換算", 24) # X

print("=== アイリスオーヤマ価格改定の決定プロセス深掘り ===")
print(f"解析対象行: 4〜25行目")
print("-" * 80)

for r in range(4, 26):
    name = ws.cell(row=r, column=col_item_name).value
    final_p = ws.cell(row=r, column=col_new_price).value
    slide_p = ws.cell(row=r, column=col_slide_price).value
    askul_slide = ws.cell(row=r, column=col_askul_slide).value
    askul_p = ws.cell(row=r, column=col_competitor_as).value
    new_s = ws.cell(row=r, column=col_new_shikiri).value
    
    if final_p is None:
        continue
        
    try:
        final_p = float(final_p)
        slide_p = float(slide_p) if slide_p is not None else 0.0
        askul_slide = float(askul_slide) if askul_slide is not None else 0.0
        askul_p = float(askul_p) if askul_p is not None else 0.0
        new_s = float(new_s) if new_s is not None else 0.0
    except ValueError:
        continue
        
    print(f"行 {r}: {name}")
    print(f"  [決定売価1]: {final_p}円 (新仕切: {new_s}円, 粗利率: {(final_p - new_s)/final_p*100:.1f}%)")
    print(f"  [参考指標値]:")
    print(f"    - スライド売価 (現行売価×原価値上率): {slide_p}円")
    print(f"    - アスクル換算売価 (競合価格): {askul_p}円")
    print(f"    - アスクルスライド (競合価格×原価値上率): {askul_slide}円")
    
    # 決定基準の推定
    matches = []
    if abs(final_p - slide_p) <= 2:
        matches.append("スライド売価に近似")
    if abs(final_p - askul_p) <= 2:
        matches.append("アスクル換算売価に対抗")
    if abs(final_p - askul_slide) <= 2:
        matches.append("アスクルスライド値に近似")
    # 切り捨て・丸め考慮
    if abs(final_p - int(askul_slide)) == 0:
        matches.append("アスクルスライド値の端数切り捨て")
    if abs(final_p - int(slide_p)) == 0:
        matches.append("スライド売価の端数切り捨て")
        
    print(f"  => 推定ロジック: {', '.join(matches) if matches else '手動による特殊調整（他社比較等）'}")
    print("-" * 80)

wb.close()
