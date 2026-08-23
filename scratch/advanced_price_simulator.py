import os
import openpyxl

TARGET_DIR = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\価格改定についてのツール化検証\参考資料"

files = [
    "●20260710【114】7月末_アイリスオーヤマ価格改定_173件.xlsx",
    "20260713【797】7月末_カグクロ価格改定_30件.xlsx",
    "●20260617【887】7月末_ジェムコ価格改定_105件.xlsx",
    "●20260626【982】7月末_ジョインテックス価格改定_64件.xlsx",
    "●20260701【291】7月末_今村紙工価格改定_47件.xlsx",
    "●20260702【740】7月末_共和価格改定_27件_売価2あり.xlsx",
    "●20260708【9323】7月末_福井価格改定_22件.xlsx"
]

def apply_specialized_rounding(val, file_type):
    """
    ファイルタイプ（メーカー）ごとの特化型丸めアルゴリズム
    """
    if val is None or val <= 0:
        return 0
        
    val = round(val)
    
    # 1. 家具特化型 (カグクロ)
    # 下二桁は常に00円、下三桁は800円、500円、000円に寄せる
    if file_type == "カグクロ":
        hundreds = val % 1000
        base = (val // 1000) * 1000
        # 800, 500, 000 に最も近いものを選択
        diffs = [
            (abs(hundreds - 0), 0),
            (abs(hundreds - 500), 500),
            (abs(hundreds - 800), 800),
            (abs(hundreds - 1000), 1000)
        ]
        chosen_diff, chosen_val = min(diffs, key=lambda x: x[0])
        if chosen_val == 1000:
            return base + 1000
        return base + chosen_val
        
    # 2. 小売・店舗特化型 (福井など)
    # 下二桁を 98円 または 80円 に寄せる
    elif file_type == "福井":
        tens = val % 100
        base = (val // 100) * 100
        if abs(tens - 98) < abs(tens - 80):
            return base + 98
        else:
            return base + 80
            
    # 3. 汎用・一般事務用品 (アイリス、ジェムコ、ジョインテックス等)
    else:
        if val < 100:
            return round(val / 10) * 10
        elif val < 1000:
            # 下二桁を 80, 00, 50, 90 に丸める
            tens = val % 100
            base = (val // 100) * 100
            targets = [0, 50, 80, 90, 100]
            chosen_val = min(targets, key=lambda x: abs(tens - x))
            if chosen_val == 100:
                return base + 100
            return base + chosen_val
        else:
            # 1,000円以上は下二桁を 80, 00, 50 に丸める
            tens = val % 100
            base = (val // 100) * 100
            targets = [0, 50, 80, 100]
            chosen_val = min(targets, key=lambda x: abs(tens - x))
            if chosen_val == 100:
                return base + 100
            return base + chosen_val

def simulate_advanced_pricing():
    print("=== 特化型丸めロジックを適用した高精度シミュレーション ===")
    
    total_valid = 0
    exact_matches = 0
    near_matches_5 = 0
    near_matches_10 = 0
    
    for filename in files:
        filepath = os.path.join(TARGET_DIR, filename)
        if not os.path.exists(filepath):
            continue
            
        # ファイルタイプ判定
        file_type = "汎用"
        if "カグクロ" in filename:
            file_type = "カグクロ"
        elif "福井" in filename:
            file_type = "福井"
            
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            ws = wb.active
        except Exception:
            continue
            
        # ヘッダー検出
        header_row = 3
        col_mapping = {}
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col).value
            if cell_val:
                col_mapping[str(cell_val).replace("\n", "").strip()] = col
                
        col_old_price = col_mapping.get("売価", 15)
        col_old_shikiri = col_mapping.get("仕切", 16)
        col_new_shikiri = col_mapping.get("仕切換算", col_mapping.get("新仕切", col_mapping.get("改定仕切", 24)))
        col_new_price = col_mapping.get("売価1", col_mapping.get("新売価", col_mapping.get("新）売価", 32)))
        col_comp_as = col_mapping.get("AS換算売価", col_mapping.get("換算金額", col_mapping.get("ｱｽｸﾙｽﾗｲﾄﾞ", 29)))
        col_slide_p = col_mapping.get("スライド売価", 26)
        col_askul_slide_p = col_mapping.get("ｱｽｸﾙｽﾗｲﾄﾞ", 30)
        
        file_valid = 0
        file_exact = 0
        file_near_5 = 0
        
        for r in range(4, ws.max_row + 1):
            old_p = ws.cell(row=r, column=col_old_price).value
            actual_new_p = ws.cell(row=r, column=col_new_price).value
            old_s = ws.cell(row=r, column=col_old_shikiri).value
            new_s = ws.cell(row=r, column=col_new_shikiri).value
            comp_p = ws.cell(row=r, column=col_comp_as).value
            slide_p = ws.cell(row=r, column=col_slide_p).value
            askul_slide = ws.cell(row=r, column=col_askul_slide_p).value
            
            if old_p is None or actual_new_p is None:
                continue
                
            try:
                old_p = float(old_p)
                actual_new_p = float(actual_new_p)
                old_s = float(old_s) if old_s is not None else 0.0
                new_s = float(new_s) if new_s is not None else 0.0
                comp_p = float(comp_p) if (comp_p is not None and str(comp_p).strip() not in ["-", ""]) else None
                slide_p = float(slide_p) if (slide_p is not None and str(slide_p).strip() not in ["-", ""]) else None
                askul_slide = float(askul_slide) if (askul_slide is not None and str(askul_slide).strip() not in ["-", ""]) else None
            except ValueError:
                continue
                
            candidate_price = None
            
            # ロジック適用
            # もし競合スライドがあり、粗利を確保できるなら競合スライドをベースにする
            # なければ自社スライド
            if askul_slide is not None and askul_slide > 0:
                # 競合価格が高すぎる場合（スライド価格の1.2倍を超えるなど）は自社スライドに戻す
                if slide_p is not None and askul_slide > slide_p * 1.2:
                    candidate_price = slide_p
                else:
                    candidate_price = askul_slide
            elif slide_p is not None and slide_p > 0:
                candidate_price = slide_p
            else:
                rate = new_s / old_s if old_s > 0 else 1.0
                candidate_price = old_p * rate
                
            # 特化型丸めの適用
            simulated_price = apply_specialized_rounding(candidate_price, file_type)
            
            # 粗利下限チェック (15%)
            if simulated_price > 0:
                margin = (simulated_price - new_s) / simulated_price
                if margin < 0.15:
                    simulated_price = apply_specialized_rounding(new_s / 0.85, file_type)
            
            file_valid += 1
            total_valid += 1
            
            error = abs(simulated_price - actual_new_p)
            if error < 0.1:
                file_exact += 1
                exact_matches += 1
            if error <= 5:
                file_near_5 += 1
                near_matches_5 += 1
            if error <= 10:
                near_matches_10 += 1
                
        wb.close()
        
        if file_valid > 0:
            print(f"■ {filename} ({file_type}モード):")
            print(f"  解析数: {file_valid}行")
            print(f"  完全一致率: {file_exact/file_valid*100:.1f}% ({file_exact}件)")
            print(f"  誤差5円以内率: {file_near_5/file_valid*100:.1f}% ({file_near_5}件)")
            print("-" * 50)
            
    print("\n================総合結果=================")
    print(f"総検証データ行数: {total_valid}行")
    print(f"完全一致率: {exact_matches/total_valid*100:.1f}% ({exact_matches}件)")
    print(f"誤差5円以内率 (ほぼ手作業と同等): {near_matches_5/total_valid*100:.1f}% ({near_matches_5}件)")
    print(f"誤差10円以内率: {near_matches_10/total_valid*100:.1f}% ({near_matches_10}件)")
    print("=========================================")

if __name__ == "__main__":
    simulate_advanced_pricing()
