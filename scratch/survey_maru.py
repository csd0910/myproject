import openpyxl
import os

# デスクトップの残業調査フォルダ内のファイルを対象にする
search_dir = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\残業調査"
for filename in os.listdir(search_dir):
    if filename.endswith((".xlsx", ".xlsm")):
        path = os.path.join(search_dir, filename)
        print(f"\n--- Checking: {filename} ---")
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            if '管理部提出用' in wb.sheetnames:
                ws = wb['管理部提出用']
                found = False
                for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    for c_idx, val in enumerate(row, 1):
                        if val == '〇' or val == '○':
                            print(f"  Found '{val}' at Row {r_idx}, Col {c_idx}")
                            found = True
                if not found:
                    print("  No '〇' found in this sheet.")
            else:
                print("  '管理部提出用' sheet not found.")
        except Exception as e:
            print(f"  Error: {e}")
