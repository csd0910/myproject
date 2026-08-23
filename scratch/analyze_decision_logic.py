import os
import openpyxl
import numpy as np

TARGET_DIR = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\価格改定についてのツール化検証\参考資料"

# 分析対象ファイルリスト
files = [
    "●20260710【114】7月末_アイリスオーヤマ価格改定_173件.xlsx",
    "20260713【797】7月末_カグクロ価格改定_30件.xlsx",
    "●20260701【291】7月末_今村紙工価格改定_47件.xlsx",
    "●20260702【740】7月末_共和価格改定_27件_売価2あり.xlsx",
    "●20260708【9323】7月末_福井価格改定_22件.xlsx",
    "●20260617【887】7月末_ジェムコ価格改定_105件.xlsx",
    "●20260626【982】7月末_ジョインテックス価格改定_64件.xlsx"
]

def analyze_file(filename):
    filepath = os.path.join(TARGET_DIR, filename)
    if not os.path.exists(filepath):
        return None
        
    try:
        # 値を読み取るために data_only=True
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        print(f"Error loading {filename}: {e}")
        return None
        
    # ヘッダー検出
    header_row = 3
    col_mapping = {}
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row, column=col).value
        if cell_val:
            cell_val_clean = str(cell_val).replace("\n", "").replace(" ", "").strip()
            col_mapping[cell_val_clean] = col
            
    # 表記揺れ対応の列インデックス取得
    def find_col(possible_names):
        for name in possible_names:
            if name in col_mapping:
                return col_mapping[name]
        return None
        
    col_old_price = find_col(["売価"])
    col_old_shikiri = find_col(["仕切"])
    col_new_shikiri = find_col(["仕切換算", "新仕切", "改定仕切"]) or find_col(["仕切"]) + 8 # 推定ズレ
    # アイリス等ではWかX列（仕切換算=24）
    if "仕切換算" in col_mapping:
        col_new_shikiri = col_mapping["仕切換算"]
    elif "仕切" in col_mapping and "売価1" in col_mapping:
        # 仕切と売価1の間にある「仕切」系の列を探す
        for col_name, col_idx in col_mapping.items():
            if "仕切" in col_name and col_idx > col_mapping["仕切"] and col_idx < col_mapping["売価1"]:
                col_new_shikiri = col_idx
                break

    col_new_price = find_col(["売価1", "決定売価", "新売価"])
    col_comp_as = find_col(["AS換算売価", "換算金額", "アスクル換算売価"])
    col_retail_old = find_col(["小売価格(税抜)", "定価"])
    col_retail_new = find_col(["小売価格(税抜)", "定価"]) # 同一列で処理される場合あり
    
    # 手動フォールバック（アイリス、今村、カグクロ等の典型構造）
    if not col_old_price: col_old_price = 15 # O列
    if not col_old_shikiri: col_old_shikiri = 16 # P列
    if not col_new_shikiri: col_new_shikiri = 24 # X列
    if not col_new_price: col_new_price = 32 # AF列
    if not col_comp_as: col_comp_as = 29 # AC列
    if not col_retail_old: col_retail_old = 14 # N列
    
    stats = {
        "total_rows": 0,
        "valid_rows": 0,
        "comp_対抗": 0,
        "原価スライド": 0,
        "現行据え置き": 0,
        "粗利下限張り付き": 0,
        "その他手動": 0,
        "値上げ件数": 0,
        "値下げ件数": 0,
        "据え置き件数": 0
    }
    
    # 4行目からスキャン
    for r in range(4, ws.max_row + 1):
        old_p = ws.cell(row=r, column=col_old_price).value
        new_p = ws.cell(row=r, column=col_new_price).value
        old_s = ws.cell(row=r, column=col_old_shikiri).value
        new_s = ws.cell(row=r, column=col_new_shikiri).value
        comp_p = ws.cell(row=r, column=col_comp_as).value
        
        if old_p is None or new_p is None:
            continue
            
        stats["total_rows"] += 1
        try:
            old_p = float(old_p)
            new_p = float(new_p)
            old_s = float(old_s) if old_s is not None else 0.0
            new_s = float(new_s) if new_s is not None else 0.0
        except ValueError:
            continue
            
        stats["valid_rows"] += 1
        
        # 値上げ・値下げ・据え置きカウント
        if new_p > old_p:
            stats["値上げ件数"] += 1
        elif new_p < old_p:
            stats["値下げ件数"] += 1
        else:
            stats["据え置き件数"] += 1
            
        # 決定ロジックの分類
        # 1. 現行据え置き
        if abs(new_p - old_p) < 0.1:
            stats["現行据え置き"] += 1
            continue
            
        # 2. 競合（アスクル）対抗 (競合価格-5円〜+5円の範囲)
        has_comp = False
        if comp_p is not None and str(comp_p).strip() not in ["-", ""]:
            try:
                comp_p = float(comp_p)
                has_comp = True
            except ValueError:
                pass
                
        if has_comp and abs(new_p - comp_p) <= 5:
            stats["comp_対抗"] += 1
            continue
            
        # 3. 原価スライド (原価の上昇率と、売価の上昇率の差が±2%以内)
        if old_s > 0 and old_p > 0:
            shikiri_rate = new_s / old_s
            price_rate = new_p / old_p
            if abs(shikiri_rate - price_rate) <= 0.02:
                stats["原価スライド"] += 1
                continue
                
        # 4. 粗利率下限（粗利率が10%, 15%, 20%, 25%などのキリの良い数字に張り付いているか）
        if new_p > 0:
            new_margin = (new_p - new_s) / new_p
            # 10%, 15%, 20%, 25%, 30%付近
            is_flat_margin = False
            for target_m in [0.10, 0.15, 0.20, 0.25, 0.30]:
                if abs(new_margin - target_m) <= 0.005:
                    is_flat_margin = True
                    break
            if is_flat_margin:
                stats["粗利下限張り付き"] += 1
                continue
                
        # 5. その他手動判定
        stats["black_box_手動"] = stats.get("black_box_手動", 0) + 1
        stats["その他手動"] += 1

    wb.close()
    return stats

print("=== 過去の価格改定データから意思決定ロジックを抽出・可視化 ===")
for f in files:
    res = analyze_file(f)
    if res:
        print(f"\n■ ファイル名: {f}")
        print(f"  総データ行数: {res['total_rows']} (解析成功: {res['valid_rows']}件)")
        print(f"  価格改定の内訳: 値上げ={res['値上げ件数']}件, 値下げ={res['値下げ件数']}件, 据置={res['据え置き件数']}件")
        
        valid = res['valid_rows'] if res['valid_rows'] > 0 else 1
        print("  【意思決定パターンの構成比】")
        print(f"    1. 現行据え置き: {res['現行据え置き']}件 ({res['現行据え置き']/valid*100:.1f}%)")
        print(f"    2. 競合価格対抗: {res['comp_対抗']}件 ({res['comp_対抗']/valid*100:.1f}%)")
        print(f"    3. 原価連動スライド: {res['原価スライド']}件 ({res['原価スライド']/valid*100:.1f}%)")
        print(f"    4. 粗利下限張り付き: {res['粗利下限張り付き']}件 ({res['粗利下限張り付き']/valid*100:.1f}%)")
        print(f"    5. カン・個別調整（その他）: {res['その他手動']}件 ({res['その他手動']/valid*100:.1f}%)")
