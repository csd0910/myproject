import os
import openpyxl

TARGET_DIR = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\価格改定についてのツール化検証\参考資料"
TARGET_FILE = "●20260710【114】7月末_アイリスオーヤマ価格改定_173件_自動計算推奨版.xlsx"

wb = openpyxl.load_workbook(os.path.join(TARGET_DIR, TARGET_FILE), data_only=True)
ws = wb.active

# 列定義（プロトタイプで自動検出されたもの）
col_old_price = 15 # O列 (売価)
col_old_shikiri = 16 # P列 (仕切)
col_new_shikiri_calc = 24 # X列 (仕切換算)
col_competitor_as = 29 # AC列 (AS換算売価)
col_new_price1 = 32 # AF列 (売価1)
col_item_name = 8 # H列 (商品名)

print("=== 値上がり（旧売価より新推奨売価が高くなった）の具体例 ===")
count = 0
for r in range(4, ws.max_row + 1):
    old_p = ws.cell(row=r, column=col_old_price).value
    new_p = ws.cell(row=r, column=col_new_price1).value
    old_s = ws.cell(row=r, column=col_old_shikiri).value
    new_s = ws.cell(row=r, column=col_new_shikiri_calc).value
    comp_p = ws.cell(row=r, column=col_competitor_as).value
    name = ws.cell(row=r, column=col_item_name).value
    
    if old_p is not None and new_p is not None:
        try:
            old_p = float(old_p)
            new_p = float(new_p)
            diff = new_p - old_p
            if diff > 0 and count < 10:
                print(f"行 {r}: {name}")
                print(f"  現行売価: {old_p}円 (仕切: {old_s}円) -> 新推奨売価: {new_p}円 (新仕切: {new_s}円) [差額: +{diff}円]")
                print(f"  他社(アスクル)価格: {comp_p}円")
                # 原因の判定
                if comp_p is not None and str(comp_p).strip() not in ["-", ""]:
                    comp_val = float(comp_p)
                    if new_p == comp_val - 1:
                        print("  => 原因: 他社価格（値上がり後、または自社より高い）に合わせて対抗値上げ")
                    else:
                        print("  => 原因: 原価高騰による粗利率15%キープのための強制値上げ")
                else:
                    print("  => 原因: 競合なしのため、仕切価格の値上げ率(原価スライド)に連動した値上げ")
                print("-" * 50)
                count += 1
        except ValueError:
            continue

wb.close()
