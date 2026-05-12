import openpyxl
import os

search_dir = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\残業調査"
for filename in os.listdir(search_dir):
    if filename.endswith((".xlsx", ".xlsm")):
        path = os.path.join(search_dir, filename)
        print(f"\n--- Checking: {filename} ---")
        try:
            # 読み込みを速くするため、data_only=True
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            if '管理部提出用' in wb.sheetnames:
                ws = wb['管理部提出用']
                found = False
                # 最初の50行だけ見る
                for r_idx, row in enumerate(ws.iter_rows(max_row=50, values_only=True), 1):
                    for c_idx, val in enumerate(row, 1):
                        if str(val) in ['〇', '○']:
                            print(f"  Found '{val}' at Row {r_idx}, Col {c_idx}")
                            found = True
                if not found:
                    print("  No '〇' or '○' found in the first 50 rows.")
            else:
                print("  '管理部提出用' sheet not found.")
        except Exception as e:
            print(f"  Error: {e}")
