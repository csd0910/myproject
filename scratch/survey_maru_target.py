import openpyxl
import os

path = r'C:\Users\フォーレスト026\Desktop\伊藤作業用\残業調査\20260427_ログ表示結果-20260416175217 - コピー.xlsx'
try:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['管理部提出用']
    print(f'Checking sheet: {ws.title}')
    found = False
    for r_idx, row in enumerate(ws.iter_rows(max_row=300, values_only=True), 1):
        for c_idx, val in enumerate(row, 1):
            if str(val) in ['〇', '○']:
                col_letter = openpyxl.utils.get_column_letter(c_idx)
                print(f"  Found '{val}' at Row {r_idx}, Col {c_idx} ({col_letter}列)")
                found = True
    if not found:
        print("  No '〇' or '○' found in the first 300 rows.")
except Exception as e:
    print(f"  Error: {e}")
