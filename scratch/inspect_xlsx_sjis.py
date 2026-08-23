import os
import openpyxl
from openpyxl.utils import get_column_letter

TARGET_DIR = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\価格改定についてのツール化検証\参考資料"
OUTPUT_FILE = r"c:\Users\フォーレスト026\MyProject\scratch\inspect_result.txt"

files = [
    "●20260710【114】7月末_アイリスオーヤマ価格改定_173件.xlsx",
    "2026.08.01改定_ﾌｼﾞｺﾋﾟｱﾝ_契約条件変更フォーマット 1件.xlsx",
    "20260713【797】7月末_カグクロ価格改定_30件.xlsx",
    "●20260701【291】7月末_今村紙工価格改定_47件.xlsx"
]

def inspect_file(filepath, out_f):
    out_f.write(f"\n========================================\nFile: {os.path.basename(filepath)}\n")
    if not os.path.exists(filepath):
        out_f.write("File does not exist.\n")
        return
        
    try:
        wb_val = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        wb_form = openpyxl.load_workbook(filepath, data_only=False, read_only=True)
    except Exception as e:
        out_f.write(f"Error loading workbook: {e}\n")
        return
    
    for sheetname in wb_val.sheetnames:
        out_f.write(f"\n--- Sheet: {sheetname} ---\n")
        ws_val = wb_val[sheetname]
        ws_form = wb_form[sheetname]
        
        max_r = ws_val.max_row
        max_c = ws_val.max_column
        out_f.write(f"Dimensions: {max_r} rows x {max_c} cols\n")
        
        # Read header (first 5 rows)
        rows_val = list(ws_val.iter_rows(max_row=min(5, max_r or 5), max_col=max_c, values_only=True))
        rows_form = list(ws_form.iter_rows(max_row=min(5, max_r or 5), max_col=max_c, values_only=True))
        
        for r_idx in range(len(rows_val)):
            out_f.write(f"Row {r_idx + 1}:\n")
            for c_idx in range(len(rows_val[r_idx])):
                col_letter = get_column_letter(c_idx + 1)
                val = rows_val[r_idx][c_idx]
                form = rows_form[r_idx][c_idx]
                
                if val is not None or form is not None:
                    if isinstance(form, str) and form.startswith('='):
                        out_f.write(f"  {col_letter}{r_idx+1}: [Formula] {form} (Value: {val})\n")
                    else:
                        out_f.write(f"  {col_letter}{r_idx+1}: {val}\n")
                        
        # Inspect interest columns in row 4-10
        interest_cols = list(range(13, 18)) + list(range(29, 33)) + list(range(72, 77)) + list(range(17, 29))
        interest_cols = sorted(list(set(interest_cols)))
        
        out_f.write("\nChecking formulas in row 4-10 for interest columns:\n")
        limit_r = min(10, max_r or 10)
        for r in range(4, limit_r + 1):
            for c in interest_cols:
                if c <= max_c:
                    col_letter = get_column_letter(c)
                    val = ws_val.cell(row=r, column=c).value
                    form = ws_form.cell(row=r, column=c).value
                    if isinstance(form, str) and form.startswith('='):
                        out_f.write(f"  {col_letter}{r}: {form} -> {val}\n")
                    else:
                        out_f.write(f"  {col_letter}{r}: {val}\n")
            out_f.write("\n")
            
    wb_val.close()
    wb_form.close()

with open(OUTPUT_FILE, "w", encoding="utf-8") as out_f:
    for f in files:
        inspect_file(os.path.join(TARGET_DIR, f), out_f)

print("Inspection completed. Output written to inspect_result.txt")
