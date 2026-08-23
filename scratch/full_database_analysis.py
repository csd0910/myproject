import os
import openpyxl

TARGET_DIR = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\価格改定についてのツール化検証\参考資料"

files = [
    "●20260710【114】7月末_アイリスオーヤマ価格改定_173件.xlsx",
    "2026.08.01改定_ﾌｼﾞｺﾋﾟｱﾝ_契約条件変更フォーマット 1件.xlsx",
    "20260713【797】7月末_カグクロ価格改定_30件.xlsx",
    "20260716【994】7月末_三治商会価格改定_2件.xlsx",
    "●20260617【887】7月末_ジェムコ価格改定_105件.xlsx",
    "●20260624【768】時期確認_スリーエム（PB）価格改定_原価のみ変更 12件.xlsx",
    "●20260626【078】7月末_ハクバ写真M価と売価だけ改定（仕切は8月末）_1件.xlsx",
    "●20260626【982】7月末_ジョインテックス価格改定_64件.xlsx",
    "●20260701【291】7月末_今村紙工価格改定_47件.xlsx",
    "●20260702【472】7月末_カネゲン（ナカバヤシ）価格改定_22件.xlsx",
    "●20260702【740】7月末_共和価格改定_27件_売価2あり.xlsx",
    "●20260703【037】7月末_桜井価格改定_6件.xlsx",
    "●20260706【459.9259】7月末_中央物産価格改定_5件.xlsx",
    "●20260706【893】7月末_大和無線価格改定_5件.xlsx",
    "●20260708【9323】7月末_福井価格改定_22件.xlsx"
]

def round_to_nearest_10(val):
    return round(val / 10) * 10

def round_to_nearest_100(val):
    return round(val / 100) * 100

def analyze_all_lines():
    out_file = r"c:\Users\フォーレスト026\MyProject\scratch\full_lines_analysis.txt"
    out = open(out_file, "w", encoding="utf-8")
    
    out.write("==================================================\n")
    out.write("全ファイル・全行の価格決定ロジック網羅的解析\n")
    out.write("==================================================\n\n")
    
    total_processed_files = 0
    total_processed_rows = 0
    
    # 決定パターンの総合集計
    global_patterns = {
        "現行据え置き (価格差0)": 0,
        "競合価格対抗 (アスクル等 -1円〜-0円など)": 0,
        "アスクルスライド値 (切り捨て・端数処理含む)": 0,
        "自社原価スライド値 (スライド売価)": 0,
        "自社原価スライド値 (丸め調整)": 0,
        "粗利率基準 (15%, 20%, 25%, 30%張り付き)": 0,
        "その他個別調整": 0
    }
    
    for filename in files:
        filepath = os.path.join(TARGET_DIR, filename)
        if not os.path.exists(filepath):
            out.write(f"ファイルなし: {filename}\n")
            continue
            
        out.write(f"\n■ ファイル名: {filename}\n")
        total_processed_files += 1
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            ws = wb.active
        except Exception as e:
            out.write(f"  読み込みエラー: {e}\n")
            continue
            
        # ヘッダー検出 (3行目)
        header_row = 3
        # 2026.08.01改定_ﾌｼﾞｺﾋﾟｱﾝ_契約条件変更フォーマット のみ構造が違う可能性あり
        if "契約条件変更フォーマット" in filename or "ﾌｼﾞｺﾋﾟｱﾝ" in filename:
            header_row = 3
            
        col_mapping = {}
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col).value
            if cell_val:
                col_mapping[str(cell_val).replace("\n", "").strip()] = col
                
        # 列名マッピング
        col_old_price = col_mapping.get("売価", 15)
        col_old_shikiri = col_mapping.get("仕切", 16)
        col_new_shikiri = col_mapping.get("仕切換算", col_mapping.get("新仕切", col_mapping.get("改定仕切", 24)))
        col_new_price = col_mapping.get("売価1", col_mapping.get("新売価", col_mapping.get("新）売価", 32)))
        col_comp_as = col_mapping.get("AS換算売価", col_mapping.get("換算金額", col_mapping.get("ｱｽｸﾙｽﾗｲﾄﾞ", 29)))
        col_slide_p = col_mapping.get("スライド売価", 26)
        col_askul_slide_p = col_mapping.get("ｱｽｸﾙｽﾗｲﾄﾞ", 30)
        col_item_name = col_mapping.get("商品名", col_mapping.get("品目TEXT", 8))
        
        # 行のループ (4行目から最終行まで)
        file_rows = 0
        file_valid_rows = 0
        
        file_patterns = {k: 0 for k in global_patterns.keys()}
        
        for r in range(4, ws.max_row + 1):
            name = ws.cell(row=r, column=col_item_name).value
            old_p = ws.cell(row=r, column=col_old_price).value
            new_p = ws.cell(row=r, column=col_new_price).value
            old_s = ws.cell(row=r, column=col_old_shikiri).value
            new_s = ws.cell(row=r, column=col_new_shikiri).value
            comp_p = ws.cell(row=r, column=col_comp_as).value
            slide_p = ws.cell(row=r, column=col_slide_p).value
            askul_slide = ws.cell(row=r, column=col_askul_slide_p).value
            
            if old_p is None or new_p is None:
                continue
                
            file_rows += 1
            total_processed_rows += 1
            
            try:
                old_p = float(old_p)
                new_p = float(new_p)
                old_s = float(old_s) if old_s is not None else 0.0
                new_s = float(new_s) if new_s is not None else 0.0
                comp_p = float(comp_p) if (comp_p is not None and str(comp_p).strip() not in ["-", ""]) else None
                slide_p = float(slide_p) if (slide_p is not None and str(slide_p).strip() not in ["-", ""]) else None
                askul_slide = float(askul_slide) if (askul_slide is not None and str(askul_slide).strip() not in ["-", ""]) else None
            except ValueError:
                continue
                
            file_valid_rows += 1
            
            # 各種指標の算出
            # 1. 現行価格据え置き
            if abs(new_p - old_p) < 0.1:
                file_patterns["現行据え置き (価格差0)"] += 1
                global_patterns["現行据え置き (価格差0)"] += 1
                continue
                
            # 2. 競合価格対抗
            is_comp = False
            if comp_p is not None:
                # 対抗価格（競合価格-1円、または同値）
                if abs(new_p - (comp_p - 1)) <= 1 or abs(new_p - comp_p) <= 1:
                    file_patterns["競合価格対抗 (アスクル等 -1円〜-0円など)"] += 1
                    global_patterns["競合価格対抗 (アスクル等 -1円〜-0円など)"] += 1
                    continue
                    
            # 3. アスクルスライド
            if askul_slide is not None:
                # アスクルスライドの端数丸め
                if abs(new_p - askul_slide) <= 2 or abs(new_p - int(askul_slide)) == 0 or abs(new_p - round_to_nearest_10(askul_slide)) <= 2:
                    file_patterns["アスクルスライド値 (切り捨て・端数処理含む)"] += 1
                    global_patterns["アスクルスライド値 (切り捨て・端数処理含む)"] += 1
                    continue
                    
            # 4. 自社原価スライド (完全一致)
            if slide_p is not None:
                if abs(new_p - slide_p) <= 2 or abs(new_p - int(slide_p)) == 0:
                    file_patterns["自社原価スライド値 (スライド売価)"] += 1
                    global_patterns["自社原価スライド値 (スライド売価)"] += 1
                    continue
                # 5. 自社原価スライド (丸め)
                elif abs(new_p - round_to_nearest_10(slide_p)) <= 2 or abs(new_p - round_to_nearest_100(slide_p)) <= 5:
                    file_patterns["自社原価スライド値 (丸め調整)"] += 1
                    global_patterns["自社原価スライド値 (丸め調整)"] += 1
                    continue
                    
            # 6. 粗利率基準のキリの良さ
            if new_p > 0:
                margin = (new_p - new_s) / new_p
                if any(abs(margin - m) <= 0.005 for m in [0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40]):
                    file_patterns["粗利率基準 (15%, 20%, 25%, 30%張り付き)"] += 1
                    global_patterns["粗利率基準 (15%, 20%, 25%, 30%張り付き)"] += 1
                    continue
                    
            # 7. その他
            file_patterns["その他個別調整"] += 1
            global_patterns["その他個別調整"] += 1
            
        out.write(f"  総データ行数: {file_rows} (解析成功: {file_valid_rows}行)\n")
        if file_valid_rows > 0:
            for k, v in file_patterns.items():
                out.write(f"    - {k}: {v}件 ({v/file_valid_rows*100:.1f}%)\n")
        out.write("-" * 50 + "\n")
        wb.close()
        
    out.write("\n==================================================\n")
    out.write("総合集計結果 (全ファイル全行合計)\n")
    out.write(f"処理ファイル数: {total_processed_files}\n")
    out.write(f"総有効データ行数: {total_processed_rows}\n")
    out.write("==================================================\n")
    for k, v in global_patterns.items():
        rate = v / total_processed_rows * 100 if total_processed_rows > 0 else 0
        out.write(f"- {k}: {v}件 ({rate:.1f}%)\n")
        
    out.close()
    print("Full analysis completed. Results written to scratch/full_lines_analysis.txt")

if __name__ == "__main__":
    analyze_all_lines()
