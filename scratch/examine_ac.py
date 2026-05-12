import openpyxl

path = r'C:\Users\フォーレスト026\Desktop\伊藤作業用\残業調査\20260427_ログ表示結果-20260416175217 - コピー.xlsx'
try:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['管理部提出用']
    # AC列(29列目)の項目名
    header_ac = ws.cell(row=1, column=29).value
    print(f"Header of AC Column (Col 29): {header_ac}")
    
    # AC列に何か入っている行を探す
    for r_idx in range(2, 2301):
        val = ws.cell(row=r_idx, column=29).value
        if val:
            print(f"Row {r_idx}, AC Value: '{val}' (Type: {type(val)})")
            # 最初の5件くらい見つかったら止める
            if r_idx > 1000: break # 安全のため
except Exception as e:
    print(f"Error: {e}")
