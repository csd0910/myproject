import openpyxl

path = r'C:\Users\フォーレスト026\Desktop\伊藤作業用\残業調査\20260427_ログ表示結果-20260416175217 - コピー.xlsx'
try:
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Scanning Sheet: {sheet_name} ---")
        # シート全体の有効範囲を確認
        print(f"  Dimensions: {ws.dimensions}")
        
        # 全セルをスキャン
        for r_idx, row in enumerate(ws.iter_rows(max_row=200, values_only=True), 1):
            for c_idx, val in enumerate(row, 1):
                if val and str(val).strip() in ['〇', '○']:
                    col_letter = openpyxl.utils.get_column_letter(c_idx)
                    print(f"  [FOUND] '{val}' at Row {r_idx}, Col {c_idx} ({col_letter})")
except Exception as e:
    print(f"Error: {e}")
