import openpyxl
from openpyxl.styles import PatternFill

file_new = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\残業調査\管理部提出用リスト_20260610175823.xlsx"
file_old = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\残業調査\管理部提出用リスト_20260610172639.xlsx"
output = r"C:\Users\フォーレスト026\Desktop\伊藤作業用\残業調査\管理部提出用リスト_9時例外比較結果.xlsx"

try:
    wb_new = openpyxl.load_workbook(file_new)
    ws_new = wb_new.active
    
    wb_old = openpyxl.load_workbook(file_old)
    ws_old = wb_old.active

    old_data = {}
    for row in range(2, ws_old.max_row + 1):
        date_val = str(ws_old.cell(row=row, column=1).value or "").split('（')[0].strip()
        comp = str(ws_old.cell(row=row, column=3).value or "").strip()
        on_val = str(ws_old.cell(row=row, column=5).value or "").strip()
        off_val = str(ws_old.cell(row=row, column=6).value or "").strip()
        if date_val and comp:
            old_data[(date_val, comp)] = (on_val, off_val)
            
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    diff_count = 0
    for row in range(2, ws_new.max_row + 1):
        date_val_raw = str(ws_new.cell(row=row, column=1).value or "")
        date_val = date_val_raw.split('（')[0].strip()
        comp = str(ws_new.cell(row=row, column=3).value or "").strip()
        
        on_cell = ws_new.cell(row=row, column=5)
        off_cell = ws_new.cell(row=row, column=6)
        on_val = str(on_cell.value or "").strip()
        off_val = str(off_cell.value or "").strip()
        
        old_record = old_data.get((date_val, comp))
        if old_record:
            old_on, old_off = old_record
            changed = False
            if on_val != old_on:
                on_cell.fill = yellow_fill
                changed = True
            if off_val != old_off:
                off_cell.fill = yellow_fill
                changed = True
            if changed:
                diff_count += 1
        else:
            # 完全に新規の行（日付シフトにより生まれた行）
            ws_new.cell(row=row, column=1).fill = orange_fill
            on_cell.fill = yellow_fill
            off_cell.fill = yellow_fill
            diff_count += 1

    wb_new.save(output)
    print(f"Comparison complete. {diff_count} rows have differences. Saved to {output}")
except Exception as e:
    print(f"Error: {e}")
