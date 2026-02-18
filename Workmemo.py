import tkinter as tk
from tkinter import ttk, messagebox
import datetime
import csv
import os
import json
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import japanize_matplotlib
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, Border, Side

# --- 定数とデフォルト設定 ---
CONFIG_FILE = "config.json"
DEFAULT_CSV_PATH = "work_log.csv"

DEFAULT_CONFIG = {
    "user_name": "伊藤",
    "categories": ["障害", "運用", "保守", "その他"],
    "genres": ["OA端末関連", "SCAW関連", "総務関連", "ネットワーク"],
    "departments": ["商品部", "物流部", "営業管理部", "人事部"],
    "importance": ["高", "中", "低"]
}

class ConfigManager:
    @staticmethod
    def load():
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf_8_sig") as f:
                    config = json.load(f)
                    for k, v in DEFAULT_CONFIG.items():
                        if k not in config: config[k] = v
                    return config
            except:
                pass
        return DEFAULT_CONFIG.copy()

    @staticmethod
    def save(config):
        with open(CONFIG_FILE, "w", encoding="utf_8_sig") as f:
            json.dump(config, f, ensure_ascii=False, indent=4)

class SettingsWindow(tk.Toplevel):
    def __init__(self, parent, config, on_save):
        super().__init__(parent)
        self.title("設定")
        self.geometry("380x480")
        self.config_data = config
        self.on_save = on_save
        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(main_frame, text="担当者名:").pack(anchor=tk.W)
        self.ent_user = ttk.Entry(main_frame)
        self.ent_user.insert(0, self.config_data.get("user_name", ""))
        self.ent_user.pack(fill=tk.X, pady=(0, 10))

        self.edits = {}
        fields = [("categories", "分類"), ("genres", "ジャンル"), ("departments", "部署")]
        for key, label in fields:
            ttk.Label(main_frame, text=f"{label} (カンマ区切り):").pack(anchor=tk.W)
            txt = tk.Text(main_frame, height=3, font=("Segoe UI", 9))
            txt.insert("1.0", ",".join(self.config_data.get(key, [])))
            txt.pack(fill=tk.X, pady=(0, 10))
            self.edits[key] = txt
        ttk.Button(main_frame, text="設定を保存", command=self.save).pack(pady=10)

    def save(self):
        new_config = {
            "user_name": self.ent_user.get(),
            "categories": [s.strip() for s in self.edits["categories"].get("1.0", tk.END).split(",") if s.strip()],
            "genres": [s.strip() for s in self.edits["genres"].get("1.0", tk.END).split(",") if s.strip()],
            "departments": [s.strip() for s in self.edits["departments"].get("1.0", tk.END).split(",") if s.strip()],
            "importance": self.config_data.get("importance", ["高", "中", "低"])
        }
        ConfigManager.save(new_config)
        self.on_save(new_config)
        self.destroy()

class HistoryWindow(tk.Toplevel):
    def __init__(self, parent, on_select):
        super().__init__(parent)
        self.title("履歴から選択")
        self.geometry("600x400")
        self.on_select = on_select
        self.setup_ui()

    def setup_ui(self):
        frame = ttk.Frame(self, padding="5")
        frame.pack(fill=tk.BOTH, expand=True)

        columns = ("no", "date", "category", "content", "genre")
        self.tree = ttk.Treeview(frame, columns=columns, show="headings")
        self.tree.heading("no", text="No")
        self.tree.heading("date", text="日付")
        self.tree.heading("category", text="分類")
        self.tree.heading("content", text="作業内容")
        self.tree.heading("genre", text="ジャンル")
        self.tree.column("no", width=30)
        self.tree.column("date", width=80)
        self.tree.column("category", width=60)
        self.tree.column("content", width=250)
        self.tree.column("genre", width=100)

        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.load_history()
        self.tree.bind("<Double-1>", self.select)
        ttk.Button(self, text="選択して閉じる", command=self.select).pack(pady=5)

    def load_history(self):
        if not os.path.exists(DEFAULT_CSV_PATH): return
        self.history_items = []
        try:
            with open(DEFAULT_CSV_PATH, 'r', encoding='utf_8_sig') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                for i, r in enumerate(rows[-100:][::-1]):
                    self.history_items.append(r)
                    self.tree.insert("", tk.END, values=(
                        r.get("No",""), r.get("日付",""), r.get("分類",""),
                        r.get("問題事象",""), r.get("ジャンル","")), tags=(str(i),))
        except Exception as e:
            print(f"History Load Error: {e}")

    def select(self, event=None):
        selection = self.tree.selection()
        if not selection: return
        item = self.tree.item(selection[0])
        idx = int(item['tags'][0])
        raw_data = self.history_items[idx]
        self.on_select(raw_data)
        self.destroy()

class AnalysisWindow(tk.Toplevel):
    def __init__(self, parent, config):
        super().__init__(parent)
        self.title("データ集計・分析")
        self.geometry("850x850")
        self.config = config
        self.data_rows = []
        self.setup_ui()

    def setup_ui(self):
        if not os.path.exists(DEFAULT_CSV_PATH):
            ttk.Label(self, text="データがありません").pack(pady=20)
            return

        try:
            with open(DEFAULT_CSV_PATH, 'r', encoding='utf_8_sig') as f:
                reader = csv.DictReader(f)
                self.data_rows = list(reader)
            if not self.data_rows: return

            # 操作パネル
            ctrl_frame = ttk.Frame(self, padding=10)
            ctrl_frame.pack(fill=tk.X)

            # 期間指定
            ttk.Label(ctrl_frame, text="集計期間:").pack(side=tk.LEFT, padx=5)

            now = datetime.datetime.now()
            first_day = now.replace(day=1).strftime("%Y/%m/%d")
            last_day = now.strftime("%Y/%m/%d")

            self.ent_start = ttk.Entry(ctrl_frame, width=12)
            self.ent_start.insert(0, first_day)
            self.ent_start.pack(side=tk.LEFT, padx=2)

            ttk.Label(ctrl_frame, text="～").pack(side=tk.LEFT)

            self.ent_end = ttk.Entry(ctrl_frame, width=12)
            self.ent_end.insert(0, last_day)
            self.ent_end.pack(side=tk.LEFT, padx=2)

            ttk.Button(ctrl_frame, text="表示更新", command=self.refresh_plots).pack(side=tk.LEFT, padx=10)
            ttk.Button(ctrl_frame, text="レポート(Excel)を生成", command=self.create_report).pack(side=tk.RIGHT, padx=5)

            # グラフ表示エリア
            self.plot_frame = ttk.Frame(self)
            self.plot_frame.pack(fill=tk.BOTH, expand=True)

            self.refresh_plots()

        except Exception as e:
            messagebox.showerror("エラー", f"分析中にエラーが発生しました: {e}")

    def get_filtered_data(self):
        start_str = self.ent_start.get().strip()
        end_str = self.ent_end.get().strip()

        def parse_date(s):
            for fmt in ("%Y/%m/%d", "%Y/%m/%j", "%Y/%n/%d", "%Y/%n/%j"):
                try: return datetime.datetime.strptime(s, fmt)
                except: continue
            try:
                parts = s.split("/")
                if len(parts) == 3:
                    return datetime.datetime(int(parts[0]), int(parts[1]), int(parts[2]))
            except: pass
            return None

        start_date = parse_date(start_str)
        end_date = parse_date(end_str)

        if not start_date or not end_date:
            messagebox.showerror("エラー", "日付形式が正しくありません (yyyy/mm/dd)")
            return []

        filtered = []
        for r in self.data_rows:
            d_str = r.get("日付", "")
            d = parse_date(d_str)
            if d and start_date <= d <= end_date:
                filtered.append(r)
        return filtered

    def refresh_plots(self):
        for widget in self.plot_frame.winfo_children():
            widget.destroy()

        data = self.get_filtered_data()
        if not data:
            ttk.Label(self.plot_frame, text="指定期間のデータがありません").pack(pady=50)
            return

        self.fig, axes = plt.subplots(2, 2, figsize=(10, 8))
        plt.subplots_adjust(wspace=0.6, hspace=0.4)

        self.plot_pie(axes[0,0], [r.get("ジャンル") for r in data], "ジャンル別比率")
        self.plot_pie(axes[0,1], [r.get("依頼部署") for r in data], "部署別比率")
        self.plot_pie(axes[1,0], [r.get("ステータス") for r in data], "完了・進捗状況")
        self.plot_pie(axes[1,1], [r.get("担当者") for r in data], "作業員別対応比率")

        canvas = FigureCanvasTkAgg(self.fig, master=self.plot_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def plot_pie(self, ax, items, title):
        from collections import Counter
        counts = Counter([i for i in items if i])
        if not counts:
            ax.text(0.5, 0.5, "データ不足", ha='center')
            ax.set_title(title)
            return
        labels = list(counts.keys())
        values = list(counts.values())
        ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90,
               pctdistance=1.2, labeldistance=1.4, textprops={'fontsize': 8})
        ax.set_title(title, pad=20)

    def create_report(self):
        data = self.get_filtered_data()
        if not data:
            messagebox.showinfo("情報", "対象期間のデータがありません。")
            return

        start_str = self.ent_start.get().replace("/", "")
        end_str = self.ent_end.get().replace("/", "")
        period_display = f"{self.ent_start.get()}～{self.ent_end.get()}"

        try:
            messagebox.showinfo("処理中", "データを集計し、Excelを作成します...")

            wb = Workbook()
            ws = wb.active
            ws.title = "作業レポート"
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4

            ws["A1"] = f"業務報告書 ({period_display})"
            ws["A1"].font = Font(size=14, bold=True)
            ws.merge_cells("A1:H1")
            ws["A1"].alignment = Alignment(horizontal="center")

            # グラフの挿入
            chart_path = "temp_report_chart.png"
            self.fig.savefig(chart_path, dpi=100)
            img = XLImage(chart_path)
            img.width = 550
            img.height = 400
            ws.add_image(img, "A3")

            # 集計データの記載
            ws["A25"] = "【項目別集計】"
            ws["A25"].font = Font(bold=True)
            ws["A26"] = "カテゴリ"; ws["B26"] = "件数"

            from collections import Counter
            genres = Counter([r.get("ジャンル") for r in data if r.get("ジャンル")])
            row = 27
            for g, count in genres.items():
                ws.cell(row=row, column=1, value=g)
                ws.cell(row=row, column=2, value=count)
                row += 1

            filename = f"Report_{start_str}_{end_str}.xlsx"
            wb.save(filename)
            if os.path.exists(chart_path): os.remove(chart_path)

            messagebox.showinfo("完了", f"レポートを保存しました: {filename}")
            os.startfile(filename)
        except Exception as e:
            messagebox.showerror("エラー", f"レポート作成中にエラーが発生しました: {e}")

class WorkLogApp:
    def __init__(self, root):
        self.root = root
        self.root.title("作業記録 Pro")
        self.root.attributes('-topmost', True)
        self.root.geometry("440x480")
        self.root.configure(bg="#f0f2f5")

        self.config = ConfigManager.load()
        self.start_dt = None

        self.apply_style()
        self.setup_ui()
        self.reset_inputs()
        self.bind_events()

    def apply_style(self):
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure("TFrame", background="#f0f2f5")
        self.style.configure("TLabel", background="#f0f2f5", font=("Segoe UI", 8))
        self.style.configure("Header.TLabel", font=("Segoe UI", 8, "bold"))
        self.style.configure("TButton", font=("Segoe UI", 8))
        self.style.configure("Accent.TButton", foreground="white", background="#007bff")
        self.style.configure("Summary.TButton", foreground="white", background="#6f42c1")
        self.style.map("Accent.TButton", background=[('active', '#0056b3')])
        self.style.map("Summary.TButton", background=[('active', '#5a32a3')])

    def setup_ui(self):
        self.main_frame = ttk.Frame(self.root, padding="3")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Top Bar
        top_bar = ttk.Frame(self.main_frame)
        top_bar.pack(fill=tk.X, pady=(0, 2))
        ttk.Button(top_bar, text="⚙", width=2, command=self.open_settings).pack(side=tk.LEFT)
        self.lbl_user_info = ttk.Label(top_bar, text=f"担当:{self.config.get('user_name', '')}", style="Header.TLabel")
        self.lbl_user_info.pack(side=tk.LEFT, padx=3)
        ttk.Button(top_bar, text="CSV", width=4, command=self.open_csv).pack(side=tk.LEFT, padx=1)
        ttk.Button(top_bar, text="📊集計・分析(レポート)", width=18, command=self.open_analysis, style="Summary.TButton").pack(side=tk.LEFT, padx=1)

        # Info: 分類
        info_bar = ttk.Frame(self.main_frame)
        info_bar.pack(fill=tk.X, pady=(0, 2))
        ttk.Label(info_bar, text="分類:").pack(side=tk.LEFT)
        self.combo_category = ttk.Combobox(info_bar, values=self.config.get("categories", []), width=10)
        self.combo_category.pack(side=tk.LEFT, padx=2)

        # Time Bar
        time_frame = tk.LabelFrame(self.main_frame, bg="#f0f2f5", fg="#555", font=("Segoe UI", 7, "bold"), padx=2, pady=0)
        time_frame.pack(fill=tk.X, pady=(0, 2))
        self.btn_start = ttk.Button(time_frame, text="開始", width=5, command=self.record_start)
        self.btn_start.pack(side=tk.LEFT, padx=2)
        self.lbl_times = ttk.Label(time_frame, text="--:-- ～ --:-- (0:00)")
        self.lbl_times.pack(side=tk.LEFT, padx=5)
        ttk.Button(time_frame, text="👔履歴", width=7, command=self.open_history).pack(side=tk.RIGHT, padx=2)

        # Attributes Grid
        attr_frame = ttk.Frame(self.main_frame)
        attr_frame.pack(fill=tk.X)
        self.combos = {}
        fields = [
            ("genre", "ジャンル:", self.config.get("genres", []), 0, 0),
            ("dept", "依頼部署:", self.config.get("departments", []), 0, 2),
            ("imp", "重要度:", self.config.get("importance", []), 1, 0)
        ]
        for key, label, vals, r, c in fields:
            ttk.Label(attr_frame, text=label).grid(row=r, column=c, sticky=tk.W)
            cb = ttk.Combobox(attr_frame, values=vals, width=14)
            cb.grid(row=r, column=c+1, sticky=tk.W, padx=(0, 5), pady=1)
            self.combos[key] = cb

        ttk.Label(attr_frame, text="依頼主:").grid(row=1, column=2, sticky=tk.W)
        self.ent_client = ttk.Entry(attr_frame, width=17)
        self.ent_client.grid(row=1, column=3, sticky=tk.W, pady=1)

        ttk.Label(attr_frame, text="完了予定:").grid(row=2, column=0, sticky=tk.W)
        self.ent_target_date = ttk.Entry(attr_frame, width=15)
        self.ent_target_date.grid(row=2, column=1, sticky=tk.W, pady=1)
        today = datetime.date.today().strftime("%Y/%m/%d")
        self.ent_target_date.insert(0, today)

        # Text Area
        ttk.Label(self.main_frame, text="作業内容:", style="Header.TLabel").pack(anchor=tk.W)
        self.txt_problem = tk.Text(self.main_frame, height=4, font=("Segoe UI", 9), bd=1)
        self.txt_problem.pack(fill=tk.X, pady=(0, 2))
        ttk.Label(self.main_frame, text="対応内容:", style="Header.TLabel").pack(anchor=tk.W)
        self.txt_solution = tk.Text(self.main_frame, height=4, font=("Segoe UI", 9), bd=1)
        self.txt_solution.pack(fill=tk.X, pady=(0, 2))

        # Footer Buttons
        btn_frame = ttk.Frame(self.main_frame)
        btn_frame.pack(fill=tk.X, pady=(2, 0))

        ttk.Button(btn_frame, text="入力クリア", command=self.reset_inputs).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2))
        self.btn_save = ttk.Button(btn_frame, text="完了・保存 (Ctrl+Enter)", command=self.save_log, style="Accent.TButton")
        self.btn_save.pack(side=tk.LEFT, fill=tk.X, expand=True)

    def bind_events(self):
        self.root.bind('<Control-Return>', lambda e: self.save_log())
        self.root.bind('<Control-s>', lambda e: self.temp_save())
        self.root.bind('<Configure>', self.check_snap)

    def open_csv(self):
        if os.path.exists(DEFAULT_CSV_PATH): os.startfile(DEFAULT_CSV_PATH)
        else: messagebox.showinfo("情報", "CSVがありません")

    def open_history(self):
        HistoryWindow(self.root, self.on_history_select)

    def on_history_select(self, d):
        self.combo_category.set(d.get("分類",""))
        self.combos["genre"].set(d.get("ジャンル",""))
        self.combos["dept"].set(d.get("依頼部署",""))
        self.combos["imp"].set(d.get("重要度",""))
        self.ent_client.delete(0, tk.END)
        self.ent_client.insert(0, d.get("依頼主",""))
        self.txt_problem.delete("1.0", tk.END)
        self.txt_problem.insert("1.0", d.get("問題事象",""))
        self.txt_solution.delete("1.0", tk.END)
        self.txt_solution.insert("1.0", d.get("対応内容",""))

    def open_analysis(self):
        AnalysisWindow(self.root, self.config)

    def check_snap(self, event=None):
        if not hasattr(self, '_snap_active'): self._snap_active = False
        if self._snap_active: return
        SNAP_DIST = 30
        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        x, y = self.root.winfo_x(), self.root.winfo_y()
        w, h = self.root.winfo_width(), self.root.winfo_height()
        nx, ny = x, y
        if abs(x) < SNAP_DIST: nx = 0
        elif abs(x + w - sw) < SNAP_DIST: nx = sw - w
        if abs(y) < SNAP_DIST: ny = 0
        elif abs(y + h - sh) < SNAP_DIST: ny = sh - h
        if nx != x or ny != y:
            self._snap_active = True
            self.root.geometry(f"+{nx}+{ny}")
            self.root.after(100, lambda: setattr(self, '_snap_active', False))

    def open_settings(self):
        SettingsWindow(self.root, self.config, self.update_config)

    def update_config(self, new_config):
        self.config = new_config
        self.lbl_user_info.config(text=f"担当:{self.config.get('user_name', '')}")
        self.combo_category.config(values=self.config.get("categories", []))
        self.combos["genre"].config(values=self.config.get("genres", []))
        self.combos["dept"].config(values=self.config.get("departments", []))

    def record_start(self):
        self.start_dt = datetime.datetime.now()
        self.lbl_times.config(text=f"{self.start_dt.strftime('%H:%M')} ～ 進行中...")
        self.btn_start.configure(state="disabled")

    def temp_save(self): messagebox.showinfo("一時保存", "内容を保持しました")

    def save_log(self):
        if not self.start_dt: return
        end_dt = datetime.datetime.now()
        duration = end_dt - self.start_dt

        next_no = 1
        if os.path.exists(DEFAULT_CSV_PATH):
            try:
                with open(DEFAULT_CSV_PATH, 'r', encoding='utf_8_sig') as f:
                    reader = csv.reader(f)
                    lines = list(reader)
                    if len(lines) > 1:
                        last_no = lines[-1][0]
                        if last_no.isdigit(): next_no = int(last_no) + 1
            except: pass

        category = self.combo_category.get()
        genre = self.combos["genre"].get()
        dept = self.combos["dept"].get()
        imp = self.combos["imp"].get()
        client = self.ent_client.get()
        target_date = self.ent_target_date.get()
        problem = self.txt_problem.get("1.0", tk.END).strip()
        solution = self.txt_solution.get("1.0", tk.END).strip()

        data = [
            next_no, category, self.config["user_name"],
            self.start_dt.strftime("%Y/%m/%d"), self.start_dt.strftime("%H:%M:%S"),
            end_dt.strftime("%H:%M:%S"), f"{int(duration.total_seconds() // 60)}分",
            target_date, end_dt.strftime("%Y/%m/%d"),
            genre, dept, client, imp, "完了", problem, solution
        ]

        try:
            file_exists = os.path.isfile(DEFAULT_CSV_PATH)
            with open(DEFAULT_CSV_PATH, 'a', newline='', encoding='utf_8_sig') as f:
                writer = csv.writer(f)
                if not file_exists:
                    writer.writerow(["No","分類","担当者","日付","発生時刻","終了時刻","実作業時間","終了予定日","完了日付","ジャンル","依頼部署","依頼主","重要度","ステータス","問題事象","対応内容"])
                writer.writerow(data)
            messagebox.showinfo("成功", f"記録しました(No.{next_no})")
            self.reset_inputs()
        except PermissionError: messagebox.showerror("エラー", "CSVを閉じてください")

    def reset_inputs(self):
        self.start_dt = None
        self.btn_start.configure(state="normal")
        self.lbl_times.config(text="--:-- ～ --:-- (0:00)")
        self.combo_category.set('')
        for cb in self.combos.values(): cb.set('')
        self.ent_client.delete(0, tk.END)
        self.txt_problem.delete("1.0", tk.END)
        self.txt_solution.delete("1.0", tk.END)

if __name__ == "__main__":
    root = tk.Tk()
    app = WorkLogApp(root)
    root.mainloop()
