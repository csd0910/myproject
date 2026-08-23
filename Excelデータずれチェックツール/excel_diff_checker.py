import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from datetime import datetime
import threading

TARGET_HEADERS = [
    "品目cd(em310)",
    "注文番号(em310)",
    "カタチラ中分類cd(em311)",
    "カタチラ小分類cd(em311)",
    "カタチラ細分類cd(em311)",
    "品目小分類cd(em163)",
    "品目細分類cd(em163)"
]

def normalize_val(val):
    if val is None:
        return ""
    s = str(val)
    if s.endswith(".0"):
        s = s[:-2]
    return s.strip().replace('\n', '').replace('\r', '')

def get_unique_code(sheet, row_idx, col_indices):
    values = []
    for col_idx in col_indices:
        if col_idx is not None:
            val = sheet.cell(row=row_idx, column=col_idx).value
            values.append(normalize_val(val))
        else:
            values.append("")
    return "_".join(values)

def get_header_indices(sheet):
    max_col = sheet.max_column
    headers = [str(sheet.cell(row=1, column=i).value).strip() for i in range(1, max_col + 1)]
    
    indices = []
    for target in TARGET_HEADERS:
        try:
            idx = headers.index(target) + 1
            indices.append(idx)
        except ValueError:
            indices.append(None)
    return indices

def process_file(filepath, sheet_name):
    wb = openpyxl.load_workbook(filepath)
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active
        
    max_col = sheet.max_column
    max_row = sheet.max_row
    
    target_col = max_col + 1
    col_indices = get_header_indices(sheet)
    
    sheet.cell(row=1, column=target_col).value = "UniqueCode"
    
    code_dict = {}
    for row_idx in range(2, max_row + 1):
        code = get_unique_code(sheet, row_idx, col_indices)
        sheet.cell(row=row_idx, column=target_col).value = code
        code_dict[code] = row_idx
        
    wb.save(filepath)
    wb.close()
    
    return code_dict

def run_processing(old_file, old_sheet, new_file, new_sheet, app):
    # exeの実行フォルダを取得
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    
    try:
        app.update_status("作業前ファイルの処理中...")
        old_codes = process_file(old_file, old_sheet)
        
        app.update_status("作業後ファイルの処理中...")
        new_codes = process_file(new_file, new_sheet)
        
        app.update_status("差分を比較中...")
        old_code_set = set(old_codes.keys())
        diff_codes = {}
        
        for code, row_idx in new_codes.items():
            if code not in old_code_set:
                diff_codes[code] = row_idx
                
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if not diff_codes:
            result_file = os.path.join(base_dir, f"判定結果_差分なし_{timestamp}.txt")
            with open(result_file, "w", encoding="utf-8") as f:
                f.write("判定結果：差分なし\n")
                f.write(f"作業前ファイル: {os.path.basename(old_file)} (シート: {old_sheet})\n")
                f.write(f"作業後ファイル: {os.path.basename(new_file)} (シート: {new_sheet})\n")
                f.write("-" * 50 + "\n")
                f.write("作業後のファイル内に、予期せぬ新しいコード（行のズレ等）は見つかりませんでした。\n")
            app.root.after(0, lambda: messagebox.showinfo("完了", f"差分なし。結果を以下に出力しました。\n{result_file}"))
        else:
            result_file = os.path.join(base_dir, f"判定結果_差分あり_{timestamp}.txt")
            with open(result_file, "w", encoding="utf-8") as f:
                f.write("判定結果：差分あり（ズレや新規発生の疑い）\n")
                f.write(f"作業前ファイル: {old_file} (シート: {old_sheet})\n")
                f.write(f"作業後ファイル: {new_file} (シート: {new_sheet})\n")
                f.write("-" * 50 + "\n")
                f.write("以下のコードは古いファイルに存在せず、新しいファイルで発生しています。\n")
                for code, row_idx in diff_codes.items():
                    f.write(f"新しいファイルの {row_idx} 行目: コード[{code}]\n")
            app.root.after(0, lambda: messagebox.showwarning("差分あり", f"ズレや新規データが {len(diff_codes)} 件見つかりました。\n結果を以下に出力しました。\n{result_file}"))
            
    except Exception as e:
        import traceback
        err_msg = traceback.format_exc()
        app.root.after(0, lambda e=e: messagebox.showerror("実行エラー", f"処理中にエラーが発生しました:\n{e}"))
    finally:
        app.root.after(0, app.finish_processing)

class DiffCheckerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excelデータずれチェックツール")
        self.root.geometry("680x380")
        self.root.eval('tk::PlaceWindow . center')
        
        style = ttk.Style()
        if 'clam' in style.theme_names():
            style.theme_use('clam')
            
        self.old_file_path = tk.StringVar()
        self.new_file_path = tk.StringVar()
        self.old_sheet_name = tk.StringVar()
        self.new_sheet_name = tk.StringVar()
        
        self.create_widgets()
        
    def create_widgets(self):
        # 説明ラベル
        ttk.Label(self.root, text="比較したい2つのExcelファイル（作業前・作業後）を選択してください。").pack(pady=(15, 10))
        
        # --- 作業前エリア ---
        frame_old = ttk.Frame(self.root)
        frame_old.pack(fill=tk.X, padx=15, pady=(5, 15))
        
        # ファイル選択行
        row_old_file = ttk.Frame(frame_old)
        row_old_file.pack(fill=tk.X, pady=2)
        ttk.Label(row_old_file, text="作業前ファイル:", width=15).pack(side=tk.LEFT)
        ttk.Entry(row_old_file, textvariable=self.old_file_path, width=58).pack(side=tk.LEFT, padx=5)
        ttk.Button(row_old_file, text="参照...", command=lambda: self.select_file(self.old_file_path, self.combo_old_sheet, self.old_sheet_name)).pack(side=tk.LEFT)
        
        # シート選択行
        row_old_sheet = ttk.Frame(frame_old)
        row_old_sheet.pack(fill=tk.X, pady=(5, 5))
        ttk.Label(row_old_sheet, text="対象シート:", width=15).pack(side=tk.LEFT)
        self.combo_old_sheet = ttk.Combobox(row_old_sheet, textvariable=self.old_sheet_name, state="readonly", width=30)
        self.combo_old_sheet.pack(side=tk.LEFT, padx=5)
        
        # セパレータ（区切り線）
        ttk.Separator(self.root, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=20, pady=5)
        
        # --- 作業後エリア ---
        frame_new = ttk.Frame(self.root)
        frame_new.pack(fill=tk.X, padx=15, pady=(15, 5))
        
        # ファイル選択行
        row_new_file = ttk.Frame(frame_new)
        row_new_file.pack(fill=tk.X, pady=2)
        ttk.Label(row_new_file, text="作業後ファイル:", width=15).pack(side=tk.LEFT)
        ttk.Entry(row_new_file, textvariable=self.new_file_path, width=58).pack(side=tk.LEFT, padx=5)
        ttk.Button(row_new_file, text="参照...", command=lambda: self.select_file(self.new_file_path, self.combo_new_sheet, self.new_sheet_name)).pack(side=tk.LEFT)
        
        # シート選択行
        row_new_sheet = ttk.Frame(frame_new)
        row_new_sheet.pack(fill=tk.X, pady=(5, 5))
        ttk.Label(row_new_sheet, text="対象シート:", width=15).pack(side=tk.LEFT)
        self.combo_new_sheet = ttk.Combobox(row_new_sheet, textvariable=self.new_sheet_name, state="readonly", width=30)
        self.combo_new_sheet.pack(side=tk.LEFT, padx=5)
        
        # Run Button
        self.btn_run = ttk.Button(self.root, text="チェック実行", command=self.start_processing)
        self.btn_run.pack(pady=(20, 5))
        
        # Status Label
        self.lbl_status = ttk.Label(self.root, text="", foreground="blue")
        self.lbl_status.pack()

    def select_file(self, var_path, combobox, var_sheet):
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            var_path.set(filepath)
            # シート名を取得してプルダウンを更新
            self.update_sheet_names(filepath, combobox, var_sheet)
            
    def update_sheet_names(self, filepath, combobox, var_sheet):
        self.update_status("シート名を読み込んでいます...")
        self.root.update()
        try:
            # 高速化のため read_only で読み込む
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheetnames = wb.sheetnames
            wb.close()
            
            combobox['values'] = sheetnames
            if sheetnames:
                var_sheet.set(sheetnames[0])
            self.update_status("シートの読み込みが完了しました。")
        except Exception as e:
            combobox['values'] = []
            var_sheet.set("")
            messagebox.showerror("読み込みエラー", f"シート名の取得に失敗しました:\n{e}")
            self.update_status("")
            
    def update_status(self, text):
        self.lbl_status.config(text=text)
        
    def start_processing(self):
        old_f = self.old_file_path.get()
        new_f = self.new_file_path.get()
        old_s = self.old_sheet_name.get()
        new_s = self.new_sheet_name.get()
        
        if not old_f or not new_f:
            messagebox.showerror("エラー", "作業前と作業後の両方のファイルを選択してください。")
            return
        if not os.path.exists(old_f) or not os.path.exists(new_f):
            messagebox.showerror("エラー", "指定されたファイルが存在しません。")
            return
        if not old_s or not new_s:
            messagebox.showerror("エラー", "比較するシートを選択してください。")
            return
            
        self.btn_run.config(state=tk.DISABLED)
        self.update_status("処理を開始します...")
        
        thread = threading.Thread(target=run_processing, args=(old_f, old_s, new_f, new_s, self))
        thread.start()
        
    def finish_processing(self):
        self.btn_run.config(state=tk.NORMAL)
        self.update_status("処理完了")

if __name__ == "__main__":
    root = tk.Tk()
    app = DiffCheckerApp(root)
    root.mainloop()
