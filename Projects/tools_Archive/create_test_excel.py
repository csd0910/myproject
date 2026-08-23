import openpyxl
import random
from datetime import datetime, timedelta

wb = openpyxl.Workbook()

# ==========================================
# 1. 商品マスタ（VLOOKUP参照用）
# ==========================================
ws_master = wb.active
ws_master.title = "商品マスタ"
ws_master.append(["商品コード", "商品カテゴリ", "商品名", "単価"])

products = [
    ("A001", "PC", "ノートPC_Standard", 120000),
    ("A002", "PC", "ノートPC_Pro", 180000),
    ("A003", "PC", "デスクトップPC", 150000),
    ("B001", "周辺機器", "モニター24インチ", 25000),
    ("B002", "周辺機器", "モニター27インチ", 35000),
    ("B003", "周辺機器", "ワイヤレスマウス", 3000),
    ("B004", "周辺機器", "メカニカルキーボード", 12000),
    ("C001", "ソフトウェア", "Officeライセンス", 20000),
    ("C002", "ソフトウェア", "セキュリティソフト", 5000),
    ("D001", "サプライ", "コピー用紙A4", 500),
]
for p in products:
    ws_master.append(p)

# ==========================================
# 2. 売上データ（大容量：VLOOKUPテスト用）
# ==========================================
ws_sales = wb.create_sheet(title="売上データ_作業用")
ws_sales.append(["日付", "担当者", "商品コード", "販売数", "商品カテゴリ(ここをVLOOKUP)", "商品名(ここをVLOOKUP)", "単価(ここをVLOOKUP)", "売上金額(単価×販売数)"])

staff = ["山田", "佐藤", "鈴木", "田中", "高橋", "伊藤"]
start_date = datetime(2026, 7, 1)

# 500行のダミーデータを生成
for i in range(500):
    d = start_date + timedelta(days=random.randint(0, 30))
    s = random.choice(staff)
    p = random.choice(products)[0] # 商品コード
    qty = random.randint(1, 10)
    ws_sales.append([d.strftime("%Y/%m/%d"), s, p, qty])

# ==========================================
# 3. 担当者別集計表（SUMIFSテスト用）
# ==========================================
ws_summary = wb.create_sheet(title="担当者別集計表")
ws_summary.append(["担当者", "総販売数(SUMIFS)", "総売上金額(SUMIFS)"])
for s in staff:
    ws_summary.append([s])
    
ws_summary.append([])
ws_summary.append(["※ここに『売上データ_作業用』からコピペしてくるテスト用スペース"])

# ==========================================
# 4. 別システム抽出データ（シート間コピペテスト用）
# ==========================================
ws_raw = wb.create_sheet(title="外部システム抽出RAW")
ws_raw.append(["抽出日: 2026/07/27", "システム: WebSalesDB"])
ws_raw.append(["Date", "StaffName", "ItemCode", "Quantity"])
for i in range(50):
    d = start_date + timedelta(days=random.randint(0, 5))
    s = random.choice(staff)
    p = random.choice(products)[0]
    qty = random.randint(1, 5)
    ws_raw.append([d.strftime("%Y/%m/%d"), s, p, qty])

wb.save("DXテスト用_大容量Excel.xlsx")
print("DXテスト用_大容量Excel.xlsx を作成しました。")
