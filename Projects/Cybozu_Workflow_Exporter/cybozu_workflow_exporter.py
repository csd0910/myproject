import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import sqlite3
import csv
import os
import re
import threading
import time
import urllib.request
from datetime import datetime

import sys
import subprocess

try:
    from selenium import webdriver
    from selenium.webdriver.edge.options import Options
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
except ImportError:
    pass

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

class CybozuWorkflowExporterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Cybozu Workflow Exporter V2.0")
        self.root.geometry("600x500")
        self.driver = None
        self.is_extracting = False
        self.create_widgets()
        
    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        config_frame = ttk.LabelFrame(main_frame, text="設定", padding=5)
        config_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(config_frame, text="保存先(Gドライブ):").grid(row=0, column=0, sticky=tk.W)
        self.save_dir_var = tk.StringVar(value=r"G:\マイドライブ\サイボウズワークフローバックアップ")
        ttk.Entry(config_frame, textvariable=self.save_dir_var, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(config_frame, text="参照", command=self.browse_dir).grid(row=0, column=2)
        
        ttk.Label(config_frame, text="URL:").grid(row=1, column=0, sticky=tk.W)
        self.url_var = tk.StringVar(value="https://forestway.cybozu.com/login")
        ttk.Entry(config_frame, textvariable=self.url_var, width=50).grid(row=1, column=1, padx=5)
        
        btn_box = ttk.Frame(main_frame)
        btn_box.pack(pady=10)
        
        self.btn_db = ttk.Button(btn_box, text="ステップ①：初期化テスト", command=self.init_system)
        self.btn_db.pack(side=tk.LEFT, padx=5)
        
        self.btn_extract = ttk.Button(btn_box, text="ステップ②：ブラウザ起動＆抽出開始", command=self.start_extraction)
        self.btn_extract.pack(side=tk.LEFT, padx=5)
        
        self.btn_delete = ttk.Button(btn_box, text="ステップ③：不要データ一括削除", command=self.delete_unneeded)
        self.btn_delete.pack(side=tk.LEFT, padx=5)

        self.btn_stop = ttk.Button(btn_box, text="中断", command=self.stop_extraction, state=tk.DISABLED)
        self.btn_stop.pack(side=tk.LEFT, padx=5)
        
        # 進捗バーと残り時間表示
        prog_frame = ttk.Frame(main_frame, padding=5)
        prog_frame.pack(fill=tk.X)
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(prog_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=2)
        self.time_info_var = tk.StringVar(value="進捗: 待機中")
        ttk.Label(prog_frame, textvariable=self.time_info_var, font=("Meiryo", 9)).pack(side=tk.TOP, anchor=tk.W)

        self.log_text = tk.Text(main_frame, height=14, width=70, font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True, pady=5)

    def browse_dir(self):
        d = filedialog.askdirectory()
        if d: self.save_dir_var.set(os.path.normpath(d))

    def log(self, msg):
        self.log_text.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_text.see(tk.END)
        self.root.update()

    def init_system(self):
        save_dir = self.save_dir_var.get()
        if not os.path.exists(save_dir):
            try:
                os.makedirs(save_dir, exist_ok=True)
            except Exception as e:
                self.log(f"エラー: フォルダ作成失敗。{e}")
                return

        db_path = os.path.join(save_dir, "workflow_manager.db")
        try:
            conn = sqlite3.connect(db_path)
            c = conn.cursor()
            c.execute('''CREATE TABLE IF NOT EXISTS processed_workflows (
                            id INTEGER PRIMARY KEY, category TEXT, form_name TEXT, title TEXT, status TEXT, downloaded_at TEXT)''')
            conn.commit()
            conn.close()
        except Exception as e:
            self.log(f"エラー: DB作成失敗。{e}")
            return

        excel_path = os.path.join(save_dir, "ワークフロー台帳.xlsx")
        if not os.path.exists(excel_path):
            try:
                wb = openpyxl.Workbook()
                ws_sent = wb.active
                ws_sent.title = "送信一覧"
                ws_sent.append(["削除フラグ", "番号", "申請日", "カテゴリ", "状況", "申請者/処理者", "件名", "添付ファイル", "フォルダへのリンク"])
                
                ws_recept = wb.create_sheet(title="受信一覧")
                ws_recept.append(["削除フラグ", "番号", "申請日", "カテゴリ", "状況", "申請者/処理者", "件名", "添付ファイル", "フォルダへのリンク"])
                
                wb.save(excel_path)
            except Exception as e:
                self.log(f"エラー: Excel作成失敗。{e}")
                return

        self.log(">>> [ステップ① 完了] 初期化成功！ステップ②へ進んでください。")

    def stop_extraction(self):
        self.is_extracting = False
        self.log("中断シグナルを送信しました。現在の処理が終わるまでお待ちください。")

    def delete_unneeded(self):
        save_dir = self.save_dir_var.get()
        excel_path = os.path.join(save_dir, "ワークフロー台帳.xlsx")
        db_path = os.path.join(save_dir, "workflow_manager.db")
        if not os.path.exists(excel_path):
            self.log("エラー: ワークフロー台帳.xlsx が見つかりません。")
            return
            
        import shutil
        import openpyxl
        
        self.log("\n--- ステップ③：不要データの一括削除処理を開始します ---")
        self.btn_db.config(state=tk.DISABLED)
        self.btn_extract.config(state=tk.DISABLED)
        self.btn_delete.config(state=tk.DISABLED)
        
        def run_delete():
            try:
                wb = openpyxl.load_workbook(excel_path)
                conn = sqlite3.connect(db_path)
                c = conn.cursor()
                
                delete_count = 0
                target_flags = ["〇", "○", "削除", "x", "X", "×"]
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # 下から上へループ（行削除時のズレ防止）
                    for row_idx in range(ws.max_row, 1, -1):
                        flag_cell = ws.cell(row=row_idx, column=1).value
                        if flag_cell and str(flag_cell).strip() in target_flags:
                            item_id = ws.cell(row=row_idx, column=2).value
                            folder_link = ws.cell(row=row_idx, column=9).value
                            
                            # 1. フォルダの完全削除
                            if folder_link and os.path.exists(folder_link):
                                try:
                                    shutil.rmtree(folder_link)
                                    self.log(f"  -> フォルダ削除完了: [No.{item_id}]")
                                except Exception as e:
                                    self.log(f"  -> フォルダ削除エラー [No.{item_id}]: {e}")
                            
                            # 2. DBから削除
                            if item_id:
                                c.execute("DELETE FROM processed_workflows WHERE id = ?", (item_id,))
                            
                            # 3. Excelから行を削除
                            ws.delete_rows(row_idx)
                            delete_count += 1
                            
                if delete_count > 0:
                    conn.commit()
                    wb.save(excel_path)
                    self.log(f"\n>>> [完了] 計 {delete_count} 件の不要データ（バックアップフォルダ・台帳記録）を完全に削除しました！")
                else:
                    self.log("\n>>> [完了] 削除フラグ（〇や削除など）が設定されているデータは見つかりませんでした。")
                    
            except Exception as e:
                self.log(f"!! 削除中にエラーが発生しました: {e}")
                self.log("※台帳ファイル(Excel)が開いたままになっている場合は、一度閉じてから再度実行してください。")
            finally:
                if 'conn' in locals(): conn.close()
                self.btn_db.config(state=tk.NORMAL)
                self.btn_extract.config(state=tk.NORMAL)
                self.btn_delete.config(state=tk.NORMAL)
                
        threading.Thread(target=run_delete, daemon=True).start()

    def start_extraction(self):
        save_dir = self.save_dir_var.get()
        excel_path = os.path.join(save_dir, "ワークフロー台帳.xlsx")
        if not os.path.exists(excel_path):
            self.log("エラー: 先に「ステップ①」を実行して台帳を作成してください。")
            return

        self.is_extracting = True
        self.btn_extract.config(state=tk.DISABLED)
        self.btn_stop.config(state=tk.NORMAL)
        threading.Thread(target=self._extraction_loop, args=(save_dir,), daemon=True).start()

    def _extraction_loop(self, save_dir):
        app_start_time = time.time()
        start_datetime_str = datetime.now().strftime('%Y/%m/%d %H:%M:%S')
        try:
            self.log("Edgeブラウザを起動します...")
            options = Options()
            options.add_experimental_option('excludeSwitches', ['enable-logging'])
            self.driver = webdriver.Edge(options=options)
            self.driver.get(self.url_var.get())
            
            self.log("サイボウズにログインしてください。（ログイン後、自動で抽出が始まります）")
            WebDriverWait(self.driver, 300).until(EC.url_contains("ag.cgi"))
            self.log("ログインを確認しました！抽出処理を開始します。")
            time.sleep(2)
            
            base_url = self.driver.current_url.split('?')[0]
            categories = [
                {"name": "送信一覧", "url_param": "page=WorkFlowSent"},
                {"name": "受信一覧", "url_param": "page=WorkFlowRecept"}
            ]
            
            db_path = os.path.join(save_dir, "workflow_manager.db")
            
            total_new_items = 0
            all_target_items = []
            
            for cat in categories:
                if not self.is_extracting: break
                self.log(f"\n--- 【{cat['name']}】 のリスト解析を開始 ---")
                
                target_list_url = f"{base_url}?{cat['url_param']}"
                self.driver.get(target_list_url)
                time.sleep(2)
                
                list_data = []
                page_count = 1
                seen_ids = set()
                
                while self.is_extracting:
                    self.log(f"  > {page_count}ページ目を解析中...")
                    
                    js_script = """
                    // サイボウズ特有のデータテーブル（.dataListSummary, .vt など）に絞ることで余分な行（52件など）を排除
                    var rows = document.querySelectorAll("table.dataListSummary tbody tr, table.vt tbody tr, table.dataList tbody tr");
                    if (rows.length === 0) rows = document.querySelectorAll("table tbody tr"); // フォールバック
                    var results = [];
                    Array.from(rows).forEach(tr => {
                        var tds = tr.querySelectorAll("td");
                        if(tds.length >= 5) {
                            var aTag = tds[1].querySelector("a");
                            if(!aTag) return;
                            var fullText = tds[1].innerText;
                            var parts = fullText.split("\\n");
                            var folderTitle = fullText.replace(/\\n/g, '_').replace(/[()（）]/g, "").trim();
                            results.push({
                                id: tds[0].innerText.trim(),
                                form: parts[0] || "",
                                title: (parts[1] || parts[0] || "").replace(/[()（）]/g, "").trim(),
                                folder_title: folderTitle,
                                status: tds[2].innerText.trim(),
                                person: tds[3].innerText.trim(),
                                date: tds[4].innerText.trim(),
                                link: aTag.getAttribute("href")
                            });
                        }
                    });
                    
                    // 次のページへのリンク要素を直接返す
                    var nextLinkElement = null;
                    var links = document.querySelectorAll("a");
                    for(var i=0; i<links.length; i++) {
                        if(links[i].innerText.includes("次の") || links[i].innerText.includes("次へ")) {
                            nextLinkElement = links[i];
                            break;
                        }
                    }
                    return { items: results, nextLink: nextLinkElement };
                    """
                    page_result = self.driver.execute_script(js_script)
                    
                    if not page_result or len(page_result['items']) == 0:
                        break # データが0件なら確実に終了
                        
                    # 重複チェック（無限ループ回避の最強の安全網）
                    current_new_items = 0
                    for item in page_result['items']:
                        if item['id'] not in seen_ids:
                            seen_ids.add(item['id'])
                            list_data.append(item)
                            current_new_items += 1
                            
                    # このページで新しいアイテムが1件も無ければ、同じページに留まっているので終了
                    if current_new_items == 0:
                        self.log("  > 新しいデータが見つからないため、リスト取得を終了します。")
                        break
                        
                    # 1ページは最大50件。50件未満なら確実に最終ページ。
                    if len(page_result['items']) < 50:
                        break
                        
                    # 次のページがあればクリックして遷移する（Email Exporterと同じ手法）
                    if page_result['nextLink']:
                        self.driver.execute_script("arguments[0].click();", page_result['nextLink'])
                        time.sleep(2) # 遷移待ち
                        page_count += 1
                    else:
                        break
                
                self.log(f"計 {len(list_data)} 件のデータをリストから検知しました。ダウンロード対象を選別します...")
                
                cat_targets = 0
                for item in list_data:
                    # 決裁、完了、却下、取り消し を対象とする
                    if item['status'] not in ["決裁", "完了", "却下", "取り消し"]:
                        continue
                    conn = sqlite3.connect(db_path)
                    c = conn.cursor()
                    c.execute("SELECT status FROM processed_workflows WHERE id = ?", (item['id'],))
                    row = c.fetchone()
                    conn.close()
                    if row and row[0] == item['status']:
                        continue
                    item['category_name'] = cat['name']
                    all_target_items.append(item)
                    cat_targets += 1
                self.log(f"  -> 【{cat['name']}】からの新規データは {cat_targets} 件です。")
                
            # --- ループ終了（両カテゴリの収集完了） ---
            
            total_targets = len(all_target_items)
            if total_targets == 0:
                self.log("\n>>> [完了] 未処理の新規データはありませんでした！")
                self.progress_var.set(100)
                self.time_info_var.set("進捗: すべて完了しました")
                self.btn_extract.config(state=tk.NORMAL)
                self.btn_stop.config(state=tk.DISABLED)
                return
                
            self.log(f"\n--- 送信・受信合わせて全 {total_targets} 件のダウンロードを開始します ---")
            self.progress_var.set(0)
            start_time = time.time()
            
            excel_path = os.path.join(save_dir, "ワークフロー台帳.xlsx")
            if os.path.exists(excel_path):
                wb = openpyxl.load_workbook(excel_path)
            else:
                wb = openpyxl.Workbook()
                ws_sent = wb.active
                ws_sent.title = "送信一覧"
                ws_sent.append(["削除フラグ", "番号", "申請日", "カテゴリ", "状況", "申請者/処理者", "件名", "添付ファイル", "フォルダへのリンク"])
                ws_recept = wb.create_sheet(title="受信一覧")
                ws_recept.append(["削除フラグ", "番号", "申請日", "カテゴリ", "状況", "申請者/処理者", "件名", "添付ファイル", "フォルダへのリンク"])
                
            if True:
                for i, item in enumerate(all_target_items):
                    cat_name = item['category_name']
                    if cat_name not in wb.sheetnames:
                        ws = wb.create_sheet(title=cat_name)
                        ws.append(["削除フラグ", "番号", "申請日", "カテゴリ", "状況", "申請者/処理者", "件名", "添付ファイル", "フォルダへのリンク"])
                    ws = wb[cat_name]
                    if not self.is_extracting: break
                    
                    # 進捗と時間の計算
                    progress_pct = (i / total_targets) * 100
                    self.progress_var.set(progress_pct)
                    
                    elapsed = time.time() - start_time
                    avg_time = elapsed / i if i > 0 else 2.5 # 初回は2.5秒と仮定
                    remain_sec = int((total_targets - i) * avg_time)
                    total_sec = int(total_targets * avg_time)
                    
                    time_text = f"【全体】進捗: {i}/{total_targets} 件 ({progress_pct:.1f}%) | 残り: 約{remain_sec//60}分{remain_sec%60}秒 (全体: 約{total_sec//60}分{total_sec%60}秒)"
                    self.time_info_var.set(time_text)
                    self.root.update()
                    
                    # DBを再度開く（保存処理用）
                    conn = sqlite3.connect(db_path)
                    c = conn.cursor()
                        
                    self.log(f"新規取得: [No.{item['id']}] {item['title']} ({item['status']})")
                    
                    # フォルダ名のサニタイズ（禁止文字除去）とWindows仕様対策
                    safe_title = re.sub(r'[\\/:*?"<>|]', '_', item['folder_title']).strip()
                    # Windowsでフォルダ末尾のピリオド・スペースが消えてエラーになる現象を回避
                    safe_title = safe_title.rstrip('. ')
                    # パス長制限エラー回避のため長過ぎる場合はカット
                    if len(safe_title) > 80:
                        safe_title = safe_title[:80].rstrip('. ')
                    folder_name = f"[No.{item['id']}]_{safe_title}"
                    
                    # 送信・受信のカテゴリごとに親フォルダを分ける
                    cat_dir = os.path.join(save_dir, cat_name)
                    os.makedirs(cat_dir, exist_ok=True)
                    
                    folder_path = os.path.join(cat_dir, folder_name)
                    os.makedirs(folder_path, exist_ok=True)
                    
                    # 詳細ページへ遷移
                    detail_url = item['link'] if item['link'].startswith('http') else f"{base_url.rsplit('/', 1)[0]}/{item['link']}"
                    self.driver.get(detail_url)
                    time.sleep(1.5)
                    
                    # 詳細情報の抽出
                    detail_js = """
                    var result = { attachments: [], log_text: "" };
                    var fileLinks = document.querySelectorAll("a[href*='Download'], a.file");
                    fileLinks.forEach(a => {
                        result.attachments.push({ name: a.innerText.trim(), url: a.href });
                    });
                    // 必要なテーブル（申請内容と進行状況）だけを綺麗に抽出
                    var logText = "";
                    var tables = document.querySelectorAll("table.vt, table.dataList, table.dataListSummary");
                    if (tables.length > 0) {
                        tables.forEach(function(tbl) {
                            var rows = tbl.querySelectorAll("tr");
                            rows.forEach(function(tr) {
                                var rowText = [];
                                var cells = tr.querySelectorAll("th, td");
                                cells.forEach(function(cell) {
                                    rowText.push(cell.innerText.trim().replace(/\\n+/g, ' '));
                                });
                                logText += rowText.join(" | ") + "\\n";
                            });
                            logText += "\\n";
                        });
                    }
                    
                    if (!logText) {
                        // テーブルがない場合のフォールバック（画面全体から不要なメニューを除去）
                        var main = document.querySelector("#contents") || document.querySelector("#main") || document.querySelector(".vr_module") || document.body;
                        var txt = main.innerText;
                        
                        // 「一覧から削除する」の直後から本文が始まるため、そこから先を切り出す
                        if (txt.includes("一覧から削除する")) {
                            var parts = txt.split("一覧から削除する");
                            txt = parts[parts.length - 1].trim();
                        } else if (txt.includes("申請データの詳細")) {
                            txt = txt.split("申請データの詳細")[1].trim();
                            if (txt.includes("印刷用画面")) {
                                var parts = txt.split("印刷用画面");
                                txt = parts[parts.length - 1].trim();
                            }
                        }
                        logText = txt;
                    }
                    
                    // 末尾の不要なフッター（ワークフロー一覧へ）を除去
                    if (logText.includes("ワークフロー一覧へ")) {
                        logText = logText.split("ワークフロー一覧へ")[0].trim();
                    }
                    
                    result.log_text = logText;
                    return result;
                    """
                    detail_data = self.driver.execute_script(detail_js)
                    
                    # テキストログの保存
                    log_file_path = os.path.join(folder_path, "詳細画面ログ.txt")
                    with open(log_file_path, "w", encoding="utf-8", errors="replace") as lf:
                        lf.write(f"取得日時: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}\n")
                        lf.write(f"番号: {item['id']}\n状態: {item['status']}\n")
                        lf.write("="*40 + "\n")
                        lf.write(detail_data.get("log_text", ""))
                        
                    # 添付ファイルのダウンロード（Cookieを引き継いで直接DL）
                    cookies = self.driver.get_cookies()
                    cookie_str = "; ".join([f"{c['name']}={c['value']}" for c in cookies])
                    
                    attached_names = []
                    for att in detail_data.get("attachments", []):
                        self.log(f"  -> 添付ファイルDL中: {att['name']}")
                        try:
                            req = urllib.request.Request(att['url'], headers={'Cookie': cookie_str})
                            with urllib.request.urlopen(req) as response:
                                safe_att_name = re.sub(r'[\\/:*?"<>|]', '_', att['name']).strip()
                                # ファイル名末尾のピリオド・スペースも念のため除去
                                safe_att_name = safe_att_name.rstrip('. ')
                                att_path = os.path.join(folder_path, safe_att_name)
                                with open(att_path, 'wb') as f:
                                    f.write(response.read())
                                attached_names.append(safe_att_name)
                        except Exception as e:
                            self.log(f"  -> DLエラー: {e}")
                            
                    # DBとCSV台帳の更新
                    now_str = datetime.now().strftime('%Y/%m/%d %H:%M:%S')
                    c.execute("REPLACE INTO processed_workflows (id, category, form_name, title, status, downloaded_at) VALUES (?, ?, ?, ?, ?, ?)",
                              (item['id'], cat_name, item['form'], item['title'], item['status'], now_str))
                    conn.commit()
                    conn.close()
                    
                    att_str = " / ".join(attached_names)
                    row_data = ["", item['id'], item['date'], cat_name, item['status'], item['person'], item['folder_title'], att_str, folder_path]
                    ws.append(row_data)
                    
                    # フォルダへのリンクをハイパーリンク化（クリックで開けるようにする）
                    cell = ws.cell(row=ws.max_row, column=9)
                    cell.hyperlink = folder_path
                    cell.style = "Hyperlink"
                    
                    wb.save(excel_path)
                        
                    self.log(f"  -> 保存完了！")
                    total_new_items += 1
            
            self.log("\n>>> [完了] 一覧画面にある未処理データの抽出がすべて完了しました！")
            self.progress_var.set(100)
            self.time_info_var.set("進捗: すべて完了しました")
            self.root.update()
            
            if total_new_items > 0:
                try:
                    import socket
                    pc_name = socket.gethostname()
                    now_str = datetime.now().strftime('%Y%m%d%H%M')
                    filename = f"{pc_name}_{now_str}.txt"
                    nas_dir = r"\\10.85.33.230\01_全社共有\システム統括部\業改室\★大宮システム部\（NAS）伊藤\サイボウズワークフロー抽出使用状況"
                    os.makedirs(nas_dir, exist_ok=True)
                    
                    # 起動から完了までの実処理時間
                    actual_elapsed_sec = int(time.time() - app_start_time)
                    actual_elapsed_str = f"{actual_elapsed_sec // 60}分 {actual_elapsed_sec % 60}秒"
                    
                    time_saved_sec = total_new_items * 210
                    time_saved_min = time_saved_sec // 60
                    time_saved_hours = time_saved_min // 60
                    time_saved_min_rem = time_saved_min % 60
                    
                    with open(os.path.join(nas_dir, filename), "w", encoding="utf-8") as f:
                        f.write(f"PC名: {pc_name}\n")
                        f.write(f"抽出開始: {start_datetime_str}\n")
                        f.write(f"抽出完了: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}\n")
                        f.write(f"実処理時間: {actual_elapsed_str}\n")
                        f.write(f"抽出件数: {total_new_items}件\n")
                        f.write(f"削減効果: 手作業比較で約 {time_saved_sec}秒 ({time_saved_hours}時間{time_saved_min_rem}分) の業務時間を削減しました。\n")
                    self.log(f"◆ NASへ使用実績を記録しました（実処理時間: {actual_elapsed_str} / 削減効果: {time_saved_hours}時間{time_saved_min_rem}分）")
                except Exception as e:
                    self.log(f"NASへの使用実績記録に失敗: {e}")
        except Exception as e:
            self.log(f"!! 停止: {e}")
        finally:
            self.is_extracting = False
            self.btn_extract.config(state=tk.NORMAL)
            self.btn_stop.config(state=tk.DISABLED)

if __name__ == "__main__":
    root = tk.Tk()
    app = CybozuWorkflowExporterApp(root)
    root.mainloop()
