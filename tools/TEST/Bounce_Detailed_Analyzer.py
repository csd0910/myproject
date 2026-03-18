# =============================================================================
# Bounce_Detailed_Analyzer.py  ―  完全統合版 v4
#
#  【旧機能を完全継承】
#    - Excelカラー色分け（カテゴリ別）
#    - 統計情報シート（理由説明＋推奨対応）
#    - tqdm による進捗バー（件数・%・速度・残時間）
#    - 期間絞り込みフィルター（YYYYMMDD 形式入力対応）
#    - 処理前の総件数カウント＆完了後サマリー表示
#    - 用途判定 / 注文ID等のカスタムID抽出
#
#  【新機能を追加】
#    - EMLフォルダ / 単一EML / mbox を自動判別して読み込み
#    - 二段構えハイブリッド解析
#        第一段階: multipart/report → message/delivery-status (RFC構造解析)
#        第二段階: 全文・全ヘッダー正規表現スキャン（Google/キャリア対応）
#    - 日本語判定結果: 【不在】【満杯】【拒否】【一時的】など
#    - 責任分界点: 相手先サーバー / 自社/中継サーバー
#    - DMARC 個別列（SPF・DKIM は旧版から継承）
#    - スパム指標: X-OCN-SPAM-CHECK 等の独自ヘッダーを集約
#    - Message-ID 列追加
#
#  出力列順（旧版画像の列順に準拠＋新機能列を追加）:
#    受信日時 / バウンス先(宛先) / 用途 / 元の件名 /
#    エラー概要 / ブロック分類 / 日本語判定結果(*) / 責任分界点(*) /
#    注文/ユーザーID / Status / SPF判定 / DKIM/認証結果 / DMARC(*) /
#    スパム判定 / リモートMTA / 詳細診断コード / Message-ID(*) / 本文抜粋(一部)
#    ※ (*) は新規追加列
#
#  出力: Excel (.xlsx)  ―  カラー色分け・統計シート・tqdm進捗バー付き
# =============================================================================

import os
import sys
import re
import glob
import email
import mailbox
import datetime
import traceback
from email import policy
from email.header import decode_header as _decode_header
from email.utils import parsedate_to_datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from tqdm import tqdm


# ─────────────────────────────────────────────────────────────────────────────
# 0.  定数 / 設定
# ─────────────────────────────────────────────────────────────────────────────

OUTPUT_FILENAME = "Detailed_Bounce_Analysis.xlsx"

# 抽出する最大経路数（Received ヘッダーの最大ホップ数）
MAX_ROUTES = 6

# デフォルトmboxパス（空の場合は対話入力）
DEFAULT_MBOX_PATH = r"C:\Users\FMV\AppData\Roaming\Thunderbird\Profiles\pwdo1tlx.default-release\Mail\c300ls1v.mwprem.net\Inbox"

# ─── カテゴリ別カラー定義（旧機能完全継承）─────────────────────────────────
COLORS = {
    "Microsoft Block":         PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
    "Gmail Block":             PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid"),
    "Blacklist":               PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "Mailbox Full":            PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
    "Address Invalid":         PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
    "DNS Error":               PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid"),
    "Auth Failure (SPF/DKIM)": PatternFill(start_color="E8CCFF", end_color="E8CCFF", fill_type="solid"),
    "Timeout/Network":         PatternFill(start_color="CCFFE0", end_color="CCFFE0", fill_type="solid"),
    "Other":                   None,
}

BOLD_FONT   = Font(bold=True)
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F4F8F", end_color="2F4F8F", fill_type="solid")

# ─── セル単位エラー色定義 ──────────────────────────────────────────────────────
_FILL_RED    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 赤：5xx / fail
_FILL_ORANGE = PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")  # オレンジ：警告
_FILL_YELLOW = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # 黄：4xx / softfail
_FILL_GREEN  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 緑：pass

# ─── 経路列を動的生成（MAX_ROUTES ホップ分） ────────────────────────────────
_ROUTE_HEADERS = []
_ROUTE_WIDTHS  = []
for _ri in range(1, MAX_ROUTES + 1):
    _ROUTE_HEADERS += [f"経路{_ri}_ドメイン", f"経路{_ri}_IP"]
    _ROUTE_WIDTHS  += [32, 18]

# ─── 出力列定義（旧版画像の列順に準拠 ＋ 新機能列）────────────────────────
HEADERS = [
    "受信日時",                              # 旧版列1
    "バウンス先(宛先)",                      # 旧版列2
    "用途",                                  # 旧版列3
    "元の件名",                              # 旧版列4
    "エラー概要",                            # 旧版列5
    "ブロック分類",                          # 旧版列6
    "日本語判定結果",                        # ★ 新規
    "責任分界点",                            # ★ 新規
    "注文/ユーザーID",                       # 旧版列7
    "Status",                                # 旧版列8
    "SPF判定",                               # 旧版列9（生の値）
    "SPF判定_説明",                          # ★ 新規（日本語説明）
    "DKIM/認証結果",                         # 旧版列10（生の値）
    "DKIM判定_説明",                         # ★ 新規（日本語説明）
    "DMARC",                                 # ★ 新規（生の値）
    "DMARC判定_説明",                        # ★ 新規（日本語説明）
    "スパム判定",                            # 旧版列11
    "リモートMTA",                           # 旧版列12
    "詳細診断コード (Diagnostic-Code)",      # 旧版列13
    "Message-ID",                            # ★ 新規
] + _ROUTE_HEADERS + [
    "本文抜粋(一部)",                        # 旧版列14
]

# ─── 列幅設定（HEADERS と同順）───────────────────────────────────────────────
COL_WIDTHS = [
    20,  # 受信日時
    30,  # バウンス先(宛先)
    14,  # 用途
    28,  # 元の件名
    35,  # エラー概要
    20,  # ブロック分類
    22,  # 日本語判定結果
    16,  # 責任分界点
    16,  # 注文/ユーザーID
    10,  # Status
    14,  # SPF判定
    30,  # SPF判定_説明
    14,  # DKIM/認証結果
    30,  # DKIM判定_説明
    10,  # DMARC
    30,  # DMARC判定_説明
    30,  # スパム判定
    28,  # リモートMTA
    40,  # 詳細診断コード
    32,  # Message-ID
] + _ROUTE_WIDTHS + [50]  # 経路列 + 本文抜粋

# ─── スパム関連ヘッダー一覧 ──────────────────────────────────────────────────
SPAM_HEADERS = [
    "X-Spam-Status", "X-Spam-Score", "X-Spam-Flag",
    "X-OCN-SPAM-CHECK", "X-Yahoo-Postmaster",
    "X-Gm-Message-State", "X-Spam-Level", "X-Mailer-Spam-Status",
]


# ─────────────────────────────────────────────────────────────────────────────
# 1.  文字コードユーティリティ
# ─────────────────────────────────────────────────────────────────────────────

def safe_decode_bytes(data: bytes, hint_charset: str = None) -> str:
    """bytes → str。ISO-2022-JP / CP932 / UTF-8 混在を完全対応。"""
    if data is None:
        return ""
    candidates = []
    if hint_charset:
        candidates.append(hint_charset)
    candidates += ["utf-8", "iso-2022-jp", "cp932", "shift_jis", "euc-jp", "latin-1"]
    for enc in candidates:
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("utf-8", errors="replace")


def decode_mime_header(raw_value: str) -> str:
    """=?ISO-2022-JP?B?...?= 等のMIMEエンコードをデコードして文字化けを防ぐ。"""
    if not raw_value:
        return ""
    try:
        parts = _decode_header(raw_value)
        result = []
        for fragment, charset in parts:
            if isinstance(fragment, bytes):
                result.append(safe_decode_bytes(fragment, charset))
            else:
                result.append(str(fragment))
        return "".join(result).strip()
    except Exception:
        return str(raw_value)


def clean_for_excel(text) -> str:
    """Excelで扱えない制御文字を除去し、セル上限に合わせる。"""
    if not text:
        return ""
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", str(text))
    return cleaned[:32000]


# ─────────────────────────────────────────────────────────────────────────────
# 2.  エラー分類器（旧機能完全継承＋強化）
# ─────────────────────────────────────────────────────────────────────────────

class ErrorClassifier:
    """カテゴリ分類・説明・推奨対応を一元管理。"""

    EXPLANATIONS = {
        "Microsoft Block": {
            "desc":   "Outlook.com/Hotmailによるブロック。送信ドメインの評判やSPF/DKIM未設定が原因。",
            "action": "送信IPがブラックリストに載っていないかSNDS(Microsoft)で確認し、SPF/DKIMを設定してください。",
        },
        "Gmail Block": {
            "desc":   "Googleのポリシーによるブロック。主にガイドライン未遵守やスパム判定。",
            "action": "Gmail送信者ガイドラインを確認し、SPF/DKIM/DMARCが正しく設定されているか確認してください。",
        },
        "Blacklist": {
            "desc":   "外部ブラックリスト（Spamhaus等）にIPやドメインが登録されています。",
            "action": "MxToolbox等で登録状況を確認し、解除申請を行ってください。",
        },
        "Mailbox Full": {
            "desc":   "受信者のメールボックスが容量不足です。一時的なエラー。",
            "action": "受信者に別手段で連絡し、容量不足を伝えてください。",
        },
        "Address Invalid": {
            "desc":   "宛先メールアドレスが存在しない、または無効です（User Unknown等）。",
            "action": "アドレスの入力ミスを確認し、存在しない場合は送信リストから除外してください。",
        },
        "DNS Error": {
            "desc":   "受信側のドメイン解決に失敗（MXレコードが見つからない等）。",
            "action": "相手ドメインが有効期限切れでないか、サーバーがダウンしていないか確認してください。",
        },
        "Auth Failure (SPF/DKIM)": {
            "desc":   "なりすまし防止（SPF/DKIM）の認証に失敗し、拒否されました。",
            "action": "自社のDNS設定（SPF/DKIMレコード）が正しいか再確認してください。",
        },
        "Timeout/Network": {
            "desc":   "通信タイムアウトやネットワーク接続エラー。",
            "action": "時間を置いて再送するか、自社メールサーバーの負荷状況を確認してください。",
        },
        "Other": {
            "desc":   "その他の理由。詳細は診断コードを確認してください。",
            "action": "診断コードの内容を元に個別に調査してください。",
        },
    }

    PATTERNS = {
        "Microsoft Block":         [r"protection\.outlook\.com", r"550 5\.7\.511", r"550 5\.7\.606",
                                    r"banned sender", r"microsoft"],
        "Gmail Block":             [r"google\.com", r"gsmtp", r"aspmx\.l\.google"],
        "Blacklist":               [r"blocked", r"\bRBL\b", r"Spamhaus", r"SpamRATS", r"SPFBL",
                                    r"spamcop", r"blacklist"],
        "Mailbox Full":            [r"quota", r"mailbox full", r"over.?quota", r"552.?5\.2\.2", r"storage"],
        "Address Invalid":         [r"user unknown", r"no such user", r"does not exist",
                                    r"550.?5\.1\.1", r"mailbox unavailable", r"invalid address"],
        "DNS Error":               [r"dns.?error", r"MX record", r"host not found",
                                    r"550.?5\.1\.2", r"name or service not known"],
        "Auth Failure (SPF/DKIM)": [r"spf.?fail", r"dkim.?fail", r"550.?5\.7\.23",
                                    r"550.?5\.7\.26", r"authentication.?failed"],
        "Timeout/Network":         [r"timed.?out", r"connection.?lost", r"421.?4\.4\.2",
                                    r"connection.?refused"],
    }

    @classmethod
    def classify(cls, content: str) -> str:
        for category, regex_list in cls.PATTERNS.items():
            for pattern in regex_list:
                if re.search(pattern, content, re.IGNORECASE):
                    return category
        return "Other"


# ─────────────────────────────────────────────────────────────────────────────
# 3.  日本語ステータスラベル（新機能）
# ─────────────────────────────────────────────────────────────────────────────

_SMTP_EXACT_CODES = {
    "101": "無効なSSLまたはTLS設定",
    "111": "リモートSMTPサーバーへの接続エラー",
    "221": "ドメインサービスが送信チャネルを閉じています",
    "250": "要求されたメールアクションが完了しOK",
    "420": "受信者のファイアウォール/フィルターがメールをブロック",
    "421": "ドメインサービスが利用不可、送信チャネルを閉じています",
    "422": "受信者のメールボックス容量超過",
    "431": "ストレージ不足またはメモリ不足",
    "432": "ドメインサービスが利用不可、送信チャネルを閉じています",
    "441": "断続的なネットワーク接続、サーバー応答なし",
    "442": "配信開始したが接続が切断",
    "446": "ホップ数が多すぎる（ループの可能性）",
    "447": "タイムアウト発生、サーバー接続確認",
    "449": "DNS問題発生、SMTPコネクタのスマートホスト設定確認",
    "450": "メールボックス利用不可、要求拒否",
    "451": "ローカル処理エラー、再試行",
    "452": "システムストレージ不足",
    "453": "メールなし",
    "454": "TLS一時的に利用不可、認証に暗号化必要",
    "455": "MAIL FROMまたはRCPT TOコマンドのパラメータ解釈不可",
    "458": "ノードへのメッセージキュー不可",
    "459": "ノード利用不可",
    "471": "アンチスパムフィルターまたはファイアウォールで送信阻止",
    "500": "構文エラー、コマンド認識不可",
    "501": "パラメータまたは引数の構文エラー",
    "502": "コマンド未実装",
    "503": "コマンドシーケンス不良",
    "504": "コマンドパラメータ未実装",
    "510": "受信者アドレス確認",
    "511": "メールボックスが見つからない",
    "512": "ドメインが見つからない、ホスト不明",
    "513": "リレー拒否またはアドレス形式不正",
    "515": "宛先メールボックスアドレス無効",
    "517": "送信者メール属性に問題、プロパティ確認",
    "521": "ドメインがメール受信拒否",
    "522": "受信者メールボックス容量超過",
    "523": "サーバー制限超過、メッセージサイズ大",
    "530": "アクセス拒否、認証必要",
    "531": "メールシステム満杯",
    "533": "リモートサーバーのディスク容量不足",
    "534": "認証メカニズム弱すぎる、メッセージサイズ大",
    "535": "同一IPの複数サーバー、認証必要",
    "538": "認証メカニズムに暗号化必要",
    "540": "メールアドレスにDNSサーバーなし",
    "541": "ホスト応答なし",
    "542": "接続不良",
    "543": "ルーティングサーバー障害、経路なし",
    "546": "メールループ",
    "547": "配信タイムアウト",
    "550": "メールボックス利用不可",
    "551": "ローカルユーザーでない、転送パス試行",
    "552": "ストレージ割り当て超過",
    "553": "メールボックス名利用不可",
    "554": "トランザクション失敗",
    "555": "プロトコルバージョン不正",
    "556": "受信サーバーが処理できないほどメッセージ大"
}

def get_japanese_status(smtp_code: str, diag_text: str) -> str:
    """SMTPコードと診断テキストから日本語ラベルを生成（一覧表対応版）。"""
    combined = f"{smtp_code} {diag_text}".lower()

    # ベースの分類ロジック
    if re.search(r"5\.[01]\.[01]|user unknown|no such user|does not exist|mailbox unavailable", combined):
        base_status = "【不在】宛先アドレスが存在しません"
    elif re.search(r"5\.2\.2|over.?quota|mailbox full|storage|quota.?exceed|552", combined):
        base_status = "【満杯】相手の容量不足です"
    elif re.search(r"5\.7\.[0-9]|blocked|spam|reject|policy|dmarc|spf.?fail|dkim.?fail|ban", combined):
        base_status = "【拒否】拒否設定または認証不備による拒否です"
    elif re.search(r"(?<!\d)4\.\d\.\d|421|450|451|452", combined):
        base_status = "【一時的】再送対象エラーです"
    elif re.search(r"(?<!\d)5\.\d\.\d|5[0-9][0-9]", combined):
        base_status = "【恒久的】配信不能エラーです"
    else:
        base_status = "【不明】詳細確認が必要です"

    # 新規：本文や診断ログから3桁のコードを見つけて詳細化
    m_code = re.search(r"\b([1-5]\d{2})\b", combined)
    if m_code:
        code_3d = m_code.group(1)
        if code_3d in _SMTP_EXACT_CODES:
            # 分類が「不明」「配信不能(詳細なし)」「一時的(詳細なし)」などのざっくりしたものなら、具体的な理由をメインにする
            if "詳細確認が必要" in base_status or "配信不能エラー" in base_status or "再送対象エラー" in base_status:
                return f"【{code_3d}エラー】{_SMTP_EXACT_CODES[code_3d]}"
            else:
                # 既に不在や満杯で特定できている場合も、補足を付けることで説得力を向上
                return f"{base_status}（{code_3d}: {_SMTP_EXACT_CODES[code_3d]}）"

    return base_status


# ─────────────────────────────────────────────────────────────────────────────
# 4.  責任分界点の判定（新機能）
# ─────────────────────────────────────────────────────────────────────────────

def get_error_origin(remote_mta: str, diag_text: str, recipient_domain: str) -> str:
    """Remote-MTA や診断テキストから責任分界点を自動判定。"""
    if remote_mta and remote_mta.strip():
        return "相手先サーバー"
    if re.search(r"said:|remote server|mx\.|mail\.|smtp\.", diag_text, re.I):
        return "相手先サーバー"
    if recipient_domain and re.search(re.escape(recipient_domain), diag_text, re.I):
        return "相手先サーバー"
    return "自社/中継サーバー"


# ─────────────────────────────────────────────────────────────────────────────
# 5.  認証ヘッダー解析（SPF/DKIM/DMARC）
# ─────────────────────────────────────────────────────────────────────────────

def parse_auth_results(auth_header: str, received_spf: str = "") -> tuple:
    """
    Authentication-Results から SPF / DKIM / DMARC を個別抽出。
    旧版の「SPF判定」「DKIM/認証結果」列の値も互換性のある形式で返す。
    Returns: (spf_result, dkim_result, dmarc_result, spf_display, dkim_display)
    """
    def _pick(key, text):
        m = re.search(rf"{key}=(\w+)", text, re.I)
        return m.group(1).lower() if m else "-"

    spf_result   = _pick("spf",   auth_header)
    dkim_result  = _pick("dkim",  auth_header)
    dmarc_result = _pick("dmarc", auth_header)

    # Received-SPF からも補完
    if spf_result == "-" and received_spf:
        m = re.search(r"(pass|fail|softfail|neutral|none|permerror|temperror)", received_spf, re.I)
        if m:
            spf_result = m.group(1).lower()

    # 旧版の表示形式（Received-SPF 全文 or 短縮結果）
    spf_display  = received_spf.strip() if received_spf else spf_result
    dkim_display = auth_header.strip()   if auth_header  else dkim_result

    return spf_result, dkim_result, dmarc_result, spf_display, dkim_display


# ─────────────────────────────────────────────────────────────────────────────
# 5b.  認証結果の日本語説明（新機能）
# ─────────────────────────────────────────────────────────────────────────────

def get_auth_explanation(result: str) -> str:
    """SPF / DKIM / DMARC の生の結果値を日本語で説明する。"""
    r = (result or "").lower().strip(" .")
    if r == "pass":
        return "✅ 正常：送信ドメインが認証済み"
    if r == "fail":
        return "❌ 失敗：なりすまし疑い（送信元不一致）"
    if r == "softfail":
        return "⚠️ 警告：送信元が部分的に不一致（要注意）"
    if r == "neutral":
        return "➡️ 中立：明示的な判定なし"
    if r in ("none", "-", ""):
        return "ℹ️ 未設定：レコードなし または 情報なし"
    if r in ("permerror", "temperror"):
        return "⚠️ エラー：認証チェック中にエラー発生"
    return f"ℹ️ その他({result})"


# ─────────────────────────────────────────────────────────────────────────────
# 5c.  経路情報抽出（Received ヘッダー解析）（新機能）
# ─────────────────────────────────────────────────────────────────────────────

def extract_mail_routes(message) -> list:
    """
    Received: ヘッダーを全件取得しメールの通過経路を返す。
    最古の経路（送信元に近い）が先頭になるよう逆順で処理。

    Returns: [{"domain": "...", "ip": "..."}, ...]  最大 MAX_ROUTES 件
    """
    all_received = message.get_all("Received") or []
    routes = []
    for rec in reversed(all_received):          # 古い順（送信元→受信）に並べ替え
        rec_clean = re.sub(r"[\r\n]+", " ", rec)
        # "from hostname" 部分を取得
        m_host = re.search(r"\bfrom\s+([\w.\-]+)", rec_clean, re.I)
        # [x.x.x.x] または (x.x.x.x) 形式の IP を取得
        m_ip   = re.search(r"[\[\(](\d{1,3}(?:\.\d{1,3}){3})[\]\)]", rec_clean)
        domain = m_host.group(1).strip() if m_host else ""
        ip     = m_ip.group(1).strip()   if m_ip   else ""
        if domain or ip:
            routes.append({"domain": domain, "ip": ip})
    return routes[:MAX_ROUTES]


# ─────────────────────────────────────────────────────────────────────────────
# 6.  スパム指標集約
# ─────────────────────────────────────────────────────────────────────────────

def get_spam_summary(message) -> str:
    """複数のスパム関連ヘッダーを集約（旧版の X-Spam-Status 等）。"""
    # まず旧版互換の単純な値を先に試みる
    simple = message.get("X-Spam-Status", "")
    if not simple:
        simple = message.get("X-Spam-Flag", "")

    parts = []
    for h in SPAM_HEADERS:
        val = message.get(h)
        if val:
            val_clean = decode_mime_header(val).strip()
            if val_clean:
                parts.append(f"{h}: {val_clean}")

    # 旧版と同様に主要値を先に、残りを追記
    if parts:
        return " | ".join(parts)
    return simple or "No"


# ─────────────────────────────────────────────────────────────────────────────
# 7.  旧機能の継承：用途・カスタムID
# ─────────────────────────────────────────────────────────────────────────────

def get_purpose(text: str) -> str:
    """メールの用途を判別（旧機能完全継承）。"""
    if any(k in text for k in ["注文", "購入", "Order", "決済", "購入完了"]):
        return "注文・決済報告"
    if any(k in text for k in ["登録", "入会", "Signup", "承認", "会員", "パスワード"]):
        return "会員関連・認証"
    if any(k in text for k in ["マガジン", "ニュース", "Newsletter", "配信", "告知"]):
        return "メルマガ・広報"
    if any(k in text for k in ["予約", "Appointment", "受付"]):
        return "予約管理"
    return "その他一般通知"


def get_custom_id(text: str) -> str:
    """注文番号・ユーザーIDを抽出（旧機能完全継承）。"""
    patterns = [
        r"(?:注文番号|Order ID|管理ID|Transaction ID|決済ID)[:：]\s*([A-Za-z0-9-]+)",
        r"(?:ユーザーID|User ID|会員ID)[:：]\s*([A-Za-z0-9-]+)",
        r"\[ID:(\d+)\]",
    ]
    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            return m.group(1)
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# 8.  本文テキスト取得（共通）
# ─────────────────────────────────────────────────────────────────────────────

def get_full_text(message) -> str:
    """マルチパートを再帰走査し text/plain を結合して返す。"""
    texts = []
    if message.is_multipart():
        for part in message.walk():
            if part.get_content_type() == "text/plain":
                try:
                    raw = part.get_payload(decode=True)
                    if raw:
                        texts.append(safe_decode_bytes(raw, part.get_content_charset()))
                except Exception:
                    pass
    else:
        try:
            raw = message.get_payload(decode=True)
            if raw:
                texts.append(safe_decode_bytes(raw, message.get_content_charset()))
        except Exception:
            pass
    return "\n".join(texts)


# ─────────────────────────────────────────────────────────────────────────────
# 9.  第一段階：RFC構造解析（multipart/report → delivery-status）
# ─────────────────────────────────────────────────────────────────────────────

def _parse_ds_content(content: str) -> dict:
    """delivery-status テキストから各フィールドを抽出。"""
    result = {
        "final_recipient": "", "status": "", "remote_mta": "",
        "diagnostic_code": "", "action": "",
    }
    m = re.search(r"Final-Recipient:\s*rfc822;\s*<?(.+?)>?\s*$", content, re.I | re.M)
    if m:
        result["final_recipient"] = m.group(1).strip()
    m = re.search(r"^Status:\s*(.+)", content, re.I | re.M)
    if m:
        result["status"] = m.group(1).strip()
    m = re.search(r"Remote-MTA:\s*(?:rfc822;\s*)?(.+)", content, re.I)
    if m:
        result["remote_mta"] = m.group(1).strip()
    m = re.search(r"Diagnostic-Code:\s*(?:smtp;\s*)?(.+?)(?=\r?\n[^\s]|\Z)", content, re.I | re.S)
    if m:
        result["diagnostic_code"] = re.sub(r"\s+", " ", m.group(1)).strip()
    m = re.search(r"^Action:\s*(.+)", content, re.I | re.M)
    if m:
        result["action"] = m.group(1).strip()
    return result


def stage1_rfc_parse(message) -> dict:
    """第一段階：multipart/report → message/delivery-status を優先解析。"""
    empty = {
        "final_recipient": "", "status": "", "remote_mta": "",
        "diagnostic_code": "", "action": "", "orig_to_rfc822": "",
    }
    if not message.is_multipart():
        return empty

    result = dict(empty)
    for part in message.walk():
        ctype = part.get_content_type()
        if ctype == "message/delivery-status":
            payload = part.get_payload()
            if isinstance(payload, list):
                content = "\n".join(str(p) for p in payload)
            elif isinstance(payload, bytes):
                content = safe_decode_bytes(payload)
            else:
                content = str(payload) if payload else ""
            result.update(_parse_ds_content(content))
        elif ctype in ("message/rfc822", "text/rfc822-headers"):
            sub = part.get_payload()
            if isinstance(sub, list) and sub:
                orig = sub[0]
                if hasattr(orig, "get"):
                    result["orig_to_rfc822"] = decode_mime_header(orig.get("To", ""))
    return result


# ─────────────────────────────────────────────────────────────────────────────
# 10. 第二段階：正規表現全文スキャン（Google/キャリア等のフォールバック）
# ─────────────────────────────────────────────────────────────────────────────

_TO_PATTERNS = [
    r"Final-Recipient:\s*rfc822;\s*<?([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})>?",
    r"To:\s*<?([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})>?",
    r"([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})[^\n]*への配信に失敗",
    r"<([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})>",
]
_STATUS_PATTERNS = [
    r"Status:\s*(\d\.\d\.\d)",
    r"said:\s*(5\d{2}[^\n]+)",
    r"\b(5\d{2})\s+5\.\d\.\d",
    r"\b(4\d{2})\s+4\.\d\.\d",
]
_DIAG_PATTERNS = [
    r"Diagnostic-Code:\s*(?:smtp;\s*)?(.+?)(?=\r?\n[^\s]|\Z)",
    r"said:\s+(.+?)(?:\s+\(in reply to|\Z)",
    r"Reason:\s*(.+)",
    r"error:\s*(.+)",
]
_ERR_SUMMARY_PATTERNS = [
    r"said:\s*(5\d{2}.+?)(?:\s+\(in reply to|\Z)",
    r"(5\d{2} .+)",
    r"(4\d{2} .+)",
]


def stage2_regex_scan(message, full_text: str) -> dict:
    """第二段階：全ヘッダー＋本文を正規表現でスキャン。"""
    header_text = "\n".join(f"{k}: {v}" for k, v in message.items())
    scan_target = header_text + "\n" + full_text

    result = {
        "final_recipient": "", "status": "", "diagnostic_code": "",
        "remote_mta": "", "error_summary": "",
    }

    for pat in _TO_PATTERNS:
        m = re.search(pat, scan_target, re.I | re.M)
        if m:
            result["final_recipient"] = m.group(1).strip()
            break
    for pat in _STATUS_PATTERNS:
        m = re.search(pat, scan_target, re.I | re.M)
        if m:
            result["status"] = m.group(1).strip()
            break
    for pat in _DIAG_PATTERNS:
        m = re.search(pat, scan_target, re.I | re.S)
        if m:
            raw = m.group(1) if m.lastindex else m.group(0)
            result["diagnostic_code"] = re.sub(r"\s+", " ", raw).strip()[:500]
            break
    for pat in _ERR_SUMMARY_PATTERNS:
        m = re.search(pat, scan_target, re.I | re.M)
        if m:
            result["error_summary"] = m.group(1).strip()[:120]
            break
    m = re.search(r"Remote-MTA:\s*(?:rfc822;\s*)?(.+)", scan_target, re.I)
    if m:
        result["remote_mta"] = m.group(1).strip()

    return result


# ─────────────────────────────────────────────────────────────────────────────
# 11. ハイブリッド解析コア（1通のメールを処理）
# ─────────────────────────────────────────────────────────────────────────────

def analyze_single_message(message) -> dict | None:
    """
    1通のメールをハイブリッド解析して行データ辞書を返す。
    バウンスと判定できなければ None を返す。
    """
    # ── 日付 ──────────────────────────────────────────────────
    date_str = decode_mime_header(message.get("Date", ""))
    parsed_date = ""
    dt_obj = None
    if date_str:
        try:
            dt_obj = parsedate_to_datetime(date_str)
            parsed_date = dt_obj.strftime("%Y-%m-%d %H:%M:%S %Z")
        except Exception:
            parsed_date = date_str

    # ── Message-ID / 件名 ─────────────────────────────────────
    message_id   = decode_mime_header(message.get("Message-ID", "")).strip()
    orig_subject = decode_mime_header(message.get("Subject", ""))

    # ── 全文テキスト ──────────────────────────────────────────
    full_text = get_full_text(message)

    # ── 第一段階（RFC構造解析） ────────────────────────────────
    s1 = stage1_rfc_parse(message)

    # ── 第二段階（正規表現スキャン）───────────────────────────
    s2 = stage2_regex_scan(message, full_text)

    # ── マージ（第一段階優先、不足を第二段階で補完）────────────
    final_recipient = (s1["final_recipient"]
                       or s1.get("orig_to_rfc822", "")
                       or s2["final_recipient"])
    status_code     = s1["status"]          or s2["status"]
    diag_code       = s1["diagnostic_code"] or s2["diagnostic_code"]
    remote_mta      = s1["remote_mta"]      or s2["remote_mta"]

    # エラー概要（旧版の"エラー概要"列に対応する短縮版）
    error_summary = s2["error_summary"]
    if not error_summary and diag_code:
        # Diagnostic-Code の先頭1行をエラー概要として使用
        error_summary = diag_code.split("\n")[0][:120]

    # ── バウンス判定 ──────────────────────────────────────────
    if not final_recipient and not status_code and not diag_code:
        is_bounce = re.search(
            r"Mail Delivery|Undeliverable|bounce|failure notice|Return-Path.*<>|"
            r"配信失敗|不達|エラー",
            orig_subject + full_text[:300], re.I
        )
        if not is_bounce:
            return None

    # ── カテゴリ分類（色分け用）───────────────────────────────
    diag_total = f"{diag_code} {full_text[:500]}"
    category   = ErrorClassifier.classify(diag_total)

    # ── 日本語判定 ────────────────────────────────────────────
    japanese_status = get_japanese_status(status_code, diag_code)

    # ── 責任分界点 ────────────────────────────────────────────
    recipient_domain = final_recipient.split("@", 1)[1] if "@" in final_recipient else ""
    error_origin     = get_error_origin(remote_mta, diag_code, recipient_domain)

    # ── 認証結果 ──────────────────────────────────────────────
    auth_header  = decode_mime_header(message.get("Authentication-Results", ""))
    received_spf = decode_mime_header(message.get("Received-SPF", ""))
    spf_result, dkim_result, dmarc_result, spf_display, dkim_display = \
        parse_auth_results(auth_header, received_spf)

    # ── 認証結果の日本語説明（生の値の直後に説明列）────────────
    spf_expl   = get_auth_explanation(spf_result)
    dkim_expl  = get_auth_explanation(dkim_result)
    dmarc_expl = get_auth_explanation(dmarc_result)

    # ── スパム指標 ────────────────────────────────────────────
    spam_status = get_spam_summary(message)

    # ── 用途・カスタムID（旧機能）────────────────────────────
    purpose   = get_purpose(f"{orig_subject} {full_text}")
    custom_id = get_custom_id(full_text)

    # ── 経路情報（Received ヘッダーから抽出）─────────────────
    routes = extract_mail_routes(message)

    # ── 経路列の辞書を生成 ────────────────────────────────────
    route_dict = {}
    for _i in range(MAX_ROUTES):
        route_dict[f"経路{_i+1}_ドメイン"] = routes[_i]["domain"] if _i < len(routes) else ""
        route_dict[f"経路{_i+1}_IP"]      = routes[_i]["ip"]     if _i < len(routes) else ""

    return {
        # 出力列に対応するキー（HEADERS と同名）
        "受信日時":                         parsed_date,
        "バウンス先(宛先)":                 final_recipient,
        "用途":                             purpose,
        "元の件名":                         orig_subject,
        "エラー概要":                       error_summary,
        "ブロック分類":                     category,
        "日本語判定結果":                   japanese_status,
        "責任分界点":                       error_origin,
        "注文/ユーザーID":                  custom_id,
        "Status":                           status_code,
        "SPF判定":                          spf_display,        # 生の値
        "SPF判定_説明":                     spf_expl,           # 日本語説明
        "DKIM/認証結果":                    dkim_display,       # 生の値
        "DKIM判定_説明":                    dkim_expl,          # 日本語説明
        "DMARC":                            dmarc_result,       # 生の値
        "DMARC判定_説明":                   dmarc_expl,         # 日本語説明
        "スパム判定":                       spam_status,
        "リモートMTA":                      remote_mta,
        "詳細診断コード (Diagnostic-Code)": diag_code[:1000] if diag_code else "",
        "Message-ID":                       message_id,
        **route_dict,                                           # 経路1〜N のドメイン/IP
        "本文抜粋(一部)":                   full_text[:2000],
        # 内部管理用（出力列にはない）
        "_dt_obj":    dt_obj,
        "_category":  category,
        "_spf_raw":   spf_result,       # セル色判定用
        "_dkim_raw":  dkim_result,      # セル色判定用
        "_dmarc_raw": dmarc_result,     # セル色判定用
        "_spf_fail":  "fail" in spf_result or "fail" in dkim_result,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 12. ソースローダー（EMLフォルダ / 単一EML / mbox を自動判別）
# ─────────────────────────────────────────────────────────────────────────────

def _parse_eml_file(filepath: str):
    """単一EMLをパース。文字コード候補を順に試みる。"""
    for enc in ("utf-8", "iso-2022-jp", "cp932", "latin-1"):
        try:
            with open(filepath, "r", encoding=enc, errors="replace") as f:
                return email.message_from_file(f, policy=policy.compat32)
        except Exception:
            continue
    try:
        with open(filepath, "rb") as f:
            return email.message_from_bytes(f.read(), policy=policy.compat32)
    except Exception as e:
        print(f"[Warning] EMLパース失敗: {filepath} -> {e}")
        return None


def _load_messages(input_path: str):
    """入力パスを自動判別してメッセージを yield する。"""
    if os.path.isdir(input_path):
        files = glob.glob(os.path.join(input_path, "**", "*.eml"), recursive=True)
        if not files:
            files = [f for f in glob.glob(os.path.join(input_path, "**", "*"), recursive=True)
                     if os.path.isfile(f)]
        for fp in files:
            msg = _parse_eml_file(fp)
            if msg:
                yield msg
    elif os.path.isfile(input_path):
        ext = os.path.splitext(input_path)[1].lower()
        if ext == ".eml":
            msg = _parse_eml_file(input_path)
            if msg:
                yield msg
        else:
            try:
                mbox = mailbox.mbox(input_path, create=False)
                for msg in mbox:
                    yield msg
            except Exception as e:
                print(f"[Warning] mbox読み込みエラー: {e}")
    else:
        raise FileNotFoundError(f"パスが見つかりません: {input_path}")


def _count_messages(input_path: str) -> int:
    """処理前に総メール件数を数える（tqdm の total 用）。"""
    try:
        if os.path.isdir(input_path):
            files = glob.glob(os.path.join(input_path, "**", "*.eml"), recursive=True)
            if not files:
                files = [f for f in glob.glob(os.path.join(input_path, "**", "*"), recursive=True)
                         if os.path.isfile(f)]
            return len(files)
        elif os.path.isfile(input_path):
            ext = os.path.splitext(input_path)[1].lower()
            if ext == ".eml":
                return 1
            else:
                mbox = mailbox.mbox(input_path, create=False)
                return len(mbox)
    except Exception:
        pass
    return 0


# ─────────────────────────────────────────────────────────────────────────────
# 13. Excel出力（色分け・統計シート付き）
# ─────────────────────────────────────────────────────────────────────────────

def _apply_header_style(ws):
    """ヘッダー行スタイルを適用。"""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _apply_cell_style(ws, row_idx: int, row_data: dict):
    """エラーがある項目のセルにだけ色・太字を適用（セル単位）。"""
    col_map = {h: idx + 1 for idx, h in enumerate(HEADERS)}

    def _set_fill(col_name, fill):
        ci = col_map.get(col_name)
        if ci:
            ws.cell(row=row_idx, column=ci).fill = fill

    def _set_bold(col_name):
        ci = col_map.get(col_name)
        if ci:
            ws.cell(row=row_idx, column=ci).font = BOLD_FONT

    # ── ブロック分類 → カテゴリ色（その列のみ）
    category = row_data.get("_category", "Other")
    cat_fill = COLORS.get(category)
    if cat_fill:
        _set_fill("ブロック分類", cat_fill)

    # ── エラー概要 → 5xx系:赤  4xx系:黄
    err = str(row_data.get("エラー概要", "") or "")
    if re.search(r"\b5\d{2}\b", err):
        _set_fill("エラー概要", _FILL_RED)
        _set_bold("エラー概要")
    elif re.search(r"\b4\d{2}\b", err):
        _set_fill("エラー概要", _FILL_YELLOW)

    # ── 日本語判定 → 【拒否】【不在】【恒久的】時はオレンジ
    jpstatus = str(row_data.get("日本語判定結果", "") or "")
    if any(k in jpstatus for k in ["【拒否】", "【不在】", "【恒久的】"]):
        _set_fill("日本語判定結果", _FILL_ORANGE)
        _set_bold("日本語判定結果")

    # ── SPF
    spf_raw = str(row_data.get("_spf_raw", "") or "").lower()
    if spf_raw == "fail":
        _set_fill("SPF判定",       _FILL_RED)
        _set_fill("SPF判定_説明",  _FILL_RED)
    elif spf_raw == "softfail":
        _set_fill("SPF判定",       _FILL_YELLOW)
        _set_fill("SPF判定_説明",  _FILL_YELLOW)
    elif spf_raw == "pass":
        _set_fill("SPF判定",       _FILL_GREEN)

    # ── DKIM
    dkim_raw = str(row_data.get("_dkim_raw", "") or "").lower()
    if dkim_raw == "fail":
        _set_fill("DKIM/認証結果", _FILL_RED)
        _set_fill("DKIM判定_説明", _FILL_RED)
    elif dkim_raw == "pass":
        _set_fill("DKIM/認証結果", _FILL_GREEN)

    # ── DMARC
    dmarc_raw = str(row_data.get("_dmarc_raw", "") or "").lower()
    if dmarc_raw == "fail":
        _set_fill("DMARC",         _FILL_RED)
        _set_fill("DMARC判定_説明", _FILL_RED)
    elif dmarc_raw == "pass":
        _set_fill("DMARC",         _FILL_GREEN)

    # ── スパム判定 → YES/検出時にオレンジ
    spam = str(row_data.get("スパム判定", "") or "").lower()
    if spam and spam not in ("no", "-") and not spam.startswith("no"):
        _set_fill("スパム判定", _FILL_ORANGE)


def _write_stats_sheet(wb, stats: dict):
    """統計情報シートを書き出す（旧機能完全継承）。"""
    ws = wb.create_sheet("統計情報")
    ws.append(["カテゴリー", "件数", "届かなかった理由の説明", "推奨される対応"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for cat, count in sorted(stats.items(), key=lambda x: -x[1]):
        info = ErrorClassifier.EXPLANATIONS.get(cat, {"desc": "-", "action": "-"})
        ws.append([cat, count, info["desc"], info["action"]])
        fill = COLORS.get(cat)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 35


# ─────────────────────────────────────────────────────────────────────────────
# 14. メインエントリーポイント
# ─────────────────────────────────────────────────────────────────────────────

def analyze_bounce_emails(
    input_path: str,
    output_path: str = None,
    start_date: datetime.datetime = None,
    end_date: datetime.datetime = None,
) -> str:
    """
    バウンスメールを解析してExcel（色分け・統計付き）に出力する。

    Parameters
    ----------
    input_path  : EMLフォルダ / 単一EMLファイル / mboxファイルのパス
    output_path : 出力先xlsxパス（省略時はスクリプトと同フォルダ）
    start_date  : 期間フィルター開始日（timezone-aware datetime）
    end_date    : 期間フィルター終了日（timezone-aware datetime）

    Returns
    -------
    str : 出力されたExcelファイルの絶対パス
    """
    if not output_path:
        base_dir = (
            os.path.dirname(sys.executable)
            if getattr(sys, "frozen", False)
            else os.path.dirname(os.path.abspath(__file__))
        )
        output_path = os.path.join(base_dir, OUTPUT_FILENAME)

    # ── ヘッダー表示 ─────────────────────────────────────────
    print("\n" + "=" * 62)
    print("  バウンスメール解析スクリプト  完全統合版 v3")
    print("=" * 62)
    print(f"  入力 : {input_path}")
    print(f"  出力 : {output_path}")
    if start_date:
        end_label = end_date.strftime("%Y-%m-%d") if end_date else "（上限なし）"
        print(f"  期間 : {start_date.strftime('%Y-%m-%d')} ～ {end_label}")
    print("-" * 62)

    # ── 総件数をカウント（tqdm の total 表示のため） ──────────
    print("  [1/3] メール件数を確認中...", end="", flush=True)
    total_count = _count_messages(input_path)
    print(f" {total_count:,} 件")

    # ── Workbook 初期化 ───────────────────────────────────────
    print("  [2/3] 解析中...", flush=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "詳細分析"
    ws.append(HEADERS)
    _apply_header_style(ws)
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    stats            = {cat: 0 for cat in ErrorClassifier.EXPLANATIONS}
    total            = 0
    analyzed         = 0
    skipped_period   = 0
    skipped_nobounce = 0
    errors           = 0

    # ── tqdm 進捗バー付きメインループ ───────────────────────
    bar_fmt = (
        "  {l_bar}{bar}| "
        "{n_fmt}/{total_fmt} 通 "
        "[経過:{elapsed} 残:{remaining} 速度:{rate_fmt}]"
    )
    with tqdm(
        _load_messages(input_path),
        total=total_count or None,
        unit="通",
        bar_format=bar_fmt,
        ncols=80,
        colour="cyan",
        dynamic_ncols=True,
    ) as pbar:
        for message in pbar:
            total += 1
            try:
                row_data = analyze_single_message(message)

                if row_data is None:
                    skipped_nobounce += 1
                    pbar.set_postfix_str("スキップ（バウンスなし）", refresh=False)
                    continue

                # ── 期間フィルター ────────────────────────────
                dt_obj = row_data.get("_dt_obj")
                if dt_obj:
                    try:
                        if dt_obj.tzinfo is None:
                            dt_obj = dt_obj.replace(tzinfo=datetime.timezone.utc)
                        if start_date and dt_obj < start_date:
                            skipped_period += 1
                            pbar.set_postfix_str("期間外", refresh=False)
                            continue
                        if end_date and dt_obj > end_date:
                            skipped_period += 1
                            pbar.set_postfix_str("期間外", refresh=False)
                            continue
                    except Exception:
                        pass

                # ── 統計カウント ──────────────────────────────
                category = row_data["_category"]
                stats[category] = stats.get(category, 0) + 1
                analyzed += 1

                # ── tqdm リアルタイム表示 ─────────────────────
                addr_disp = (row_data.get("バウンス先(宛先)") or "?")[:28]
                pbar.set_postfix_str(
                    f"✓{analyzed}件 | {addr_disp} [{category}]",
                    refresh=False,
                )

                # ── 行データをExcelに書き込み ─────────────────
                row_values = [clean_for_excel(row_data.get(h, "")) for h in HEADERS]
                ws.append(row_values)

                # ── セル単位スタイル適用（エラー項目のみ色付け） ─
                _apply_cell_style(ws, ws.max_row, row_data)

            except Exception as e:
                errors += 1
                pbar.set_postfix_str(f"⚠ エラー({errors}件)", refresh=False)
                if errors <= 5:
                    tqdm.write(f"  [Warning] 解析エラー: {e}")

    # ── 統計シート書き出し & 保存 ─────────────────────────────
    print(f"\n  [3/3] Excelファイルを保存中...", flush=True)
    _write_stats_sheet(wb, stats)
    try:
        wb.save(output_path)
    except PermissionError:
        alt_path = output_path.replace(
            ".xlsx", f"_{datetime.datetime.now().strftime('%H%M%S')}.xlsx"
        )
        wb.save(alt_path)
        output_path = alt_path
        print(f"  [Info] ファイルが使用中のため別名で保存: {alt_path}")

    # ── 最終サマリー表示 ──────────────────────────────────────
    print("\n" + "=" * 62)
    print("  ▼ 解析完了サマリー")
    print("-" * 62)
    print(f"  処理したメール総数  : {total:>6,} 通")
    print(f"  バウンス検出数      : {analyzed:>6,} 件  ← Excelに出力")
    print(f"  期間外スキップ      : {skipped_period:>6,} 件")
    print(f"  非バウンススキップ  : {skipped_nobounce:>6,} 件")
    print(f"  解析エラー          : {errors:>6,} 件")
    print("-" * 62)
    print("  ▼ カテゴリ別集計")
    for cat, cnt in sorted(stats.items(), key=lambda x: -x[1]):
        if cnt > 0:
            bar_len = min(30, int(cnt / max(analyzed, 1) * 30))
            bar_str = "█" * bar_len + "░" * (30 - bar_len)
            pct     = cnt / analyzed * 100 if analyzed else 0
            print(f"  {cat:<25s} {cnt:>4,} 件 ({pct:5.1f}%) |{bar_str}|")
    print("=" * 62)
    print(f"  出力ファイル: {output_path}")
    print("=" * 62 + "\n")
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# 15. 対話式エントリーポイント（直接実行時）
# ─────────────────────────────────────────────────────────────────────────────

def _input_date(prompt: str) -> datetime.datetime | None:
    """YYYYMMDD または YYYY-MM-DD 形式で日付入力を受け付ける。"""
    while True:
        val = input(prompt).strip()
        if not val:
            return None
        for fmt in ("%Y%m%d", "%Y-%m-%d"):
            try:
                dt = datetime.datetime.strptime(val, fmt)
                return dt.replace(tzinfo=datetime.timezone.utc)
            except ValueError:
                continue
        print("  形式エラー: YYYYMMDD または YYYY-MM-DD で入力してください。")


if __name__ == "__main__":
    try:
        print("=" * 62)
        print("  最強バウンスメール解析スクリプト  完全統合版 v4")
        print("  (EML/mbox 対応 ｜ カラー色分け ｜ 経路抽出 ｜ 進捗表示 ｜ 期間フィルター)")
        print("=" * 62)

        # ── 期間入力 ──────────────────────────────────────────
        print("\n【期間フィルター】（指定しない場合は Enter でスキップ）")
        start_dt = _input_date("  開始日 (例: 20260101 または 2026-01-01) > ")
        end_dt   = _input_date("  終了日 (例: 20260226 または 2026-02-26) > ")
        if end_dt:
            end_dt = end_dt.replace(hour=23, minute=59, second=59)

        # ── 入力パスの取得 ────────────────────────────────────
        if len(sys.argv) >= 2:
            target = sys.argv[1].strip().strip('"')
        else:
            print("\n【入力ソース】EMLフォルダ / EMLファイル / mboxファイルのパスを入力")
            print("  ※ ドラッグ＆ドロップも可")
            target = input("> ").strip().strip('"')

        if not target:
            target = DEFAULT_MBOX_PATH
            print(f"  デフォルトパスを使用: {target}")

        if not os.path.exists(target):
            print(f"\n[Error] パスが見つかりません: {target}")
            print("Press Enter to exit.")
            input()
            sys.exit(1)

        out = sys.argv[2].strip().strip('"') if len(sys.argv) >= 3 else None

        result_path = analyze_bounce_emails(target, out, start_dt, end_dt)
        print(f"✅ 完了！ Excelファイルを開いてください: {result_path}")

    except Exception as e:
        print(f"\n[Critical Error] {e}")
        traceback.print_exc()

    finally:
        print("\n続行するには Enter キーを押してください...")
        try:
            input()
        except EOFError:
            pass
