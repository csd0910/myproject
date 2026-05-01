import openpyxl
wb = openpyxl.load_workbook('data/WorkLog_Unified.xlsx')
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"Sheet: {sheet}")
    rows = list(ws.iter_rows(values_only=True))
    if rows:
        print(f"  Headers: {rows[0]}")
        print(f"  Rows: {len(rows)-1}")
