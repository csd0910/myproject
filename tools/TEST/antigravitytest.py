import openpyxl

def copy_excel_data(source_filepath: str, destination_filepath: str):
    """
    この関数は、指定されたソースExcelファイルからデータを読み込み、
    新しいExcelファイルにその内容を完全にコピーします。
    具体的には、ソースファイルの全シートの全セルデータを宛先ファイルに転送します。

    Args:
        source_filepath (str): コピー元のExcelファイルのパス。
        destination_filepath (str): コピー先のExcelファイルのパス。

    処理フロー:
    1.  **ファイルオープン**: `openpyxl.load_workbook()` を使用して、指定されたソースExcelファイルを開きます。
    2.  **新規ワークブック作成**: `openpyxl.Workbook()` を使用して、データを書き込むための新しいExcelワークブックをメモリ上に作成します。
    3.  **デフォルトシート削除**: 新規作成されたワークブックには 'Sheet' という名前のデフォルトシートが自動的に作成されるため、
        これを削除し、ソースファイルの内容のみを反映できるように準備します。
    4.  **シートのイテレーション**: ソースワークブック内の各シート名を順に処理します。
    5.  **シートの作成とデータコピー**:
        *   ソースワークブックの各シートに対応する新しいシートを、同じ名前で宛先ワークブックに作成します。
        *   ソースシートの全行と全セルをイテレートし、各セルの値 (`cell.value`) を取得します。
        *   取得した値を、宛先シートの対応するセル (`destination_sheet[cell.coordinate]`) に書き込みます。
            `cell.coordinate` を使用することで、セルの位置情報（例: 'A1', 'B2'）を正確に引き継ぎます。
    6.  **ファイル保存**: 全てのデータコピーが完了した後、`destination_wb.save()` を使用して、
        作成された宛先ワークブックを指定されたパスに保存します。
    7.  **エラーハンドリング**:
        *   `FileNotFoundError`: 指定されたソースファイルが見つからない場合に捕捉し、エラーメッセージを出力します。
        *   `Exception`: その他の予期せぬエラーが発生した場合に捕捉し、詳細なエラーメッセージを出力します。

    この関数は、Excelファイルの構造（シート名、セルの位置と値）を保持したまま、
    内容を別のファイルに複製するシナリオで有用です。
    """
    try:
        # ソースワークブックを開く
        source_wb = openpyxl.load_workbook(source_filepath)
        # 宛先ワークブックを新規作成
        destination_wb = openpyxl.Workbook()

        # デフォルトで作成されるシートを削除
        if 'Sheet' in destination_wb.sheetnames:
            destination_wb.remove(destination_wb['Sheet'])

        # ソースワークブックの各シートをイテレート
        for sheet_name in source_wb.sheetnames:
            source_sheet = source_wb[sheet_name]
            # 宛先ワークブックに新しいシートを作成
            destination_sheet = destination_wb.create_sheet(title=sheet_name)

            # ソースシートの行と列をイテレートしてデータをコピー
            for row in source_sheet.iter_rows():
                for cell in row:
                    destination_sheet[cell.coordinate].value = cell.value

        # 宛先ワークブックを保存
        destination_wb.save(destination_filepath)
        print(f"'{source_filepath}' から '{destination_filepath}' へのデータコピーが完了しました。")

    except FileNotFoundError:
        print(f"エラー: ファイルが見つかりません - {source_filepath}")
    except Exception as e:
        print(f"エラーが発生しました: {e}")
