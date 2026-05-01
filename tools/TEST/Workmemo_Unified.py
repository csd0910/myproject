import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import datetime
import os
import sys
import json
import threading

try:
    import pystray
    from PIL import Image, ImageDraw
except ImportError:
    pystray = None

try:
    import keyboard
except ImportError:
    keyboard = None

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

import ctypes
import time

# --- ファイルパス設定 ---
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DATA_DIR = os.path.join(BASE_DIR, "data")
if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

CONFIG_FILE = os.path.join(DATA_DIR, "config_unified.json")

MODES = [
    "他部署受付 - 作業内容管理・進捗",
    "ヘルプデスク - 作業内容管理・進捗",
    "社内機器・ソフト・システムの保守 - 作業内容管理・進捗"
]

DEFAULT_CONFIG = {
    "user_name": "伊藤",
    "hotkey": "ctrl+shift+w",
    "last_mode": "他部署受付 - 作業内容管理・進捗",
    "excel_path": os.path.join(DATA_DIR, "WorkLog_Unified.xlsx"),
    "status_list": ["未完了", "進行中", "完了", "保留", "中止"],
    "helpdesk_categories": ["障害", "運用", "保守", "その他"],
    "helpdesk_genres": ["SCAW関連", "OA端末関連", "基幹端末関連", "ネットワーク", "サーバー", "開発", "総務関連", "業務運用", "名刺", "配送", "その他"],
    "helpdesk_depts": ["商品部", "物流部", "営業管理部", "人事部"],
    "helpdesk_importance": ["高", "中", "低"],
    "sasajima_depts": {
        "代表取締役": ["代表取締役"],
        "監査": ["監査"],
        "営業部": ["営業課", "営業推進課"],
        "カスタマーソリューション部": ["カスタマーサポート課", "サービス課"],
        "商品部": ["商品課", "商品管理課"],
        "マーケティング部": ["プロモーションマーケティング課", "デジタルマーケティング課"],
        "システム部": ["システム課", "システム運営課"],
        "管理部": ["財務経理課", "総務人事課"],
        "事業管理部": ["事業管理部"],
        "物流部": ["ロジスティクス運営課"],
        "配送部": ["配送サービス課"],
        "その他": ["拠点", "他部署"]
    }
}

class ConfigManager:
    @staticmethod
    def load():
        config = DEFAULT_CONFIG.copy()
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf_8_sig") as f:
                    loaded = json.load(f)
                    # 以前の古い組織図データが入っている場合は、最新のデフォルトで上書きする
                    if "sasajima_depts" in loaded:
                        if "代表取締役" not in loaded["sasajima_depts"] or "マーケティング部" not in loaded["sasajima_depts"]:
                            loaded["sasajima_depts"] = DEFAULT_CONFIG["sasajima_depts"]
                    for k, v in loaded.items():
                        config[k] = v
            except:
                pass
        # 起動時に必ず保存して、ユーザーがすぐに編集できるようにする
        ConfigManager.save(config)
        return config

    @staticmethod
    def save(config):
        with open(CONFIG_FILE, "w", encoding="utf_8_sig") as f:
            json.dump(config, f, ensure_ascii=False, indent=4)

class HistoryWindow(tk.Toplevel):
    def __init__(self, parent, excel_path, mode, on_select):
        super().__init__(parent)
        self.title(f"履歴選択 - {mode}")
        self.geometry("700x500")
        self.attributes('-topmost', True)
        self.excel_path = excel_path
        self.mode = mode
        self.on_select = on_select
        self.setup_ui()

    def setup_ui(self):
        # フィルタ
        filter_frame = ttk.Frame(self)
        filter_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(filter_frame, text="表示フィルタ:").pack(side=tk.LEFT)
        self.filter_var = tk.StringVar(value="すべて")
        filter_cb = ttk.Combobox(filter_frame, textvariable=self.filter_var, values=["すべて", "未完了・進行中のみ"], state="readonly", width=18)
        filter_cb.pack(side=tk.LEFT, padx=5)
        filter_cb.bind("<<ComboboxSelected>>", lambda e: self.load_history())

        frame = ttk.Frame(self)
        frame.pack(fill=tk.BOTH, expand=True)

        if self.mode == MODES[0]: # 他部署受付
            columns = ("no", "date", "dept", "subject", "status", "problem")
            headings = ("No", "日付", "依頼部署", "件名", "状況", "内容抜粋")
            widths = (40, 90, 100, 120, 80, 250)
        elif self.mode == MODES[1]: # ヘルプデスク
            columns = ("no", "date", "dept", "client", "status", "problem")
            headings = ("No", "日付", "依頼元", "依頼主", "状況", "内容抜粋")
            widths = (40, 90, 100, 80, 80, 250)
        else: # 保守
            columns = ("no", "subject", "status", "date", "background")
            headings = ("No", "課題件名", "状況", "期間", "背景抜粋")
            widths = (40, 150, 80, 150, 250)

        self.tree = ttk.Treeview(frame, columns=columns, show="headings")
        for col, head, w in zip(columns, headings, widths):
            self.tree.heading(col, text=head)
            self.tree.column(col, width=w)

        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # タグ設定 (色分け)
        self.tree.tag_configure("incomplete", foreground="#d00000", background="#fff0f0")
        self.tree.tag_configure("pending", foreground="#856404", background="#fff3cd")

        self.load_history()
        self.tree.bind("<Double-1>", self.select)
        ttk.Button(self, text="選択して入力欄に反映", command=self.select).pack(pady=10)

    def load_history(self):
        if not os.path.exists(self.excel_path): return
        # 既存リストをクリア
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.history_items = []
        try:
            wb = load_workbook(self.excel_path, data_only=True)
            if self.mode not in wb.sheetnames: return
            ws = wb[self.mode]
            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) <= 1: return
            headers = all_rows[0]
            # 直近100件を逆順で処理
            total_rows = len(all_rows)
            for i in range(total_rows - 1, 0, -1):
                if len(self.history_items) >= 100: break
                r = all_rows[i]
                row_dict = dict(zip(headers, r))
                row_dict["_row_idx"] = i + 1 # 1-indexed Excel row
                
                # フィルタチェック
                status = str(row_dict.get("状況") or row_dict.get("14.ステータス") or "")
                if self.filter_var.get() == "未完了・進行中のみ":
                    if status not in ["未完了", "進行中", "着手中"]:
                        continue

                # 表示対象のインデックスを取得
                current_idx = len(self.history_items)
                self.history_items.append(row_dict)
                
                tag = ""
                if status in ["未完了", "進行中", "着手中"]:
                    tag = "incomplete"
                elif status == "保留":
                    tag = "pending"

                if self.mode == MODES[0]:
                    d_val = str(row_dict.get("依頼部署","")).replace("\n", " / ")
                    dt_val = str(row_dict.get("依頼日",""))[:10]
                    subj_val = str(row_dict.get("課題件名","")).replace("\n", " ")
                    stat_val = str(row_dict.get("状況",""))
                    p_val = str(row_dict.get("課題内容","")).replace("\n", " ")[:50]
                    vals = (row_dict.get("No",""), dt_val, d_val, subj_val, stat_val, p_val)
                elif self.mode == MODES[1]:
                    d_full = f"{row_dict.get('10.依頼部署','')}/{row_dict.get('11.依頼課','')}".strip("/").replace("\n", " ")
                    dt_val = str(row_dict.get("3.日付",""))[:10]
                    stat_val = str(row_dict.get("14.ステータス",""))
                    problem = row_dict.get("15.質問・事象") or row_dict.get("15.質問事象・作業内容") or ""
                    p_val = str(problem).replace("\n", " ")[:50]
                    vals = (row_dict.get("No",""), dt_val, d_full, row_dict.get("12.依頼主",""), stat_val, p_val)
                else:
                    subj_val = str(row_dict.get("課題件名","")).replace("\n", " ")
                    stat_val = str(row_dict.get("状況",""))
                    dt_val = f"{str(row_dict.get('開始日',''))[:10]}～{str(row_dict.get('終了日','') or row_dict.get('期限','') or '')[:10]}"
                    p_val = str(row_dict.get("背景","")).replace("\n", " ")[:50]
                    vals = (row_dict.get("No",""), subj_val, stat_val, dt_val, p_val)
                
                self.tree.insert("", tk.END, values=vals, tags=(str(current_idx), tag))
        except Exception as e:
            print(f"History Load Error: {e}")

    def select(self, event=None):
        selection = self.tree.selection()
        if not selection: return
        idx = int(self.tree.item(selection[0])['tags'][0])
        self.on_select(self.history_items[idx])
        self.destroy()

class SettingsWindow(tk.Toplevel):
    def __init__(self, parent, config, on_save):
        super().__init__(parent)
        self.title("設定")
        self.geometry("400x500")
        self.config_data = config
        self.on_save = on_save
        self.setup_ui()

    def setup_ui(self):
        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="担当者名:", font=("Meiryo", 9, "bold")).pack(anchor=tk.W)
        self.ent_user = ttk.Entry(main_frame)
        self.ent_user.insert(0, self.config_data.get("user_name", ""))
        self.ent_user.pack(fill=tk.X, pady=(0, 15))

        ttk.Label(main_frame, text="呼出ホットキー (例: ctrl+shift+w):", font=("Meiryo", 9, "bold")).pack(anchor=tk.W)
        self.ent_hotkey = ttk.Entry(main_frame)
        self.ent_hotkey.insert(0, self.config_data.get("hotkey", "ctrl+shift+w"))
        self.ent_hotkey.pack(fill=tk.X, pady=(0, 15))

        ttk.Label(main_frame, text="状況リスト (カンマ区切り):", font=("Meiryo", 9, "bold")).pack(anchor=tk.W)
        self.txt_status = tk.Text(main_frame, height=3, font=("Meiryo", 9))
        self.txt_status.insert("1.0", ",".join(self.config_data.get("status_list", [])))
        self.txt_status.pack(fill=tk.X, pady=(0, 15))

        # --- Excel 保存先設定 ---
        e_frame = ttk.LabelFrame(main_frame, text="Excel保存先", padding=5)
        e_frame.pack(fill=tk.X, pady=(0, 15))
        self.ent_excel = ttk.Entry(e_frame)
        self.ent_excel.insert(0, self.config_data.get("excel_path", ""))
        self.ent_excel.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(e_frame, text="参照", width=5, command=self.browse_excel).pack(side=tk.LEFT, padx=5)

        # --- スタートアップ登録設定 ---
        s_frame = ttk.LabelFrame(main_frame, text="自動起動設定", padding=5)
        s_frame.pack(fill=tk.X, pady=(0, 15))
        ttk.Label(s_frame, text="PC起動時に自動で常駐:", font=("Meiryo", 8)).pack(side=tk.LEFT)
        ttk.Button(s_frame, text="登録", width=8, command=self.add_to_startup).pack(side=tk.LEFT, padx=5)
        ttk.Button(s_frame, text="解除", width=8, command=self.remove_from_startup).pack(side=tk.LEFT)

        ttk.Label(main_frame, text="※部署マスタの変更は config_unified.json を編集してください。", foreground="gray", font=("Meiryo", 8)).pack(anchor=tk.W)

        btn_save = ttk.Button(main_frame, text="設定を保存", command=self.save)
        btn_save.pack(pady=10, fill=tk.X)

    def browse_excel(self):
        path = filedialog.asksaveasfilename(
            title="保存先Excelファイルの選択",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=os.path.basename(self.ent_excel.get()) or "WorkLog_Unified.xlsx",
            confirmoverwrite=False
        )
        if path:
            self.ent_excel.delete(0, tk.END)
            self.ent_excel.insert(0, path)

    def add_to_startup(self):
        if not getattr(sys, 'frozen', False):
            messagebox.showinfo("情報", "この機能はexe版でのみ利用可能です。")
            return
        exe_path = sys.executable
        startup_dir = os.path.join(os.environ.get("APPDATA", ""), r"Microsoft\Windows\Start Menu\Programs\Startup")
        shortcut_path = os.path.join(startup_dir, "Workmemo_Unified.lnk")
        vbs_path = os.path.join(os.environ.get("TEMP", ""), "create_shortcut_unified.vbs")
        vbs_content = f'Set oWS = WScript.CreateObject("WScript.Shell")\nsLinkFile = "{shortcut_path}"\nSet oLink = oWS.CreateShortcut(sLinkFile)\noLink.TargetPath = "{exe_path}"\noLink.WindowStyle = 7\noLink.Save'
        try:
            with open(vbs_path, "w", encoding="shift_jis") as f:
                f.write(vbs_content)
            os.system(f'cscript //nologo "{vbs_path}"')
            messagebox.showinfo("完了", "スタートアップに登録しました。")
        except Exception as e:
            messagebox.showerror("エラー", f"失敗: {e}")

    def remove_from_startup(self):
        startup_dir = os.path.join(os.environ.get("APPDATA", ""), r"Microsoft\Windows\Start Menu\Programs\Startup")
        shortcut_path = os.path.join(startup_dir, "Workmemo_Unified.lnk")
        if os.path.exists(shortcut_path):
            try:
                os.remove(shortcut_path)
                messagebox.showinfo("完了", "スタートアップから解除しました。")
            except Exception as e:
                messagebox.showerror("エラー", f"失敗: {e}")
        else:
            messagebox.showinfo("情報", "登録されていません。")

    def save(self):
        new_config = self.config_data.copy()
        new_config["user_name"] = self.ent_user.get()
        new_config["hotkey"] = self.ent_hotkey.get()
        new_config["status_list"] = [s.strip() for s in self.txt_status.get("1.0", tk.END).split(",") if s.strip()]
        new_config["excel_path"] = self.ent_excel.get()
        ConfigManager.save(new_config)
        self.on_save(new_config)
        self.destroy()

class ScrollableFrame(ttk.Frame):
    def __init__(self, container, *args, **kwargs):
        super().__init__(container, *args, **kwargs)
        self.canvas = tk.Canvas(self, bg="#f0f2f5", highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_window, width=event.width)

class UnifiedApp:
    def __init__(self, root):
        self.root = root
        self.root.title("作業記録 Unified")
        self.root.attributes('-topmost', True)
        self.root.geometry("550x650")
        self.root.configure(bg="#f0f2f5")
        self.root.protocol("WM_DELETE_WINDOW", self.hide_window)

        self.config = ConfigManager.load()
        self.tray_icon = None
        self.start_dt = None
        self.last_toggle = 0
        self.current_row_idx = None
        
        self.setup_tray()
        self.setup_hotkey()

        self.apply_style()
        self.setup_ui()
        
        # 起動時にウィンドウを表示する
        self.show_window()

    # --- トレイ・ホットキー ---
    def setup_tray(self):
        if pystray is None: return
        def create_image():
            image = Image.new('RGB', (64, 64), color=(0, 120, 215))
            dc = ImageDraw.Draw(image)
            dc.rectangle((16, 16, 48, 48), fill=(255, 255, 255))
            return image
        menu = pystray.Menu(
            pystray.MenuItem("開く", self.show_window_from_tray),
            pystray.MenuItem("終了", self.quit_app)
        )
        self.tray_icon = pystray.Icon("workmemo", create_image(), "作業記録", menu)
        threading.Thread(target=self.tray_icon.run, daemon=True).start()

    def show_window_from_tray(self, icon=None, item=None):
        self.root.after(0, self.show_window)

    def show_window(self):
        self.root.deiconify()
        self.root.attributes('-topmost', True)
        self.root.lift()
        self.root.focus_force()
        self.root.after(500, lambda: self.root.attributes('-topmost', False))

    def hide_window(self):
        self.root.withdraw()

    def toggle_window(self):
        if self.root.state() == 'withdrawn' or not self.root.winfo_viewable():
            self.show_window()
        else:
            self.hide_window()

    def setup_hotkey(self):
        if keyboard is None: return
        hotkey = self.config.get("hotkey", "ctrl+shift+w").lower().strip()
        try:
            # 以前の登録を解除（エラーは無視）
            try: keyboard.remove_hotkey(hotkey)
            except: pass
            
            # ホットキー登録 (suppress=False で安定性を優先)
            keyboard.add_hotkey(hotkey, self.on_hotkey, suppress=False)
        except Exception as e:
            print(f"Hotkey Setup Error: {e}")

    def on_hotkey(self):
        # 連打対策 (0.5秒以内の連続動作を無視)
        now = time.time()
        if now - self.last_toggle < 0.5:
            return
        self.last_toggle = now
        # メインスレッドで実行
        self.root.after(0, self.toggle_window)

    def quit_app(self, icon=None, item=None):
        if self.tray_icon: self.tray_icon.stop()
        if keyboard is not None:
            try: keyboard.unhook_all()
            except: pass
        self.root.quit()
        self.root.destroy()
        os._exit(0)

    # --- UI 構築 ---
    def apply_style(self):
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure("TFrame", background="#f0f2f5")
        self.style.configure("TLabel", background="#f0f2f5", font=("Segoe UI", 9))
        self.style.configure("Header.TLabel", font=("Segoe UI", 9, "bold"))
        self.style.configure("Accent.TButton", foreground="white", background="#007bff")

    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Header
        header = ttk.Frame(main_frame)
        header.pack(fill=tk.X, pady=(0, 10))
        ttk.Button(header, text="⚙ 設定", command=self.open_settings).pack(side=tk.LEFT)
        self.lbl_user_info = ttk.Label(header, text=f"担当: {self.config.get('user_name', '')}", style="Header.TLabel")
        self.lbl_user_info.pack(side=tk.LEFT, padx=10)
        
        self.lbl_excel_path = ttk.Label(header, text=f"保存先: {os.path.basename(self.config.get('excel_path', ''))}", foreground="blue", cursor="hand2")
        self.lbl_excel_path.pack(side=tk.RIGHT)
        self.lbl_excel_path.bind("<Button-1>", lambda e: self.open_excel())
        
        # Format Selector
        mode_frame = ttk.Frame(main_frame)
        mode_frame.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(mode_frame, text="入力フォーマット:").pack(side=tk.LEFT)
        self.combo_mode = ttk.Combobox(mode_frame, values=MODES, width=40, state="readonly")
        last_mode = self.config.get("last_mode", MODES[0])
        if last_mode not in MODES: last_mode = MODES[0]
        self.combo_mode.set(last_mode)
        self.combo_mode.pack(side=tk.LEFT, padx=5)
        self.combo_mode.bind("<<ComboboxSelected>>", self.on_mode_change)

        # --- スクロール可能なメインエリア ---
        self.scroll_container = ScrollableFrame(main_frame)
        self.scroll_container.pack(fill=tk.BOTH, expand=True)
        self.content_frame = self.scroll_container.scrollable_frame

        # Footer
        footer = ttk.Frame(main_frame)
        footer.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(footer, text="クリア", command=self.on_mode_change).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)
        
        self.btn_update = ttk.Button(footer, text="上書き修正", command=lambda: self.save_log(update=True))
        self.btn_update.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)
        self.btn_update.state(["disabled"])

        ttk.Button(footer, text="新規保存", style="Accent.TButton", command=self.save_log).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)
        
        self.root.bind('<Control-Return>', lambda e: self.save_log())
        
        self.on_mode_change()

    def on_mode_change(self, event=None):
        self.current_row_idx = None
        if hasattr(self, "btn_update"):
            self.btn_update.state(["disabled"])
        
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        mode = self.combo_mode.get()
        
        if self.config.get("last_mode") != mode:
            self.config["last_mode"] = mode
            ConfigManager.save(self.config)

        if mode == MODES[0]: self.build_mode_1()
        elif mode == MODES[1]: self.build_mode_2()
        elif mode == MODES[2]: self.build_mode_3()

    # --- Mode 1: 他部署受付 ---
    def build_mode_1(self):
        f = self.content_frame
        self.inputs = {}
        
        # 部署
        row1 = ttk.Frame(f); row1.pack(fill=tk.X, pady=2)
        ttk.Label(row1, text="依頼部署 (部):").pack(side=tk.LEFT)
        depts = list(self.config.get("sasajima_depts", {}).keys())
        self.cb_dept1 = ttk.Combobox(row1, values=depts, width=15)
        self.cb_dept1.pack(side=tk.LEFT, padx=5)
        self.cb_dept1.bind("<<ComboboxSelected>>", self.update_sections)
        
        ttk.Label(row1, text="課:").pack(side=tk.LEFT)
        self.cb_sect1 = ttk.Combobox(row1, width=15)
        self.cb_sect1.pack(side=tk.LEFT, padx=5)
        self.inputs["dept1"] = self.cb_dept1
        self.inputs["sect1"] = self.cb_sect1
        
        ttk.Button(row1, text="👔履歴", width=7, command=self.open_history).pack(side=tk.RIGHT, padx=2)

        # 課題件名
        row2 = ttk.Frame(f); row2.pack(fill=tk.X, pady=2)
        ttk.Label(row2, text="課題件名:").pack(side=tk.LEFT, fill=tk.X)
        self.inputs["subject"] = ttk.Entry(row2)
        self.inputs["subject"].pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        # 期限 / 状況
        row3 = ttk.Frame(f); row3.pack(fill=tk.X, pady=2)
        ttk.Label(row3, text="期限:").pack(side=tk.LEFT)
        self.inputs["deadline"] = ttk.Entry(row3, width=12)
        self.inputs["deadline"].insert(0, datetime.date.today().strftime("%Y/%m/%d"))
        self.inputs["deadline"].pack(side=tk.LEFT, padx=5)
        
        ttk.Label(row3, text="状況:").pack(side=tk.LEFT)
        self.inputs["status"] = ttk.Combobox(row3, values=self.config.get("status_list", []), width=15)
        self.inputs["status"].pack(side=tk.LEFT, padx=5)

        # テキストエリア
        text_fields = [("課題内容", "problem", 4), ("対応内容", "solution", 4), ("備考", "note", 2), ("進捗", "progress", 2), ("効果", "effect", 2)]
        for label, key, h in text_fields:
            ttk.Label(f, text=f"{label}:", style="Header.TLabel").pack(anchor=tk.W, pady=(5,0))
            
            # スクロールバー付きテキストエリア
            t_frame = ttk.Frame(f)
            t_frame.pack(fill=tk.X)
            txt = tk.Text(t_frame, height=h, font=("Segoe UI", 9), bd=1)
            sb = ttk.Scrollbar(t_frame, orient=tk.VERTICAL, command=txt.yview)
            txt.configure(yscrollcommand=sb.set)
            txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            sb.pack(side=tk.RIGHT, fill=tk.Y)
            
            txt.bind("<Tab>", self.focus_next_widget)
            self.inputs[key] = txt

    def update_sections(self, event=None):
        dept = self.cb_dept1.get()
        sects = self.config.get("sasajima_depts", {}).get(dept, [])
        # 空白を選択肢の先頭に追加（役職者からの依頼対応）
        self.cb_sect1.config(values=[""] + sects)
        self.cb_sect1.set("")

    # --- Mode 2: ヘルプデスク ---
    def build_mode_2(self):
        f = self.content_frame
        self.inputs = {}
        self.start_dt = None
        
        # Time Entry
        time_frame = tk.LabelFrame(f, text="時刻設定", bg="#f0f2f5", padx=5, pady=5)
        time_frame.pack(fill=tk.X, pady=5)
        
        now = datetime.datetime.now()
        ttk.Label(time_frame, text="開始:").pack(side=tk.LEFT)
        self.inputs["start_time"] = ttk.Entry(time_frame, width=8)
        self.inputs["start_time"].insert(0, now.strftime("%H:%M"))
        self.inputs["start_time"].pack(side=tk.LEFT, padx=5)
        
        ttk.Label(time_frame, text="終了:").pack(side=tk.LEFT)
        self.inputs["end_time"] = ttk.Entry(time_frame, width=8)
        self.inputs["end_time"].insert(0, now.strftime("%H:%M"))
        self.inputs["end_time"].pack(side=tk.LEFT, padx=5)
        
        ttk.Button(time_frame, text="👔履歴", width=7, command=self.open_history).pack(side=tk.RIGHT, padx=2)

        # Grid settings
        grid_f = ttk.Frame(f); grid_f.pack(fill=tk.X, pady=5)
        
        # Row 0: 分類 / ジャンル
        ttk.Label(grid_f, text="分類:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.inputs["category"] = ttk.Combobox(grid_f, values=self.config.get("helpdesk_categories", []), width=12)
        self.inputs["category"].grid(row=0, column=1, padx=5, pady=2)
        
        ttk.Label(grid_f, text="ジャンル:").grid(row=0, column=2, sticky=tk.W, pady=2)
        self.inputs["genre"] = ttk.Combobox(grid_f, values=self.config.get("helpdesk_genres", []), width=12)
        self.inputs["genre"].grid(row=0, column=3, padx=5, pady=2)

        # Row 1: 依頼部署 (部 / 課)
        ttk.Label(grid_f, text="依頼部署(部):").grid(row=1, column=0, sticky=tk.W, pady=2)
        depts = list(self.config.get("sasajima_depts", {}).keys())
        self.cb_dept2 = ttk.Combobox(grid_f, values=depts, width=12)
        self.cb_dept2.grid(row=1, column=1, padx=5, pady=2)
        self.cb_dept2.bind("<<ComboboxSelected>>", self.update_sections_m2)
        self.inputs["dept2"] = self.cb_dept2
        
        ttk.Label(grid_f, text="課:").grid(row=1, column=2, sticky=tk.W, pady=2)
        self.cb_sect2 = ttk.Combobox(grid_f, width=12)
        self.cb_sect2.grid(row=1, column=3, padx=5, pady=2)
        self.inputs["sect2"] = self.cb_sect2

        # Row 2: 依頼主 / 重要度
        ttk.Label(grid_f, text="依頼主:").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.inputs["client"] = ttk.Entry(grid_f, width=15)
        self.inputs["client"].grid(row=2, column=1, padx=5, pady=2)
        
        ttk.Label(grid_f, text="重要度:").grid(row=2, column=2, sticky=tk.W, pady=2)
        self.inputs["imp"] = ttk.Combobox(grid_f, values=self.config.get("helpdesk_importance", []), width=12)
        self.inputs["imp"].grid(row=2, column=3, padx=5, pady=2)

        # Row 3: ステータス / 終了予定日
        ttk.Label(grid_f, text="ステータス:").grid(row=3, column=0, sticky=tk.W, pady=2)
        self.inputs["status"] = ttk.Combobox(grid_f, values=self.config.get("status_list", []), width=12)
        self.inputs["status"].grid(row=3, column=1, padx=5, pady=2)

        ttk.Label(grid_f, text="終了予定日:").grid(row=3, column=2, sticky=tk.W, pady=2)
        self.inputs["target_date"] = ttk.Entry(grid_f, width=15)
        self.inputs["target_date"].insert(0, datetime.date.today().strftime("%Y/%m/%d"))
        self.inputs["target_date"].grid(row=3, column=3, padx=5, pady=2)

        # Texts
        text_fields = [("質問・事象", "problem", 6), ("対応内容", "solution", 6)]
        for label, key, h in text_fields:
            ttk.Label(f, text=f"{label}:", style="Header.TLabel").pack(anchor=tk.W, pady=(5,0))
            tf = ttk.Frame(f); tf.pack(fill=tk.X)
            txt = tk.Text(tf, height=h, font=("Segoe UI", 9), bd=1)
            sb = ttk.Scrollbar(tf, orient=tk.VERTICAL, command=txt.yview)
            txt.configure(yscrollcommand=sb.set)
            txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            sb.pack(side=tk.RIGHT, fill=tk.Y)
            txt.bind("<Tab>", self.focus_next_widget)
            self.inputs[key] = txt

    def update_sections_m2(self, event=None):
        dept = self.cb_dept2.get()
        sects = self.config.get("sasajima_depts", {}).get(dept, [])
        # 空白を選択肢の先頭に追加
        self.cb_sect2.config(values=[""] + sects)
        self.cb_sect2.set("")

    # --- Mode 3: 保守 ---
    def build_mode_3(self):
        f = self.content_frame
        self.inputs = {}
        
        row1 = ttk.Frame(f); row1.pack(fill=tk.X, pady=2)
        ttk.Label(row1, text="課題件名:").pack(side=tk.LEFT)
        self.inputs["subject"] = ttk.Entry(row1)
        self.inputs["subject"].pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(row1, text="👔履歴", width=7, command=self.open_history).pack(side=tk.RIGHT, padx=2)

        row2 = ttk.Frame(f); row2.pack(fill=tk.X, pady=2)
        ttk.Label(row2, text="開始日:").pack(side=tk.LEFT)
        self.inputs["start_date"] = ttk.Entry(row2, width=12)
        self.inputs["start_date"].insert(0, datetime.date.today().strftime("%Y/%m/%d"))
        self.inputs["start_date"].pack(side=tk.LEFT, padx=5)
        
        ttk.Label(row2, text="終了日:").pack(side=tk.LEFT)
        self.inputs["end_date"] = ttk.Entry(row2, width=12)
        self.inputs["end_date"].insert(0, datetime.date.today().strftime("%Y/%m/%d"))
        self.inputs["end_date"].pack(side=tk.LEFT, padx=5)
        
        ttk.Label(row2, text="状況:").pack(side=tk.LEFT)
        self.inputs["status"] = ttk.Combobox(row2, values=self.config.get("status_list", []), width=12)
        self.inputs["status"].pack(side=tk.LEFT, padx=5)

        text_fields = [("背景", "background", 4), ("対応内容", "solution", 4), ("備考", "note", 3), ("進捗", "progress", 3)]
        for label, key, h in text_fields:
            ttk.Label(f, text=f"{label}:", style="Header.TLabel").pack(anchor=tk.W, pady=(5,0))
            tf = ttk.Frame(f); tf.pack(fill=tk.X)
            txt = tk.Text(tf, height=h, font=("Segoe UI", 9), bd=1)
            sb = ttk.Scrollbar(tf, orient=tk.VERTICAL, command=txt.yview)
            txt.configure(yscrollcommand=sb.set)
            txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            sb.pack(side=tk.RIGHT, fill=tk.Y)
            txt.bind("<Tab>", self.focus_next_widget)
            self.inputs[key] = txt

    def focus_next_widget(self, event):
        event.widget.tk_focusNext().focus()
        return "break"

    # --- 履歴管理 ---
    def open_history(self):
        mode = self.combo_mode.get()
        path = self.config.get("excel_path", os.path.join(DATA_DIR, "WorkLog_Unified.xlsx"))
        HistoryWindow(self.root, path, mode, self.on_history_select)

    def on_history_select(self, d):
        def g(key): # Helper to prevent None
            return d.get(key) if d.get(key) is not None else ""

        self.current_row_idx = d.get("_row_idx")
        if self.current_row_idx:
            self.btn_update.state(["!disabled"])

        mode = self.combo_mode.get()
        if mode == MODES[0]: # 他部署受付
            dept_full = g("依頼部署")
            if "\n" in str(dept_full):
                d1, d2 = str(dept_full).split("\n", 1)
            else:
                d1, d2 = str(dept_full), ""
            self.inputs["dept1"].set(d1)
            self.update_sections()
            self.inputs["sect1"].set(d2)
            self.inputs["subject"].delete(0, tk.END)
            self.inputs["subject"].insert(0, g("課題件名"))
            self.inputs["status"].set(g("状況"))
            self.inputs["deadline"].delete(0, tk.END)
            self.inputs["deadline"].insert(0, g("期限"))
            
            mapping = {"problem":"課題内容", "solution":"対応内容", "note":"備考", "progress":"進捗", "effect":"効果"}
            for key, col in mapping.items():
                self.inputs[key].delete("1.0", tk.END)
                self.inputs[key].insert("1.0", d.get(col, "") or "")
                
        elif mode == MODES[1]: # ヘルプデスク
            self.inputs["category"].set(g("1.分類"))
            self.inputs["genre"].set(g("9.ジャンル"))
            self.inputs["dept2"].set(g("10.依頼部署"))
            self.update_sections_m2()
            self.inputs["sect2"].set(g("11.依頼課"))
            self.inputs["client"].delete(0, tk.END)
            self.inputs["client"].insert(0, g("12.依頼主"))
            self.inputs["imp"].set(g("13.重要度"))
            self.inputs["status"].set(g("14.ステータス"))
            self.inputs["problem"].delete("1.0", tk.END)
            # 新旧両方のヘッダー名に対応
            problem = d.get("15.質問・事象") or d.get("15.質問事象・作業内容") or ""
            self.inputs["problem"].insert("1.0", g("15.質問・事象") or g("15.質問事象・作業内容"))
            self.inputs["solution"].delete("1.0", tk.END)
            self.inputs["solution"].insert("1.0", g("16.対応内容"))
            self.inputs["start_time"].delete(0, tk.END)
            self.inputs["start_time"].insert(0, g("4.発生時刻"))
            self.inputs["end_time"].delete(0, tk.END)
            self.inputs["end_time"].insert(0, g("5.終了時刻"))
            self.inputs["target_date"].delete(0, tk.END)
            self.inputs["target_date"].insert(0, g("7.終了予定"))
            
        elif mode == MODES[2]: # 保守
            self.inputs["subject"].delete(0, tk.END)
            self.inputs["subject"].insert(0, g("課題件名"))
            self.inputs["status"].set(g("状況"))
            # 旧ヘッダー(期限)と新ヘッダー(終了日)の両方に対応
            self.inputs["start_date"].delete(0, tk.END)
            self.inputs["start_date"].insert(0, g("開始日") or datetime.date.today().strftime("%Y/%m/%d"))
            self.inputs["end_date"].delete(0, tk.END)
            self.inputs["end_date"].insert(0, g("終了日") or g("期限"))
            
            mapping = {"background":"背景", "solution":"対応内容", "note":"備考", "progress":"進捗"}
            for key, col in mapping.items():
                self.inputs[key].delete("1.0", tk.END)
                self.inputs[key].insert("1.0", d.get(col, "") or "")

    # --- 保存処理 ---
    def open_excel(self):
        path = self.config.get("excel_path", os.path.join(DATA_DIR, "WorkLog_Unified.xlsx"))
        if os.path.exists(path):
            os.startfile(path)
        else:
            messagebox.showinfo("情報", "Excelファイルがまだ作成されていません。\n一度保存を行うと作成されます。")

    def change_excel_path(self):
        path = filedialog.asksaveasfilename(
            title="保存先Excelファイルの選択",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile="WorkLog_Unified.xlsx"
        )
        if path:
            self.config["excel_path"] = path
            ConfigManager.save(self.config)
            messagebox.showinfo("完了", f"保存先を変更しました:\n{path}")

    def save_log(self, update=False):
        if Workbook is None:
            messagebox.showerror("エラー", "openpyxlがインストールされていません。")
            return

        mode = self.combo_mode.get()
        path = self.config.get("excel_path", os.path.join(DATA_DIR, "WorkLog_Unified.xlsx"))
        
        try:
            now = datetime.datetime.now()
            if not os.path.exists(path):
                # (既存の新規作成ロジック)
                wb = Workbook()
                wb.remove(wb.active) # デフォルトのSheetを削除
                # 指定された順番ですべてのシートを事前に作成
                for m in MODES:
                    ws_new = wb.create_sheet(title=m)
                    if m == MODES[0]:
                        ws_new.append(["No", "依頼日", "依頼部署", "課題件名", "課題内容", "対応内容", "期限", "状況", "備考", "進捗", "効果", "年度"])
                    elif m == MODES[1]:
                        ws_new.append(["No", "1.分類", "2.担当者", "3.日付", "4.発生時刻", "5.終了時刻", "6.実作業時間", "7.終了予定", "8.完了日付", "9.ジャンル", "10.依頼部署", "11.依頼課", "12.依頼主", "13.重要度", "14.ステータス", "15.質問・事象", "16.対応内容"])
                    elif m == MODES[2]:
                        ws_new.append(["No", "課題件名", "背景", "対応内容", "開始日", "終了日", "状況", "備考", "進捗"])
            else:
                wb = load_workbook(path)
                
            if mode not in wb.sheetnames:
                ws = wb.create_sheet(title=mode)
            else:
                ws = wb[mode]

            next_no = ws.max_row # simple auto-increment (since header is row 1)

            if mode == MODES[0]:
                d1 = self.inputs["dept1"].get()
                d2 = self.inputs["sect1"].get()
                dept_full = f"{d1}\n{d2}" if d2 else d1
                data = [
                    next_no if not update else self.current_row_idx - 1,
                    now.strftime("%Y/%m/%d"),
                    dept_full,
                    self.inputs["subject"].get(),
                    self.inputs["problem"].get("1.0", tk.END).strip(),
                    self.inputs["solution"].get("1.0", tk.END).strip(),
                    self.inputs["deadline"].get(),
                    self.inputs["status"].get(),
                    self.inputs["note"].get("1.0", tk.END).strip(),
                    self.inputs["progress"].get("1.0", tk.END).strip(),
                    self.inputs["effect"].get("1.0", tk.END).strip(),
                    str(now.year) if now.month >= 4 else str(now.year - 1)
                ]
            elif mode == MODES[1]:
                start_t = self.inputs["start_time"].get()
                end_t = self.inputs["end_time"].get()
                
                duration_str = ""
                try:
                    t1 = datetime.datetime.strptime(start_t, "%H:%M")
                    t2 = datetime.datetime.strptime(end_t, "%H:%M")
                    diff = t2 - t1
                    mins = int(diff.total_seconds() / 60)
                    if mins < 0: mins += 1440
                    duration_str = f"{mins}分"
                except:
                    pass

                data = [
                    next_no if not update else self.current_row_idx - 1,
                    self.inputs["category"].get(),
                    self.config.get("user_name", ""),
                    now.strftime("%Y/%m/%d"),
                    start_t,
                    end_t,
                    duration_str,
                    self.inputs["target_date"].get(),
                    now.strftime("%Y/%m/%d"),
                    self.inputs["genre"].get(),
                    self.inputs["dept2"].get(),
                    self.inputs["sect2"].get(),
                    self.inputs["client"].get(),
                    self.inputs["imp"].get(),
                    self.inputs["status"].get(),
                    self.inputs["problem"].get("1.0", tk.END).strip(),
                    self.inputs["solution"].get("1.0", tk.END).strip()
                ]
            elif mode == MODES[2]:
                data = [
                    next_no if not update else self.current_row_idx - 1,
                    self.inputs["subject"].get(),
                    self.inputs["background"].get("1.0", tk.END).strip(),
                    self.inputs["solution"].get("1.0", tk.END).strip(),
                    self.inputs["start_date"].get(),
                    self.inputs["end_date"].get(),
                    self.inputs["status"].get(),
                    self.inputs["note"].get("1.0", tk.END).strip(),
                    self.inputs["progress"].get("1.0", tk.END).strip()
                ]

            if update and self.current_row_idx:
                # 指定行を上書き
                for col_idx, value in enumerate(data, 1):
                    ws.cell(row=self.current_row_idx, column=col_idx, value=value)
                msg = f"行 {self.current_row_idx} を上書き修正しました。"
            else:
                # 末尾に追記
                ws.append(data)
                msg = "新しい履歴として保存しました。"

            wb.save(path)
            messagebox.showinfo("完了", msg)
            
            # 保存後は状態をリセット
            self.on_mode_change()

        except Exception as e:
            messagebox.showerror("保存エラー", f"Excelの書き込みに失敗しました。\nExcelファイルを開いたままではないか確認してください。\n\n詳細: {e}")

    # --- 設定管理 ---
    def open_settings(self):
        SettingsWindow(self.root, self.config, self.update_config)

    def update_config(self, new_config):
        old_hotkey = self.config.get("hotkey")
        self.config = new_config
        self.lbl_user_info.config(text=f"担当: {self.config.get('user_name', '')}")
        self.lbl_excel_path.config(text=f"保存先: {os.path.basename(self.config.get('excel_path', ''))}")
        
        # ホットキー更新
        if new_config.get("hotkey") != old_hotkey:
            self.setup_hotkey() # 再セットアップ
        
        # ステータスコンボの更新
        if "status" in self.inputs:
            self.inputs["status"].config(values=self.config.get("status_list", []))

if __name__ == "__main__":
    # 二重起動防止 (Windows Mutex)
    kernel32 = ctypes.windll.kernel32
    mutex_name = "WorkmemoUnified_SingleInstance_Mutex"
    mutex = kernel32.CreateMutexW(None, False, mutex_name)
    if kernel32.GetLastError() == 183: # ERROR_ALREADY_EXISTS
        # すでに起動している場合は通知して終了
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("お知らせ", "アプリは既に起動しています。\n画面右下のタスクトレイ（アイコン）を確認するか、\nCtrl+Shift+W を押してください。")
        sys.exit(0)

    root = tk.Tk()
    app = UnifiedApp(root)
    root.mainloop()
